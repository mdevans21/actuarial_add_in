#!/usr/bin/env python3
"""
Populate Excel workbook with add-in function examples.
Examples are consistent with the test suite in ActuarialAddIn.Tests.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, ScatterChart, BarChart, Reference, Series
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import Marker
import os
import re

# Paths
EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', 'excel', 'actuarial_add_in.xlsx')

# Common styles
TITLE_FONT = Font(bold=True, size=14)
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
NOTE_FONT = Font(italic=True, size=10, color="666666")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Chart colors
CHART_COLORS = ["4472C4", "ED7D31", "A5A5A5", "FFC000", "5B9BD5", "70AD47"]


def set_column_widths(ws, widths):
    """Set column widths from a dictionary."""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def add_title(ws, title, row=1, col=1):
    """Add a title cell."""
    cell = ws.cell(row=row, column=col, value=title)
    cell.font = TITLE_FONT
    return row + 2


def add_note(ws, note, row, col=1):
    """Add a note/description."""
    cell = ws.cell(row=row, column=col, value=note)
    cell.font = NOTE_FONT
    return row + 1


def add_table_header(ws, headers, row, start_col=1):
    """Add a table header row."""
    for i, header in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + i, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
    return row + 1


def add_data_row(ws, values, row, start_col=1, formulas=None):
    """Add a data row with optional formulas."""
    for i, value in enumerate(values):
        cell = ws.cell(row=row, column=start_col + i)
        if formulas and i < len(formulas) and formulas[i]:
            cell.value = formulas[i]
        else:
            cell.value = value
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
    return row + 1


def create_scatter_chart_with_axes(ws, title, x_col, y_col, min_row, max_row, anchor,
                                   x_title=None, y_title=None, width=10, height=7,
                                   x_min=None, y_min=None, x_max=None, y_max=None):
    """Create a simple line chart with a single series."""
    chart = ScatterChart()
    chart.title = title
    chart.width = width
    chart.height = height
    chart.roundedCorners = False

    # Axis titles and settings
    if x_title:
        chart.x_axis.title = x_title
    if y_title:
        chart.y_axis.title = y_title

    # Axis scaling
    if x_min is not None:
        chart.x_axis.scaling.min = x_min
    if x_max is not None:
        chart.x_axis.scaling.max = x_max
    if y_min is not None:
        chart.y_axis.scaling.min = y_min
    if y_max is not None:
        chart.y_axis.scaling.max = y_max

    # Ensure axes are visible with tick marks and numbers
    chart.x_axis.tickLblPos = "low"
    chart.y_axis.tickLblPos = "low"
    chart.x_axis.delete = False
    chart.y_axis.delete = False

    xvalues = Reference(ws, min_col=x_col, min_row=min_row+1, max_row=max_row)
    yvalues = Reference(ws, min_col=y_col, min_row=min_row+1, max_row=max_row)

    from openpyxl.chart import Series
    series = Series(yvalues, xvalues, title_from_data=False)
    series.marker = Marker(symbol='none')  # No markers, just line
    series.graphicalProperties.line.solidFill = CHART_COLORS[0]
    chart.series.append(series)
    chart.legend = None  # No legend for single series

    ws.add_chart(chart, anchor)
    return chart


def create_bar_chart_with_axes(ws, title, cat_col, data_col, min_row, max_row, anchor,
                                x_title=None, y_title=None, width=12, height=8):
    """Create a bar chart with proper axes."""
    chart = BarChart()
    chart.title = title
    chart.type = "col"
    chart.grouping = "standard"
    chart.width = width
    chart.height = height
    chart.roundedCorners = False
    
    if x_title:
        chart.x_axis.title = x_title
    if y_title:
        chart.y_axis.title = y_title
    
    chart.y_axis.scaling.min = 0
    chart.x_axis.tickLblPos = "low"
    chart.y_axis.tickLblPos = "low"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    
    data = Reference(ws, min_col=data_col, min_row=min_row, max_row=max_row)
    cats = Reference(ws, min_col=cat_col, min_row=min_row+1, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    
    ws.add_chart(chart, anchor)
    return chart


def add_distribution_section(ws, row, title, x_values, pdf_formula, cdf_formula, x_label="x", notes_dict=None):
    """Add a distribution section with table and return data range info."""
    section_start = row
    row = add_note(ws, title, row)
    row = add_table_header(ws, [x_label, "PDF", "CDF", "Notes"], row)
    data_start = row
    for x in x_values:
        formulas = [None, pdf_formula.format(row=row), cdf_formula.format(row=row), None]
        notes = notes_dict.get(x, "") if notes_dict else ""
        row = add_data_row(ws, [x, "", "", notes], row, formulas=formulas)
    data_end = row - 1
    return row, section_start, data_start, data_end


def create_versions_sheet(wb):
    """Create/recreate the Versions sheet with dynamic version info formulas
    and a live, XLL-driven commit history (no populate-time git log)."""

    ws = wb.create_sheet("Versions")
    set_column_widths(ws, {'A': 22, 'B': 55, 'C': 70, 'D': 15})

    row = add_title(ws, "Actuarial Add-In Version History")

    # Version info section - formulas that read from the DLL at runtime
    ws.cell(row=row, column=1, value="Current Version:").font = HEADER_FONT
    ws.cell(row=row, column=2, value='=ACT_VERSION()')
    row += 1
    ws.cell(row=row, column=1, value="Build Date:").font = HEADER_FONT
    ws.cell(row=row, column=2, value='=ACT_BUILD_DATE()')
    row += 1
    ws.cell(row=row, column=1, value="GitHub:").font = HEADER_FONT
    ws.cell(row=row, column=2, value='=ACT_GITHUB_URL()')
    row += 2

    # Version functions demo
    ws.cell(row=row, column=1, value="Version Functions:").font = HEADER_FONT
    row += 1
    for func, formula in [
        ("ACT_VERSION()", "=ACT_VERSION()"),
        ("ACT_BUILD_DATE()", "=ACT_BUILD_DATE()"),
        ("ACT_GITHUB_URL()", "=ACT_GITHUB_URL()"),
        ("ACT_COMMIT_COUNT()", "=ACT_COMMIT_COUNT()"),
        ('ACT_COMMIT_INFO(n, field)', '=ACT_COMMIT_INFO(1, "message")'),
    ]:
        ws.cell(row=row, column=1, value=func)
        ws.cell(row=row, column=2, value=formula)
        row += 1
    row += 1

    # Commit history — live, driven by the XLL's curated release log in
    # VersionInfo.GetCommitHistory(). One anchor cell, dynamic-array spill
    # (Formula2-promoted by dump_workbook.ps1 -SaveWorkbook). The previous
    # `git log -20` snapshot drifted whenever populate_examples.py wasn't
    # re-run; this version refreshes itself on every XLL rebuild.
    ws.cell(row=row, column=1, value="Commit History (live from XLL)").font = HEADER_FONT
    row += 1
    anchor = ws.cell(row=row, column=1, value='=ACT_COMMIT_HISTORY()')
    anchor.border = THIN_BORDER
    return ws


def create_distributions_sheet(wb):
    """Create the Distributions examples sheet with ALL distributions."""
    ws = wb.create_sheet("Distributions")
    set_column_widths(ws, {'A': 12, 'B': 15, 'C': 15, 'D': 20, 'E': 5, 'F': 15})

    # Chart settings - all charts same size
    CHART_WIDTH = 8
    CHART_HEIGHT = 6
    CHART_COL = "F"  # Column for charts
    SECTION_PADDING = 15  # Minimum rows per section (for chart height)

    row = add_title(ws, "Statistical Distributions")
    row = add_note(ws, "PDF = Probability Density Function, CDF = Cumulative Distribution Function", row)
    row += 1

    distributions = []

    # 1. Normal Distribution
    section_start = row
    row = add_note(ws, "NORMAL DISTRIBUTION (mu=0, sigma=1)", row)
    row = add_table_header(ws, ["x", "PDF", "CDF", "Notes"], row)
    data_start = row
    for x in [-3, -2, -1, -0.5, 0, 0.5, 1, 2, 3]:
        formulas = [None, f'=ACT_DIST_NORMAL_PDF(A{row}, 0, 1)', f'=ACT_DIST_NORMAL_CDF(A{row}, 0, 1)', None]
        notes = "Mean" if x == 0 else ""
        row = add_data_row(ws, [x, "", "", notes], row, formulas=formulas)
    distributions.append(("Normal PDF (μ=0, σ=1)", section_start, data_start, row - 1, "x", "f(x)", None, None))
    row = max(row, section_start + SECTION_PADDING)

    # 2. Poisson Distribution
    section_start = row
    row = add_note(ws, "POISSON DISTRIBUTION (lambda=5)", row)
    row = add_table_header(ws, ["k", "PDF", "CDF", "Notes"], row)
    data_start = row
    for k in range(11):
        formulas = [None, f'=ACT_DIST_POISSON_PDF(A{row}, 5)', f'=ACT_DIST_POISSON_CDF(A{row}, 5)', None]
        notes = "Mode" if k == 5 else ""
        row = add_data_row(ws, [k, "", "", notes], row, formulas=formulas)
    distributions.append(("Poisson PMF (λ=5)", section_start, data_start, row - 1, "k", "P(X=k)", 0, None))
    row = max(row, section_start + SECTION_PADDING)

    # 3. Negative Binomial
    section_start = row
    row = add_note(ws, "NEGATIVE BINOMIAL DISTRIBUTION (r=5, p=0.3)", row)
    row = add_table_header(ws, ["k", "PDF", "CDF", "Notes"], row)
    data_start = row
    for k in range(12):
        formulas = [None, f'=ACT_DIST_NEGBIN_PDF(A{row}, 5, 0.3)', f'=ACT_DIST_NEGBIN_CDF(A{row}, 5, 0.3)', None]
        row = add_data_row(ws, [k, "", "", ""], row, formulas=formulas)
    distributions.append(("Negative Binomial PMF", section_start, data_start, row - 1, "k", "P(X=k)", 0, None))
    row = max(row, section_start + SECTION_PADDING)

    # 4. Lognormal
    section_start = row
    row = add_note(ws, "LOGNORMAL DISTRIBUTION (mu=0, sigma=1)", row)
    row = add_table_header(ws, ["x", "PDF", "CDF", "Notes"], row)
    data_start = row
    for x in [0.1, 0.25, 0.5, 0.75, 1.0, 1.5, 2.0, 3.0, 4.0, 5.0]:
        formulas = [None, f'=ACT_DIST_LOGNORM_PDF(A{row}, 0, 1)', f'=ACT_DIST_LOGNORM_CDF(A{row}, 0, 1)', None]
        notes = "Median" if x == 1.0 else ""
        row = add_data_row(ws, [x, "", "", notes], row, formulas=formulas)
    distributions.append(("Lognormal PDF (μ=0, σ=1)", section_start, data_start, row - 1, "x", "f(x)", 0, None))
    row = max(row, section_start + SECTION_PADDING)

    # 5. Gamma
    section_start = row
    row = add_note(ws, "GAMMA DISTRIBUTION (alpha=2, beta=1)", row)
    row = add_table_header(ws, ["x", "PDF", "CDF", "Notes"], row)
    data_start = row
    for x in [0.0, 0.5, 1.0, 1.5, 2.0, 3.0, 4.0, 5.0, 6.0]:
        formulas = [None, f'=ACT_DIST_GAMMA_PDF(A{row}, 2, 1)', f'=ACT_DIST_GAMMA_CDF(A{row}, 2, 1)', None]
        notes = "Mean" if x == 2.0 else ""
        row = add_data_row(ws, [x, "", "", notes], row, formulas=formulas)
    distributions.append(("Gamma PDF (α=2, β=1)", section_start, data_start, row - 1, "x", "f(x)", 0, None))
    row = max(row, section_start + SECTION_PADDING)

    # 6. Exponential
    section_start = row
    row = add_note(ws, "EXPONENTIAL DISTRIBUTION (lambda=1)", row)
    row = add_table_header(ws, ["x", "PDF", "CDF", "Notes"], row)
    data_start = row
    for x in [0.0, 0.5, 1.0, 1.5, 2.0, 2.5, 3.0, 4.0, 5.0]:
        formulas = [None, f'=ACT_DIST_EXP_PDF(A{row}, 1)', f'=ACT_DIST_EXP_CDF(A{row}, 1)', None]
        notes = "Mean" if x == 1.0 else ""
        row = add_data_row(ws, [x, "", "", notes], row, formulas=formulas)
    distributions.append(("Exponential PDF (λ=1)", section_start, data_start, row - 1, "x", "f(x)", 0, None))
    row = max(row, section_start + SECTION_PADDING)

    # 7. Weibull
    section_start = row
    row = add_note(ws, "WEIBULL DISTRIBUTION (k=2, lambda=1)", row)
    row = add_table_header(ws, ["x", "PDF", "CDF", "Notes"], row)
    data_start = row
    for x in [0.0, 0.25, 0.5, 0.75, 1.0, 1.5, 2.0, 2.5, 3.0]:
        formulas = [None, f'=ACT_DIST_WEIBULL_PDF(A{row}, 2, 1)', f'=ACT_DIST_WEIBULL_CDF(A{row}, 2, 1)', None]
        row = add_data_row(ws, [x, "", "", ""], row, formulas=formulas)
    distributions.append(("Weibull PDF (k=2, λ=1)", section_start, data_start, row - 1, "x", "f(x)", 0, None))
    row = max(row, section_start + SECTION_PADDING)

    # 8. Beta
    section_start = row
    row = add_note(ws, "BETA DISTRIBUTION (alpha=2, beta=5)", row)
    row = add_table_header(ws, ["x", "PDF", "CDF", "Notes"], row)
    data_start = row
    for x in [0.0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0]:
        formulas = [None, f'=ACT_DIST_BETA_PDF(A{row}, 2, 5)', f'=ACT_DIST_BETA_CDF(A{row}, 2, 5)', None]
        row = add_data_row(ws, [x, "", "", ""], row, formulas=formulas)
    distributions.append(("Beta PDF (α=2, β=5)", section_start, data_start, row - 1, "x", "f(x)", 0, 1))
    row = max(row, section_start + SECTION_PADDING)

    # 9. Pareto Type I
    section_start = row
    row = add_note(ws, "PARETO TYPE I DISTRIBUTION (alpha=2, xm=1)", row)
    row = add_table_header(ws, ["x", "PDF", "CDF", "Notes"], row)
    data_start = row
    for x in [1.0, 1.5, 2.0, 3.0, 4.0, 5.0, 7.0, 10.0]:
        formulas = [None, f'=ACT_DIST_PARETO_PDF(A{row}, 2, 1)', f'=ACT_DIST_PARETO_CDF(A{row}, 2, 1)', None]
        notes = "Minimum (xm)" if x == 1.0 else ""
        row = add_data_row(ws, [x, "", "", notes], row, formulas=formulas)
    distributions.append(("Pareto I PDF (α=2, xm=1)", section_start, data_start, row - 1, "x", "f(x)", 0, None))
    row = max(row, section_start + SECTION_PADDING)

    # 10. Lomax (Pareto Type II)
    section_start = row
    row = add_note(ws, "LOMAX (PARETO TYPE II) DISTRIBUTION (alpha=2, lambda=1)", row)
    row = add_table_header(ws, ["x", "PDF", "CDF", "Notes"], row)
    data_start = row
    for x in [0.0, 0.5, 1.0, 2.0, 3.0, 4.0, 5.0, 7.0, 10.0]:
        formulas = [None, f'=ACT_DIST_LOMAX_PDF(A{row}, 2, 1)', f'=ACT_DIST_LOMAX_CDF(A{row}, 2, 1)', None]
        row = add_data_row(ws, [x, "", "", ""], row, formulas=formulas)
    distributions.append(("Lomax PDF (α=2, λ=1)", section_start, data_start, row - 1, "x", "f(x)", 0, None))
    row = max(row, section_start + SECTION_PADDING)

    # 11. GPD (Generalized Pareto)
    section_start = row
    row = add_note(ws, "GENERALIZED PARETO (GPD) DISTRIBUTION (xi=0.5, sigma=1)", row)
    row = add_table_header(ws, ["x", "PDF", "CDF", "Notes"], row)
    data_start = row
    for x in [0.0, 0.25, 0.5, 1.0, 1.5, 2.0, 3.0, 4.0, 5.0]:
        formulas = [None, f'=ACT_DIST_GPD_PDF(A{row}, 0.5, 1)', f'=ACT_DIST_GPD_CDF(A{row}, 0.5, 1)', None]
        row = add_data_row(ws, [x, "", "", ""], row, formulas=formulas)
    distributions.append(("GPD PDF (ξ=0.5, σ=1)", section_start, data_start, row - 1, "x", "f(x)", 0, None))
    row = max(row, section_start + SECTION_PADDING)

    # 12. Burr Type XII
    section_start = row
    row = add_note(ws, "BURR TYPE XII DISTRIBUTION (c=2, k=1, lambda=1)", row)
    row = add_table_header(ws, ["x", "PDF", "CDF", "Notes"], row)
    data_start = row
    for x in [0.0, 0.25, 0.5, 1.0, 1.5, 2.0, 3.0, 4.0, 5.0]:
        formulas = [None, f'=ACT_DIST_BURR_PDF(A{row}, 2, 1, 1)', f'=ACT_DIST_BURR_CDF(A{row}, 2, 1, 1)', None]
        row = add_data_row(ws, [x, "", "", ""], row, formulas=formulas)
    distributions.append(("Burr XII PDF (c=2, k=1, λ=1)", section_start, data_start, row - 1, "x", "f(x)", 0, None))
    row = max(row, section_start + SECTION_PADDING)

    # 13. Inverse Gaussian (Wald)
    section_start = row
    row = add_note(ws, "INVERSE GAUSSIAN (WALD) DISTRIBUTION (mu=2, lambda=3)", row)
    row = add_table_header(ws, ["x", "PDF", "CDF", "Notes"], row)
    data_start = row
    for x in [0.25, 0.5, 1.0, 1.5, 2.0, 2.5, 3.0, 4.0, 5.0]:
        formulas = [None, f'=ACT_DIST_INVGAUSS_PDF(A{row}, 2, 3)', f'=ACT_DIST_INVGAUSS_CDF(A{row}, 2, 3)', None]
        notes = "Mean" if x == 2.0 else ""
        row = add_data_row(ws, [x, "", "", notes], row, formulas=formulas)
    distributions.append(("Inverse Gaussian PDF (μ=2, λ=3)", section_start, data_start, row - 1, "x", "f(x)", 0, None))
    row = max(row, section_start + SECTION_PADDING)

    # 14. Loglogistic
    section_start = row
    row = add_note(ws, "LOGLOGISTIC DISTRIBUTION (alpha=2, beta=3)", row)
    row = add_table_header(ws, ["x", "PDF", "CDF", "Notes"], row)
    data_start = row
    for x in [0.25, 0.5, 1.0, 1.5, 2.0, 3.0, 4.0, 5.0, 7.0]:
        formulas = [None, f'=ACT_DIST_LOGLOGISTIC_PDF(A{row}, 2, 3)', f'=ACT_DIST_LOGLOGISTIC_CDF(A{row}, 2, 3)', None]
        notes = "Median" if x == 2.0 else ""
        row = add_data_row(ws, [x, "", "", notes], row, formulas=formulas)
    distributions.append(("Loglogistic PDF (α=2, β=3)", section_start, data_start, row - 1, "x", "f(x)", 0, None))
    row = max(row, section_start + SECTION_PADDING)

    # 15. Zero-Truncated Poisson
    section_start = row
    row = add_note(ws, "ZERO-TRUNCATED POISSON (lambda=5)", row)
    row = add_table_header(ws, ["k", "PMF", "CDF", "Notes"], row)
    data_start = row
    for k in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]:
        formulas = [None, f'=ACT_DIST_ZTPOISSON_PDF(A{row}, 5)', f'=ACT_DIST_ZTPOISSON_CDF(A{row}, 5)', None]
        notes = "Mode" if k == 5 else ""
        row = add_data_row(ws, [k, "", "", notes], row, formulas=formulas)
    row = add_data_row(ws, ["Mean:", f'=ACT_DIST_ZTPOISSON_MEAN(5)', "", "E[X|X>0]"], row)
    distributions.append(("ZT Poisson PMF (λ=5)", section_start, data_start, row - 2, "k", "P(X=k)", 1, 10))
    row = max(row, section_start + SECTION_PADDING)

    # 16. Zero-Truncated Negative Binomial
    section_start = row
    row = add_note(ws, "ZERO-TRUNCATED NEGATIVE BINOMIAL (r=5, p=0.3)", row)
    row = add_table_header(ws, ["k", "PMF", "CDF", "Notes"], row)
    data_start = row
    for k in [1, 2, 3, 5, 7, 10, 15, 20, 25, 30]:
        formulas = [None, f'=ACT_DIST_ZTNEGBIN_PDF(A{row}, 5, 0.3)', f'=ACT_DIST_ZTNEGBIN_CDF(A{row}, 5, 0.3)', None]
        row = add_data_row(ws, [k, "", "", ""], row, formulas=formulas)
    row = add_data_row(ws, ["Mean:", f'=ACT_DIST_ZTNEGBIN_MEAN(5, 0.3)', "", "E[X|X>0]"], row)
    distributions.append(("ZT NegBin PMF (r=5, p=0.3)", section_start, data_start, row - 2, "k", "P(X=k)", 1, 30))
    row = max(row, section_start + SECTION_PADDING)

    # 17. Zero-Truncated Binomial
    section_start = row
    row = add_note(ws, "ZERO-TRUNCATED BINOMIAL (n=10, p=0.3)", row)
    row = add_table_header(ws, ["k", "PMF", "CDF", "Notes"], row)
    data_start = row
    for k in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]:
        formulas = [None, f'=ACT_DIST_ZTBINOM_PDF(A{row}, 10, 0.3)', f'=ACT_DIST_ZTBINOM_CDF(A{row}, 10, 0.3)', None]
        notes = "Max" if k == 10 else ""
        row = add_data_row(ws, [k, "", "", notes], row, formulas=formulas)
    row = add_data_row(ws, ["Mean:", f'=ACT_DIST_ZTBINOM_MEAN(10, 0.3)', "", "E[X|X>0]"], row)
    distributions.append(("ZT Binomial PMF (n=10, p=0.3)", section_start, data_start, row - 2, "k", "P(X=k)", 1, 10))
    row = max(row, section_start + SECTION_PADDING)

    # 18. Zero-Truncated Geometric
    section_start = row
    row = add_note(ws, "ZERO-TRUNCATED GEOMETRIC (p=0.3)", row)
    row = add_table_header(ws, ["k", "PMF", "CDF", "Notes"], row)
    data_start = row
    for k in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]:
        formulas = [None, f'=ACT_DIST_ZTGEOM_PDF(A{row}, 0.3)', f'=ACT_DIST_ZTGEOM_CDF(A{row}, 0.3)', None]
        notes = "Mode" if k == 1 else ""
        row = add_data_row(ws, [k, "", "", notes], row, formulas=formulas)
    row = add_data_row(ws, ["Mean:", f'=ACT_DIST_ZTGEOM_MEAN(0.3)', "", "E[X|X>0] = 1/p"], row)
    distributions.append(("ZT Geometric PMF (p=0.3)", section_start, data_start, row - 2, "k", "P(X=k)", 1, 10))
    row = max(row, section_start + SECTION_PADDING)

    # 19. Zero-Modified Poisson
    section_start = row
    row = add_note(ws, "ZERO-MODIFIED POISSON (lambda=5, p0=0.2)", row)
    row = add_table_header(ws, ["k", "PMF", "CDF", "Notes"], row)
    data_start = row
    for k in [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10]:
        formulas = [None, f'=ACT_DIST_ZMPOISSON_PDF(A{row}, 5, 0.2)', f'=ACT_DIST_ZMPOISSON_CDF(A{row}, 5, 0.2)', None]
        notes = "Inflated" if k == 0 else ""
        row = add_data_row(ws, [k, "", "", notes], row, formulas=formulas)
    row = add_data_row(ws, ["Mean:", f'=ACT_DIST_ZMPOISSON_MEAN(5, 0.2)', "", "(1-p0)*λ"], row)
    row = add_data_row(ws, ["Var:", f'=ACT_DIST_ZMPOISSON_VAR(5, 0.2)', "", "(1-p0)*λ*(1+p0*λ)"], row)
    distributions.append(("ZM Poisson PMF (λ=5, p0=0.2)", section_start, data_start, row - 3, "k", "P(X=k)", 0, 10))
    row = max(row, section_start + SECTION_PADDING)

    # 20. Zero-Modified Negative Binomial
    section_start = row
    row = add_note(ws, "ZERO-MODIFIED NEGATIVE BINOMIAL (r=5, p=0.3, p0=0.2)", row)
    row = add_table_header(ws, ["k", "PMF", "CDF", "Notes"], row)
    data_start = row
    for k in [0, 1, 2, 3, 5, 7, 10, 15, 20, 25]:
        formulas = [None, f'=ACT_DIST_ZMNEGBIN_PDF(A{row}, 5, 0.3, 0.2)', f'=ACT_DIST_ZMNEGBIN_CDF(A{row}, 5, 0.3, 0.2)', None]
        notes = "Inflated" if k == 0 else ""
        row = add_data_row(ws, [k, "", "", notes], row, formulas=formulas)
    row = add_data_row(ws, ["Mean:", f'=ACT_DIST_ZMNEGBIN_MEAN(5, 0.3, 0.2)', "", "(1-p0)*r*(1-p)/p"], row)
    row = add_data_row(ws, ["Var:", f'=ACT_DIST_ZMNEGBIN_VAR(5, 0.3, 0.2)', "", ""], row)
    distributions.append(("ZM NegBin PMF (r=5, p=0.3, p0=0.2)", section_start, data_start, row - 3, "k", "P(X=k)", 0, 25))
    row = max(row, section_start + SECTION_PADDING)

    # 21. Pareto III (Lomax with location)
    section_start = row
    row = add_note(ws, "PARETO III (mu=1, sigma=2, gamma=3)", row)
    row = add_table_header(ws, ["x", "PDF", "CDF", "Notes"], row)
    data_start = row
    for x in [1.0, 1.5, 2.0, 2.5, 3.0, 4.0, 5.0, 7.0, 10.0]:
        formulas = [None, f'=ACT_DIST_PARETO3_PDF(A{row}, 1, 2, 3)', f'=ACT_DIST_PARETO3_CDF(A{row}, 1, 2, 3)', None]
        notes = "Location" if x == 1.0 else ""
        row = add_data_row(ws, [x, "", "", notes], row, formulas=formulas)
    row = add_data_row(ws, ["Mean:", f'=ACT_DIST_PARETO3_MEAN(1, 2, 3)', "", "μ + σ/(γ-1)"], row)
    distributions.append(("Pareto III PDF (μ=1, σ=2, γ=3)", section_start, data_start, row - 2, "x", "f(x)", 1, 10))
    row = max(row, section_start + SECTION_PADDING)

    # 22. Pareto IV (Generalized)
    section_start = row
    row = add_note(ws, "PARETO IV (mu=0, sigma=2, gamma=0.5, alpha=3)", row)
    row = add_table_header(ws, ["x", "PDF", "CDF", "Notes"], row)
    data_start = row
    for x in [0.1, 0.5, 1.0, 1.5, 2.0, 3.0, 4.0, 5.0]:
        formulas = [None, f'=ACT_DIST_PARETO4_PDF(A{row}, 0, 2, 0.5, 3)', f'=ACT_DIST_PARETO4_CDF(A{row}, 0, 2, 0.5, 3)', None]
        row = add_data_row(ws, [x, "", "", ""], row, formulas=formulas)
    distributions.append(("Pareto IV PDF (4-param)", section_start, data_start, row - 1, "x", "f(x)", 0, 5))
    row = max(row, section_start + SECTION_PADDING)

    # 23. Lognormal-Pareto (Composite)
    section_start = row
    row = add_note(ws, "LOGNORMAL-PARETO COMPOSITE (mu=0, sigma=1, theta=2)", row)
    row = add_table_header(ws, ["x", "PDF", "CDF", "Notes"], row)
    data_start = row
    for x in [0.1, 0.5, 1.0, 1.5, 2.0, 3.0, 4.0, 5.0, 7.0, 10.0]:
        formulas = [None, f'=ACT_DIST_LNPARETO_PDF(A{row}, 0, 1, 2)', f'=ACT_DIST_LNPARETO_CDF(A{row}, 0, 1, 2)', None]
        notes = "Threshold" if x == 2.0 else ""
        row = add_data_row(ws, [x, "", "", notes], row, formulas=formulas)
    distributions.append(("Lognormal-Pareto PDF", section_start, data_start, row - 1, "x", "f(x)", 0, None))
    row = max(row, section_start + SECTION_PADDING)

    # 24. Exponential-Pareto (Composite)
    section_start = row
    row = add_note(ws, "EXPONENTIAL-PARETO COMPOSITE (rate=0.5, theta=2)", row)
    row = add_table_header(ws, ["x", "PDF", "CDF", "Notes"], row)
    data_start = row
    for x in [0.0, 0.5, 1.0, 1.5, 2.0, 3.0, 4.0, 5.0, 7.0, 10.0]:
        formulas = [None, f'=ACT_DIST_EXPPARETO_PDF(A{row}, 0.5, 2)', f'=ACT_DIST_EXPPARETO_CDF(A{row}, 0.5, 2)', None]
        notes = "Threshold" if x == 2.0 else ""
        row = add_data_row(ws, [x, "", "", notes], row, formulas=formulas)
    distributions.append(("Exponential-Pareto PDF", section_start, data_start, row - 1, "x", "f(x)", 0, None))
    row = max(row, section_start + SECTION_PADDING)

    # 25. Power-Pareto (Composite)
    section_start = row
    row = add_note(ws, "POWER-PARETO COMPOSITE (alpha=2, beta=3, theta=1)", row)
    row = add_table_header(ws, ["x", "PDF", "CDF", "Notes"], row)
    data_start = row
    for x in [0.0, 0.25, 0.5, 0.75, 1.0, 1.5, 2.0, 3.0, 4.0, 5.0]:
        formulas = [None, f'=ACT_DIST_POWPARETO_PDF(A{row}, 2, 3, 1)', f'=ACT_DIST_POWPARETO_CDF(A{row}, 2, 3, 1)', None]
        notes = "Threshold" if x == 1.0 else ""
        row = add_data_row(ws, [x, "", "", notes], row, formulas=formulas)
    distributions.append(("Power-Pareto PDF", section_start, data_start, row - 1, "x", "f(x)", 0, None))

    # Create all charts aligned with their tables
    for title, section_start, data_start, data_end, x_title, y_title, x_min, x_max in distributions:
        anchor = f"{CHART_COL}{section_start}"
        create_scatter_chart_with_axes(ws, title, x_col=1, y_col=2,
            min_row=data_start-1, max_row=data_end, anchor=anchor,
            x_title=x_title, y_title=y_title,
            width=CHART_WIDTH, height=CHART_HEIGHT,
            x_min=x_min, y_min=0, x_max=x_max)

    return ws


def create_exposure_curves_sheet(wb):
    """Create the Exposure Curves examples sheet."""
    ws = wb.create_sheet("Exposure Curves")
    set_column_widths(ws, {'A': 15, 'B': 20, 'C': 20, 'D': 20, 'E': 30, 'F': 15})

    row = add_title(ws, "Exposure Curves")
    row = add_note(ws, "Exposure curves relate damage ratio (d) to expected loss ratio G(d)", row)
    row += 1

    # MBBEFD
    row = add_note(ws, "MBBEFD CURVES (b=2, g=3) - Swiss Re parameterization", row)
    row = add_table_header(ws, ["d", "G(d)", "Notes"], row)
    mbbefd_data_start = row
    for d in [0.0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0]:
        formulas = [None, f'=ACT_EXPOSURE_MBBEFD(A{row}, 2, 3)', None]
        row = add_data_row(ws, [d, "", ""], row, formulas=formulas)
    mbbefd_data_end = row - 1
    row += 1

    # Swiss Re curves comparison - moved up to combine with chart
    row = add_note(ws, "SWISS RE CURVES COMPARISON (all curves)", row)
    row = add_table_header(ws, ["d", "Curve 1", "Curve 2", "Curve 3", "Curve 4", "Curve 5"], row)
    swissre_data_start = row
    for d in [0.0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0]:
        ws.cell(row=row, column=1, value=d).border = THIN_BORDER
        for c in range(1, 6):
            ws.cell(row=row, column=c+1, value=f'=ACT_EXPOSURE_SWISSRE(A{row}, {c})').border = THIN_BORDER
        row += 1
    swissre_data_end = row - 1
    row += 1

    # Create multi-series chart for Swiss Re curves comparison
    chart = ScatterChart()
    chart.title = "Swiss Re Exposure Curves Comparison"
    chart.width = 14
    chart.height = 10
    chart.roundedCorners = False
    chart.x_axis.title = "Damage Ratio (d)"
    chart.y_axis.title = "Loss Ratio G(d)"
    chart.x_axis.scaling.min = 0
    chart.x_axis.scaling.max = 1
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1

    # Add each Swiss Re curve as a separate series
    x_values = Reference(ws, min_col=1, min_row=swissre_data_start, max_row=swissre_data_end)
    for curve_num in range(1, 6):
        y_values = Reference(ws, min_col=curve_num + 1, min_row=swissre_data_start - 1, max_row=swissre_data_end)
        series = Series(y_values, x_values, title=f"Curve {curve_num}")
        series.graphicalProperties.line.width = 20000  # 2pt line
        chart.series.append(series)

    ws.add_chart(chart, "G4")

    # Lloyd's curves
    row = add_note(ws, "LLOYD'S CURVES COMPARISON", row)
    row = add_table_header(ws, ["d", "Y1", "Y2", "Y3", "Y4"], row)
    lloyds_data_start = row
    for d in [0.0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0]:
        ws.cell(row=row, column=1, value=d).border = THIN_BORDER
        for i, curve in enumerate(["Y1", "Y2", "Y3", "Y4"]):
            ws.cell(row=row, column=i+2, value=f'=ACT_EXPOSURE_LLOYDS(A{row}, "{curve}")').border = THIN_BORDER
        row += 1
    lloyds_data_end = row - 1
    row += 1

    # Power and Pareto
    row = add_note(ws, "OTHER EXPOSURE CURVES at d=0.5", row)
    row = add_table_header(ws, ["Function", "Formula", "Result", "Notes"], row)
    curves = [
        ("Power (n=2)", '=ACT_EXPOSURE_POWER(0.5, 2)', "G(d) = d^n"),
        ("Inverse Power (n=2)", '=ACT_EXPOSURE_INVERSE_POWER(0.5, 2)', "G(d) = 1 - (1-d)^n"),
        ("Pareto (alpha=2)", '=ACT_EXPOSURE_PARETO(0.5, 2)', "Based on Pareto severity"),
        ("Riebesell (n=0.5)", '=ACT_EXPOSURE_RIEBESELL(0.5, 0.5)', "Alternative single-parameter curve"),
    ]
    for name, formula, notes in curves:
        ws.cell(row=row, column=1, value=name).border = THIN_BORDER
        ws.cell(row=row, column=2, value=formula.replace('=', '')).border = THIN_BORDER
        ws.cell(row=row, column=3, value=formula).border = THIN_BORDER
        ws.cell(row=row, column=4, value=notes).border = THIN_BORDER
        row += 1

    return ws


def create_reinsurance_sheet(wb):
    """Create the Reinsurance examples sheet."""
    ws = wb.create_sheet("Reinsurance")
    set_column_widths(ws, {'A': 20, 'B': 20, 'C': 20, 'D': 30})

    row = add_title(ws, "Reinsurance Functions")
    row = add_note(ws, "Excess of loss layer calculations", row)
    row += 1

    # XOL Layer
    row = add_note(ws, "XOL LAYER LOSS (Attachment=1,000,000, Limit=5,000,000)", row)
    row = add_table_header(ws, ["Ground-Up Loss", "Layer Loss", "Notes"], row)
    losses = [500000, 1000000, 2000000, 4000000, 6000000, 10000000]
    notes_list = ["Below attachment", "At attachment", "Partial layer", "Partial layer", "Full limit", "Capped at limit"]
    for loss, note in zip(losses, notes_list):
        cell_a = ws.cell(row=row, column=1, value=loss)
        cell_a.border = THIN_BORDER
        cell_a.number_format = '#,##0'
        cell_b = ws.cell(row=row, column=2, value=f'=ACT_XOL_LAYER_LOSS(A{row}, 1000000, 5000000)')
        cell_b.border = THIN_BORDER
        cell_b.number_format = '#,##0'
        cell_c = ws.cell(row=row, column=3, value=note)
        cell_c.border = THIN_BORDER
        row += 1

    return ws


def create_cat_modeling_sheet(wb):
    """Create the Cat Modeling examples sheet with ELT → YLT → EP workflow."""
    ws = wb.create_sheet("Cat Modeling")
    set_column_widths(ws, {'A': 12, 'B': 14, 'C': 14, 'D': 14, 'E': 4, 'F': 14, 'G': 14, 'H': 14})

    row = add_title(ws, "Cat Modeling: ELT → YLT → EP Curves")
    row = add_note(ws, "Simulate a Year Loss Table (YLT) from an Event Loss Table (ELT) and derive OEP/AEP curves.", row)
    row += 1

    # =========================================================================
    # SECTION 1: Simple ELT Example with YLT Simulation
    # =========================================================================
    row = add_note(ws, "EXAMPLE 1: ELT TO YLT SIMULATION", row)
    row += 1

    # ELT inputs
    row = add_note(ws, "Event Loss Table (ELT) - Input", row)
    row = add_table_header(ws, ["Event", "Annual Rate", "Event Loss"], row)
    elt_start_row = row
    events = [("Event 1", 0.2, 1_000_000), ("Event 2", 0.1, 2_500_000), ("Event 3", 0.05, 10_000_000)]
    for name, rate, loss in events:
        ws.cell(row=row, column=1, value=name).border = THIN_BORDER
        ws.cell(row=row, column=2, value=rate).border = THIN_BORDER
        loss_cell = ws.cell(row=row, column=3, value=loss)
        loss_cell.border = THIN_BORDER
        loss_cell.number_format = '#,##0'
        row += 1
    elt_end_row = row - 1
    row += 2

    # YLT output
    row = add_note(ws, "Year Loss Table (YLT) - Simulated 1000 years, seed=42", row)
    row = add_table_header(ws, ["Year", "Aggregate Loss", "Max Loss", "Event Count"], row)
    ylt_start_row = row
    ylt_end_row = ylt_start_row + 1000 - 1
    cell = ws.cell(row=row, column=1)
    cell.value = f'=ACT_CAT_ELT_TO_YLT(B{elt_start_row}:B{elt_end_row}, C{elt_start_row}:C{elt_end_row}, 1000, 42, FALSE)'
    cell.border = THIN_BORDER
    row = ylt_end_row + 2

    # OEP Curve from YLT
    row = add_note(ws, "OEP Curve (from YLT max losses) - Weibull plotting position", row)
    row = add_table_header(ws, ["Return Period", "OEP Loss"], row)
    oep_start = row
    cell = ws.cell(row=row, column=1)
    cell.value = f'=ACT_CAT_YLT_OEP_CURVE(C{ylt_start_row}:C{ylt_end_row}, "WEIBULL", FALSE)'
    cell.border = THIN_BORDER
    row += 2

    # AEP Curve from YLT
    row = add_note(ws, "AEP Curve (from YLT aggregate losses)", row)
    row = add_table_header(ws, ["Return Period", "AEP Loss"], row)
    aep_start = row
    cell = ws.cell(row=row, column=1)
    cell.value = f'=ACT_CAT_YLT_AEP_CURVE(B{ylt_start_row}:B{ylt_end_row}, "WEIBULL", FALSE)'
    cell.border = THIN_BORDER
    row += 3

    # =========================================================================
    # SECTION 2: Worked Example with Manual YLT and EP Curve Functions
    # =========================================================================
    row = add_note(ws, "EXAMPLE 2: WORKED EXAMPLE WITH DETAILED ELT", row)
    row += 1

    # ELT input (raw) - Column F onwards
    col_offset = 5  # Start at column F
    row2 = 4  # Start near the top, in a separate column area

    ws.cell(row=row2, column=col_offset+1, value="Event Loss Table (Detailed)").font = HEADER_FONT
    row2 += 1
    headers = ["Event", "Rate", "Mean", "SDi", "SDc", "TIV"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=row2, column=col_offset+1+i, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    row2 += 1

    elt_rows = [
        (1, 0.10, 500, 500, 200, 100000),
        (2, 0.10, 200, 400, 100, 5000),
        (3, 0.20, 300, 200, 400, 40000),
        (4, 0.10, 100, 300, 500, 4000),
        (5, 0.20, 500, 100, 200, 2000),
        (6, 0.25, 200, 200, 500, 50000),
        (7, 0.01, 1000, 500, 600, 100000),
        (8, 0.12, 250, 300, 100, 5000),
        (9, 0.14, 1000, 500, 200, 6000),
        (10, 0.00, 10000, 1000, 500, 1000000),
    ]
    elt2_start = row2
    for values in elt_rows:
        for i, v in enumerate(values):
            cell = ws.cell(row=row2, column=col_offset+1+i, value=v)
            cell.border = THIN_BORDER
        row2 += 1
    elt2_end = row2 - 1
    row2 += 2

    # YLT (worked example output - static data for reconciliation)
    ws.cell(row=row2, column=col_offset+1, value="Sample YLT Output (10 years)").font = HEADER_FONT
    row2 += 1
    for i, h in enumerate(["Year", "Loss", "Event"]):
        cell = ws.cell(row=row2, column=col_offset+1+i, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    row2 += 1

    ylt_rows = [
        (1, 0.0, "None"),
        (2, 85.74, 5),
        (3, 0.91, 2),
        (4, 261.18, 1),
        (5, 2.69, 8),
        (6, 262.10, "1,3"),
        (7, 0.0, "None"),
        (8, 0.00, 6),
        (9, 358.33, "3,6"),
        (10, 0.0, "None"),
    ]
    ylt2_start = row2
    for year, loss, event_id in ylt_rows:
        ws.cell(row=row2, column=col_offset+1, value=year).border = THIN_BORDER
        cell = ws.cell(row=row2, column=col_offset+2, value=loss)
        cell.border = THIN_BORDER
        cell.number_format = '#,##0.00'
        ws.cell(row=row2, column=col_offset+3, value=event_id).border = THIN_BORDER
        row2 += 1
    ylt2_end = row2 - 1
    row2 += 2

    # EP Curves at specific return periods
    ws.cell(row=row2, column=col_offset+1, value="EP Curves at Return Periods").font = HEADER_FONT
    row2 += 1
    for i, h in enumerate(["Return Period", "OEP", "AEP"]):
        cell = ws.cell(row=row2, column=col_offset+1+i, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    row2 += 1

    return_periods = [500, 250, 100, 50, 25, 10, 5, 2]
    rp_start = row2
    # Capture the return-period column once we know its end row.
    rp_end = rp_start + len(return_periods) - 1
    rp_col_letter = get_column_letter(col_offset + 1)
    rp_range = f"{rp_col_letter}{rp_start}:{rp_col_letter}{rp_end}"
    for rp in return_periods:
        ws.cell(row=row2, column=col_offset+1, value=rp).border = THIN_BORDER
        # OEP / AEP at requested return periods, drawn from the simulated YLT
        # (1000 years, seed 42) sitting at C{ylt_start_row}:C{ylt_end_row} and
        # B{...}:B{...} respectively.
        oep_formula = (f"=INDEX(ACT_CAT_OEP_CURVE_RP("
                       f"C{ylt_start_row}:C{ylt_end_row}, "
                       f"{rp_range}, FALSE), {row2 - rp_start + 1}, 2)")
        aep_formula = (f"=INDEX(ACT_CAT_AEP_CURVE_RP("
                       f"B{ylt_start_row}:B{ylt_end_row}, "
                       f"{rp_range}, FALSE), {row2 - rp_start + 1}, 2)")
        cell_oep = ws.cell(row=row2, column=col_offset+2, value=oep_formula)
        cell_oep.border = THIN_BORDER
        cell_oep.number_format = '#,##0'
        cell_aep = ws.cell(row=row2, column=col_offset+3, value=aep_formula)
        cell_aep.border = THIN_BORDER
        cell_aep.number_format = '#,##0'
        row2 += 1
    rp_end = row2 - 1

    # Note: These would use ACT_CAT_OEP_CURVE_RP and ACT_CAT_AEP_CURVE_RP with actual YLT data
    ws.cell(row=row2+1, column=col_offset+1,
            value="Note: Use ACT_CAT_OEP_CURVE_RP / ACT_CAT_AEP_CURVE_RP with YLT data").font = NOTE_FONT

    return ws


def create_interpolation_sheet(wb):
    """Create the Interpolation examples sheet."""
    ws = wb.create_sheet("Interpolation")
    set_column_widths(ws, {'A': 15, 'B': 15, 'C': 20, 'D': 20, 'E': 30, 'F': 5})

    row = add_title(ws, "Linear Interpolation")
    row = add_note(ws, "Interpolate between known points with optional extrapolation", row)
    row += 1

    # Known points
    row = add_note(ws, "KNOWN DATA POINTS", row)
    row = add_table_header(ws, ["X", "Y"], row)
    points = [(1, 10), (2, 20), (3, 35), (4, 55), (5, 80)]
    data_start_row = row
    for x, y in points:
        row = add_data_row(ws, [x, y], row)
    data_end_row = row - 1
    row += 1

    # Interpolation examples
    row = add_note(ws, "INTERPOLATION EXAMPLES", row)
    row = add_table_header(ws, ["X", "Y (FLAT extrap)", "Y (GRADIENT extrap)", "Notes"], row)
    x_range = f"$A${data_start_row}:$A${data_end_row}"
    y_range = f"$B${data_start_row}:$B${data_end_row}"

    interp_start_row = row
    test_x = [0.5, 1.5, 2.5, 3.5, 4.5, 5.5, 6.0]
    notes = ["Below range", "Interpolated", "Interpolated", "Interpolated", "Interpolated", "Above range", "Above range"]
    for x, note in zip(test_x, notes):
        flat_formula = f'=ACT_INTERP({x_range}, {y_range}, A{row}, "FLAT")'
        grad_formula = f'=ACT_INTERP({x_range}, {y_range}, A{row}, "GRADIENT")'
        ws.cell(row=row, column=1, value=x).border = THIN_BORDER
        ws.cell(row=row, column=2, value=flat_formula).border = THIN_BORDER
        ws.cell(row=row, column=3, value=grad_formula).border = THIN_BORDER
        ws.cell(row=row, column=4, value=note).border = THIN_BORDER
        row += 1
    interp_end_row = row - 1

    row += 1
    row = add_note(ws, "FLAT = extrapolate using last known value; GRADIENT = extrapolate using slope from last two points", row)

    # Scatter chart with two overlaid series:
    #   1) the original (X, Y) data with a connecting line
    #   2) the interpolated test points (column C, GRADIENT extrap) with visible
    #      markers in a contrasting colour, so it's obvious the interpolation
    #      goes through the input curve.
    chart = ScatterChart()
    chart.title = "Known Data Points + Interpolated"
    chart.width = 12
    chart.height = 8
    chart.roundedCorners = False
    chart.x_axis.title = "X"
    chart.y_axis.title = "Y"
    chart.x_axis.scaling.min = 0
    chart.y_axis.scaling.min = 0
    chart.x_axis.tickLblPos = "low"
    chart.y_axis.tickLblPos = "low"
    chart.x_axis.delete = False
    chart.y_axis.delete = False

    from openpyxl.chart import Series
    # Series 1: known data, line + small marker
    xvalues = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)
    yvalues = Reference(ws, min_col=2, min_row=data_start_row, max_row=data_end_row)
    s_known = Series(yvalues, xvalues, title="Known data")
    s_known.marker = Marker(symbol='circle', size=5)
    s_known.marker.graphicalProperties.solidFill = CHART_COLORS[0]
    s_known.graphicalProperties.line.solidFill = CHART_COLORS[0]
    chart.series.append(s_known)

    # Series 2: interpolated points (GRADIENT extrap), markers only, no line.
    # Different colour + larger marker so the overlay is clearly visible.
    xq_values = Reference(ws, min_col=1, min_row=interp_start_row, max_row=interp_end_row)
    yq_values = Reference(ws, min_col=3, min_row=interp_start_row, max_row=interp_end_row)
    s_interp = Series(yq_values, xq_values, title="Interpolated (GRADIENT)")
    s_interp.marker = Marker(symbol='diamond', size=9)
    s_interp.marker.graphicalProperties.solidFill = CHART_COLORS[3] if len(CHART_COLORS) > 3 else 'FF8800'
    s_interp.graphicalProperties.line.noFill = True
    chart.series.append(s_interp)

    chart.legend.position = 'b'

    ws.add_chart(chart, "F4")

    return ws


def create_chainladder_sheet(wb):
    """Create the Chain Ladder examples sheet."""
    ws = wb.create_sheet("Chain Ladder")
    set_column_widths(ws, {'A': 10, 'B': 12, 'C': 12, 'D': 12, 'E': 12, 'F': 12,
                           'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 12, 'L': 5, 'M': 15})

    row = add_title(ws, "Chain Ladder Reserving - Taylor-Ashe Dataset")
    row = add_note(ws, "Using the Taylor-Ashe triangle from Peter England's bootstrapping presentation (England & Verrall, 1999)", row)
    row += 1

    # Input triangle
    row = add_note(ws, "INPUT TRIANGLE (Cumulative Paid Losses - Taylor-Ashe)", row)
    row = add_table_header(ws, ["AY", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10"], row)

    triangle = [
        [357848, 1124788, 1735330, 2218270, 2745596, 3319994, 3466336, 3606286, 3833515, 3901463],
        [352118, 1236139, 2170033, 3353322, 3799067, 4120063, 4647867, 4914039, 5339085, None],
        [290507, 1292306, 2218525, 3235179, 3985995, 4132918, 4628910, 4909315, None, None],
        [310608, 1418858, 2195047, 3757447, 4029929, 4381982, 4588268, None, None, None],
        [443160, 1136350, 2128333, 2897821, 3402672, 3873311, None, None, None, None],
        [396132, 1333217, 2180715, 2985752, 3691712, None, None, None, None, None],
        [440832, 1288463, 2419861, 3483130, None, None, None, None, None, None],
        [359480, 1421128, 2864498, None, None, None, None, None, None, None],
        [376686, 1363294, None, None, None, None, None, None, None, None],
        [344014, None, None, None, None, None, None, None, None, None]
    ]

    triangle_start_row = row
    for i, tri_row in enumerate(triangle):
        ws.cell(row=row, column=1, value=i+1).border = THIN_BORDER
        for j, val in enumerate(tri_row):
            cell = ws.cell(row=row, column=j+2, value=val)
            cell.border = THIN_BORDER
            if val is not None:
                cell.number_format = '#,##0'
        row += 1
    triangle_end_row = row - 1
    row += 1

    tri_range = f"B{triangle_start_row}:K{triangle_end_row}"

    # Development factors - use INDEX to extract single values
    row = add_note(ws, "DEVELOPMENT FACTORS", row)
    row = add_note(ws, "Expected: 3.491, 1.747, 1.457, 1.174, 1.104, 1.086, 1.054, 1.077, 1.018", row)
    row = add_table_header(ws, ["Period", "1-2", "2-3", "3-4", "4-5", "5-6", "6-7", "7-8", "8-9", "9-10"], row)
    factors_row = row
    ws.cell(row=row, column=1, value="Factors").border = THIN_BORDER
    for i in range(9):
        ws.cell(row=row, column=i+2, value=f"=INDEX(ACT_CL_FACTORS({tri_range}), {i+1})").border = THIN_BORDER
    row += 2

    # Ultimates and IBNR
    row = add_note(ws, "PROJECTED ULTIMATES AND IBNR", row)
    row = add_note(ws, "Expected Total IBNR: ~18,680,856", row)
    row = add_table_header(ws, ["AY", "Latest Cumulative", "Ultimate", "IBNR"], row)

    ibnr_start = row
    for i in range(10):
        ws.cell(row=row, column=1, value=i+1).border = THIN_BORDER
        # Latest Cumulative used to be hard-coded constants; restore as a real
        # formula so a triangle edit propagates without the column drifting out
        # of sync.
        cell_latest = ws.cell(row=row, column=2, value=f"=INDEX(ACT_CL_LATEST({tri_range}), {i+1})")
        cell_latest.border = THIN_BORDER
        cell_latest.number_format = '#,##0'
        cell_ult = ws.cell(row=row, column=3, value=f"=INDEX(ACT_CL_ULTIMATE({tri_range}), {i+1})")
        cell_ult.border = THIN_BORDER
        cell_ult.number_format = '#,##0'
        cell_ibnr = ws.cell(row=row, column=4, value=f"=INDEX(ACT_CL_IBNR({tri_range}), {i+1})")
        cell_ibnr.border = THIN_BORDER
        cell_ibnr.number_format = '#,##0'
        row += 1
    ibnr_end = row - 1

    # Total row
    ws.cell(row=row, column=1, value="Total").font = HEADER_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    cell_sum_latest = ws.cell(row=row, column=2, value=f"=SUM(B{ibnr_start}:B{ibnr_end})")
    cell_sum_latest.border = THIN_BORDER
    cell_sum_latest.number_format = '#,##0'
    cell_sum_latest.font = HEADER_FONT
    cell_sum_ult = ws.cell(row=row, column=3, value=f"=SUM(C{ibnr_start}:C{ibnr_end})")
    cell_sum_ult.border = THIN_BORDER
    cell_sum_ult.number_format = '#,##0'
    cell_sum_ult.font = HEADER_FONT
    cell_sum_ibnr = ws.cell(row=row, column=4, value=f"=SUM(D{ibnr_start}:D{ibnr_end})")
    cell_sum_ibnr.border = THIN_BORDER
    cell_sum_ibnr.number_format = '#,##0'
    cell_sum_ibnr.font = HEADER_FONT
    row += 2

    # Mack Standard Errors with Mack 1993 reference (computed-from-scratch on
    # the Taylor-Ashe triangle; matches our C# implementation bit-exact). The
    # earlier "E&V Reference" column contained transcribed values that don't
    # actually match Mack's published algorithm — they've been replaced with
    # the values you get by running Mack 1993 §3 directly.
    row = add_note(ws, "MACK STANDARD ERRORS (reserve uncertainty)", row)
    row = add_table_header(ws, ["AY", "Reserve SE", "Mack 1993 reference"], row)

    mack_se_reference = [0, 75535, 121699, 133549, 261406, 411010, 558317,
                         875328, 971258, 1363155]

    se_start = row
    for i in range(10):
        ws.cell(row=row, column=1, value=i+1).border = THIN_BORDER
        cell_se = ws.cell(row=row, column=2, value=f"=INDEX(ACT_MACK_RESERVE_SE({tri_range}), {i+1})")
        cell_se.border = THIN_BORDER
        cell_se.number_format = '#,##0'
        cell_ref = ws.cell(row=row, column=3, value=mack_se_reference[i])
        cell_ref.border = THIN_BORDER
        cell_ref.number_format = '#,##0'
        row += 1
    se_end = row - 1

    # Total Mack SE — matches Mack (1993) §3 algorithm on Taylor-Ashe.
    ws.cell(row=row, column=1, value="Total").font = HEADER_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=3, value=2447095).border = THIN_BORDER
    ws.cell(row=row, column=3).number_format = '#,##0'
    row += 2

    # Bootstrap Chain Ladder
    row = add_note(ws, "BOOTSTRAP CHAIN LADDER (ODP) - England & Verrall (2002)", row)
    ws.cell(row=row, column=1, value="Iterations").border = THIN_BORDER
    ws.cell(row=row, column=2, value=1000).border = THIN_BORDER
    iter_cell = f"B{row}"
    row += 1
    ws.cell(row=row, column=1, value="Seed").border = THIN_BORDER
    ws.cell(row=row, column=2, value=42).border = THIN_BORDER
    seed_cell = f"B{row}"
    row += 1

    row = add_note(ws, "BOOTSTRAP RESULTS - EV method (England & Verrall 2002, non-constant scale)", row)
    row = add_table_header(ws, ["Statistic", "Value", "E&V Reference"], row)

    # E&V (2002) non-constant scale reference values
    bootstrap_reference = {
        "Mean": 18680856,
        "StdDev": 2228677,  # E&V non-constant scale total SE
    }

    bootstrap_stats = ["Mean", "StdDev", "P1", "P5", "P10", "P25", "P50", "P75", "P90", "P95", "P99"]
    for i, stat in enumerate(bootstrap_stats):
        ws.cell(row=row, column=1, value=stat).border = THIN_BORDER
        cell_val = ws.cell(row=row, column=2, value=f"=INDEX(ACT_CL_BOOTSTRAP({tri_range}, {iter_cell}, {seed_cell}), {i+1}, 2)")
        cell_val.border = THIN_BORDER
        cell_val.number_format = '#,##0'
        ref_val = bootstrap_reference.get(stat)
        if ref_val:
            cell_ref = ws.cell(row=row, column=3, value=ref_val)
            cell_ref.border = THIN_BORDER
            cell_ref.number_format = '#,##0'
        row += 1
    row += 1

    row = add_note(ws, "BOOTSTRAP BY ORIGIN YEAR - E&V (2002) Table 3, non-constant scale", row)
    row = add_table_header(ws, ["AY", "Mean", "StdDev", "E&V SE", "Ratio", "P50", "P75", "P90"], row)

    # E&V (2002) Table 3 - non-constant scale SE by origin year
    ev_se_by_origin = [0, 43882, 109449, 141509, 256031, 398377, 529898, 735245, 809457, 1285560]
    origin_data_start = row

    for i in range(10):
        ws.cell(row=row, column=1, value=i+1).border = THIN_BORDER
        # Mean
        cell_mean = ws.cell(row=row, column=2, value=f"=INDEX(ACT_CL_BOOTSTRAP_ORIGIN({tri_range}, {iter_cell}, {seed_cell}), {i+2}, 2)")
        cell_mean.border = THIN_BORDER
        cell_mean.number_format = '#,##0'
        # StdDev
        sd_cell_ref = f"C{row}"
        cell_sd = ws.cell(row=row, column=3, value=f"=INDEX(ACT_CL_BOOTSTRAP_ORIGIN({tri_range}, {iter_cell}, {seed_cell}), {i+2}, 3)")
        cell_sd.border = THIN_BORDER
        cell_sd.number_format = '#,##0'
        # E&V Reference SE
        ev_ref = ev_se_by_origin[i]
        cell_ref = ws.cell(row=row, column=4, value=ev_ref)
        cell_ref.border = THIN_BORDER
        cell_ref.number_format = '#,##0'
        # Ratio (our SE / E&V SE)
        if ev_ref > 0:
            cell_ratio = ws.cell(row=row, column=5, value=f"={sd_cell_ref}/D{row}")
            cell_ratio.border = THIN_BORDER
            cell_ratio.number_format = '0%'
        else:
            ws.cell(row=row, column=5, value="-").border = THIN_BORDER
        # P50
        cell_p50 = ws.cell(row=row, column=6, value=f"=INDEX(ACT_CL_BOOTSTRAP_ORIGIN({tri_range}, {iter_cell}, {seed_cell}), {i+2}, 4)")
        cell_p50.border = THIN_BORDER
        cell_p50.number_format = '#,##0'
        # P75
        cell_p75 = ws.cell(row=row, column=7, value=f"=INDEX(ACT_CL_BOOTSTRAP_ORIGIN({tri_range}, {iter_cell}, {seed_cell}), {i+2}, 5)")
        cell_p75.border = THIN_BORDER
        cell_p75.number_format = '#,##0'
        # P90
        cell_p90 = ws.cell(row=row, column=8, value=f"=INDEX(ACT_CL_BOOTSTRAP_ORIGIN({tri_range}, {iter_cell}, {seed_cell}), {i+2}, 6)")
        cell_p90.border = THIN_BORDER
        cell_p90.number_format = '#,##0'
        row += 1

    # Total row with E&V reference
    ws.cell(row=row, column=1, value="Total").font = HEADER_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=4, value=2228677).border = THIN_BORDER
    ws.cell(row=row, column=4).number_format = '#,##0'
    row += 1

    # Charts - IBNR bar chart
    chart1 = BarChart()
    chart1.title = "IBNR by Accident Year"
    chart1.type = "col"
    chart1.width = 12
    chart1.height = 8
    chart1.roundedCorners = False
    chart1.x_axis.title = "Accident Year"
    chart1.y_axis.title = "IBNR"
    chart1.y_axis.scaling.min = 0
    chart1.x_axis.tickLblPos = "low"
    chart1.y_axis.tickLblPos = "low"
    chart1.x_axis.delete = False
    chart1.y_axis.delete = False
    chart1.varyColors = False  # All bars same color
    data1 = Reference(ws, min_col=4, min_row=ibnr_start-1, max_row=ibnr_end)
    cats1 = Reference(ws, min_col=1, min_row=ibnr_start, max_row=ibnr_end)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    chart1.legend = None
    ws.add_chart(chart1, "L4")

    # Ultimates bar chart
    chart2 = BarChart()
    chart2.title = "Ultimate vs Latest by Accident Year"
    chart2.type = "col"
    chart2.grouping = "clustered"
    chart2.width = 12
    chart2.height = 8
    chart2.roundedCorners = False
    chart2.x_axis.title = "Accident Year"
    chart2.y_axis.title = "Amount"
    chart2.y_axis.scaling.min = 0
    chart2.x_axis.tickLblPos = "low"
    chart2.y_axis.tickLblPos = "low"
    chart2.x_axis.delete = False
    chart2.y_axis.delete = False
    data_latest = Reference(ws, min_col=2, min_row=ibnr_start-1, max_row=ibnr_end)
    data_ultimate = Reference(ws, min_col=3, min_row=ibnr_start-1, max_row=ibnr_end)
    chart2.add_data(data_latest, titles_from_data=True)
    chart2.add_data(data_ultimate, titles_from_data=True)
    cats2 = Reference(ws, min_col=1, min_row=ibnr_start, max_row=ibnr_end)
    chart2.set_categories(cats2)
    ws.add_chart(chart2, "L20")

    # Standard Error bar chart
    chart3 = BarChart()
    chart3.title = "Mack Standard Error by Accident Year"
    chart3.type = "col"
    chart3.width = 12
    chart3.height = 8
    chart3.roundedCorners = False
    chart3.x_axis.title = "Accident Year"
    chart3.y_axis.title = "Standard Error"
    chart3.y_axis.scaling.min = 0
    chart3.x_axis.tickLblPos = "low"
    chart3.y_axis.tickLblPos = "low"
    chart3.x_axis.delete = False
    chart3.y_axis.delete = False
    chart3.varyColors = False  # All bars same color
    data3 = Reference(ws, min_col=2, min_row=se_start-1, max_row=se_end)
    cats3 = Reference(ws, min_col=1, min_row=se_start, max_row=se_end)
    chart3.add_data(data3, titles_from_data=True)
    chart3.set_categories(cats3)
    chart3.legend = None
    ws.add_chart(chart3, "L36")

    return ws


def create_copulas_sheet(wb):
    """Create the Copulas examples sheet."""
    ws = wb.create_sheet("Copulas")
    set_column_widths(ws, {'A': 18, 'B': 12, 'C': 12, 'D': 12, 'E': 25, 'F': 15})

    row = add_title(ws, "Student-t Copula")
    row = add_note(ws, "Generate correlated uniform random numbers for Monte Carlo simulation", row)
    row += 1

    # Correlation matrix - lower triangle has values, upper triangle links to lower, diagonal = 1
    row = add_note(ws, "CORRELATION MATRIX (3x3)", row)
    row = add_table_header(ws, ["", "X1", "X2", "X3"], row)
    corr_start_row = row

    # Row 1: X1 - diagonal=1, others are formulas linking to lower triangle
    ws.cell(row=row, column=1, value="X1").font = HEADER_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=2, value=1.0).border = THIN_BORDER  # Diagonal
    ws.cell(row=row, column=3, value=f"=B{row+1}").border = THIN_BORDER  # Link to C6 -> B7
    ws.cell(row=row, column=4, value=f"=B{row+2}").border = THIN_BORDER  # Link to D6 -> B8
    row += 1

    # Row 2: X2 - lower triangle value, diagonal=1, upper links
    ws.cell(row=row, column=1, value="X2").font = HEADER_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=2, value=0.9).border = THIN_BORDER  # Editable: corr(X2,X1) = 0.9
    corr_x2_x1_cell = f"B{row}"  # Store reference for chart title
    ws.cell(row=row, column=3, value=1.0).border = THIN_BORDER  # Diagonal
    ws.cell(row=row, column=4, value=f"=C{row+1}").border = THIN_BORDER  # Link to D7 -> C8
    row += 1

    # Row 3: X3 - lower triangle values, diagonal=1
    ws.cell(row=row, column=1, value="X3").font = HEADER_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=2, value=0.3).border = THIN_BORDER  # Editable: corr(X3,X1)
    ws.cell(row=row, column=3, value=0.4).border = THIN_BORDER  # Editable: corr(X3,X2)
    ws.cell(row=row, column=4, value=1.0).border = THIN_BORDER  # Diagonal
    row += 1

    corr_end_row = row - 1

    # Add color scale conditional formatting (white to red) for correlation matrix
    corr_range_cf = f"B{corr_start_row}:D{corr_end_row}"
    ws.conditional_formatting.add(corr_range_cf,
        ColorScaleRule(start_type='num', start_value=0, start_color='FFFFFF',
                       end_type='num', end_value=1, end_color='FF6B6B'))

    row += 1

    corr_range = f"B{corr_start_row}:D{corr_end_row}"

    # Parameters
    row = add_note(ws, "PARAMETERS", row)
    ws.cell(row=row, column=1, value="Degrees of freedom:").border = THIN_BORDER
    ws.cell(row=row, column=2, value=5).border = THIN_BORDER
    df_cell = f"B{row}"
    row += 1
    ws.cell(row=row, column=1, value="Number of samples:").border = THIN_BORDER
    ws.cell(row=row, column=2, value=100).border = THIN_BORDER
    n_cell = f"B{row}"
    row += 1
    ws.cell(row=row, column=1, value="Random seed:").border = THIN_BORDER
    ws.cell(row=row, column=2, value=42).border = THIN_BORDER
    seed_cell = f"B{row}"
    row += 2

    # Output - use INDEX to extract individual values (all 100 samples for chart)
    row = add_note(ws, "GENERATED SAMPLES (use INDEX to extract from array)", row)
    row = add_note(ws, f"Formula: =INDEX(ACT_COPULA_STUDENT_T({corr_range}, {df_cell}, {n_cell}, {seed_cell}), row, col)", row)
    row = add_table_header(ws, ["Sample", "U1", "U2", "U3"], row)
    samples_start_row = row

    for i in range(100):  # All 100 samples for plotting
        ws.cell(row=row, column=1, value=i+1).border = THIN_BORDER
        for j in range(3):
            ws.cell(row=row, column=j+2, value=f"=INDEX(ACT_COPULA_STUDENT_T({corr_range}, {df_cell}, {n_cell}, {seed_cell}), {i+1}, {j+1})").border = THIN_BORDER
        row += 1
    sample_end_row = row - 1

    # Scatter chart with axes
    chart = ScatterChart()
    chart.title = "Student-t Copula Samples: U1 vs U2"
    chart.width = 10
    chart.height = 10
    chart.roundedCorners = False
    chart.x_axis.title = "U1"
    chart.y_axis.title = "U2"
    chart.x_axis.scaling.min = 0
    chart.x_axis.scaling.max = 1
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    chart.x_axis.tickLblPos = "low"
    chart.y_axis.tickLblPos = "low"
    chart.x_axis.delete = False
    chart.y_axis.delete = False

    from openpyxl.chart import Series
    xvalues = Reference(ws, min_col=2, min_row=samples_start_row, max_row=sample_end_row)
    yvalues = Reference(ws, min_col=3, min_row=samples_start_row, max_row=sample_end_row)
    series = Series(yvalues, xvalues, title_from_data=False)
    series.marker = Marker(symbol='circle', size=5)
    series.marker.graphicalProperties.solidFill = CHART_COLORS[0]  # All points same color
    series.graphicalProperties.line.noFill = True
    chart.series.append(series)
    chart.legend = None

    ws.add_chart(chart, "F4")

    row += 1
    row = add_note(ws, "Note: Results are uniform[0,1] values. Apply inverse CDF to get desired marginal distributions.", row)
    row += 2

    # ─────────────────────────────────────────────────────────────────────────
    # Additional copula families. Per the alignment principle every package
    # function appears in the spreadsheet, each Archimedean copula gets its own
    # mini-section: theta input cell, 50-sample table, and a U1 vs U2 scatter.
    # Gaussian gets a 2x2 correlation matrix (df-free analogue of Student-t).
    # Layout: stacked below the Student-t section in column A-D for the data,
    # chart placed in column F at the same vertical offset.
    # ─────────────────────────────────────────────────────────────────────────

    def _emit_copula_section(title, fn_name, theta_label, theta_value,
                             extra_param_label=None, extra_param_value=None,
                             chart_anchor_col='F'):
        nonlocal row
        section_title_row = row
        # add_title defaults to row=1 if you don't pass current row — that bug
        # collapsed all four sections on top of each other on the first build.
        row = add_title(ws, title, row=row)
        n_samples = 50
        ws.cell(row=row, column=1, value=theta_label).border = THIN_BORDER
        ws.cell(row=row, column=2, value=theta_value).border = THIN_BORDER
        theta_cell = f"B{row}"
        row += 1
        if extra_param_label is not None:
            ws.cell(row=row, column=1, value=extra_param_label).border = THIN_BORDER
            ws.cell(row=row, column=2, value=extra_param_value).border = THIN_BORDER
            extra_cell = f"B{row}"
            row += 1
        else:
            extra_cell = None
        ws.cell(row=row, column=1, value="Number of samples:").border = THIN_BORDER
        ws.cell(row=row, column=2, value=n_samples).border = THIN_BORDER
        n_cell = f"B{row}"
        row += 1
        ws.cell(row=row, column=1, value="Random seed:").border = THIN_BORDER
        ws.cell(row=row, column=2, value=42).border = THIN_BORDER
        seed_cell_local = f"B{row}"
        row += 2

        row = add_table_header(ws, ["Sample", "U1", "U2"], row)
        sample_start = row
        if extra_cell is None:
            sample_args = f"{theta_cell}, {n_cell}, {seed_cell_local}"
        else:
            sample_args = f"{theta_cell}, {extra_cell}, {n_cell}, {seed_cell_local}"
        for i in range(n_samples):
            ws.cell(row=row, column=1, value=i+1).border = THIN_BORDER
            for j in range(2):
                ws.cell(row=row, column=j+2,
                    value=f"=INDEX({fn_name}({sample_args}), {i+1}, {j+1})").border = THIN_BORDER
            row += 1
        sample_end = row - 1

        # Scatter chart in column F at the section's vertical anchor
        sc = ScatterChart()
        sc.title = f"{title}: U1 vs U2"
        sc.width = 10
        sc.height = 8
        sc.roundedCorners = False
        sc.x_axis.title = "U1"; sc.y_axis.title = "U2"
        sc.x_axis.scaling.min = 0; sc.x_axis.scaling.max = 1
        sc.y_axis.scaling.min = 0; sc.y_axis.scaling.max = 1
        sc.x_axis.tickLblPos = "low"; sc.y_axis.tickLblPos = "low"
        sc.x_axis.delete = False; sc.y_axis.delete = False
        xv = Reference(ws, min_col=2, min_row=sample_start, max_row=sample_end)
        yv = Reference(ws, min_col=3, min_row=sample_start, max_row=sample_end)
        srs = Series(yv, xv, title_from_data=False)
        srs.marker = Marker(symbol='circle', size=5)
        srs.marker.graphicalProperties.solidFill = CHART_COLORS[0]
        srs.graphicalProperties.line.noFill = True
        sc.series.append(srs)
        sc.legend = None
        ws.add_chart(sc, f"{chart_anchor_col}{section_title_row}")
        row += 2

    # Gaussian copula needs a 2x2 correlation matrix (a *range* on the sheet,
    # not a scalar theta) — Excel array constants `{...}` can't reference cells.
    # Lay out a 2x2 matrix block first, then reference that range.
    section_title_row = row
    row = add_title(ws, "Gaussian Copula", row=row)
    ws.cell(row=row, column=1, value="Correlation matrix:").border = THIN_BORDER
    row += 1
    corr_top = row
    ws.cell(row=row, column=2, value=1.0).border = THIN_BORDER
    ws.cell(row=row, column=3, value=0.7).border = THIN_BORDER
    row += 1
    ws.cell(row=row, column=2, value=f"=C{corr_top}").border = THIN_BORDER
    ws.cell(row=row, column=3, value=1.0).border = THIN_BORDER
    row += 1
    corr_bot = row - 1
    g_corr_range = f"B{corr_top}:C{corr_bot}"
    ws.cell(row=row, column=1, value="Number of samples:").border = THIN_BORDER
    ws.cell(row=row, column=2, value=50).border = THIN_BORDER
    g_n_cell = f"B{row}"
    row += 1
    ws.cell(row=row, column=1, value="Random seed:").border = THIN_BORDER
    ws.cell(row=row, column=2, value=42).border = THIN_BORDER
    g_seed_cell = f"B{row}"
    row += 2

    row = add_table_header(ws, ["Sample", "U1", "U2"], row)
    g_sample_start = row
    for i in range(50):
        ws.cell(row=row, column=1, value=i+1).border = THIN_BORDER
        for j in range(2):
            ws.cell(row=row, column=j+2,
                    value=f"=INDEX(ACT_COPULA_GAUSSIAN({g_corr_range}, {g_n_cell}, {g_seed_cell}), {i+1}, {j+1})").border = THIN_BORDER
        row += 1
    g_sample_end = row - 1

    sc = ScatterChart()
    sc.title = "Gaussian Copula: U1 vs U2"
    sc.width = 10; sc.height = 8; sc.roundedCorners = False
    sc.x_axis.title = "U1"; sc.y_axis.title = "U2"
    sc.x_axis.scaling.min = 0; sc.x_axis.scaling.max = 1
    sc.y_axis.scaling.min = 0; sc.y_axis.scaling.max = 1
    sc.x_axis.tickLblPos = "low"; sc.y_axis.tickLblPos = "low"
    sc.x_axis.delete = False; sc.y_axis.delete = False
    xv = Reference(ws, min_col=2, min_row=g_sample_start, max_row=g_sample_end)
    yv = Reference(ws, min_col=3, min_row=g_sample_start, max_row=g_sample_end)
    srs = Series(yv, xv, title_from_data=False)
    srs.marker = Marker(symbol='circle', size=5)
    srs.marker.graphicalProperties.solidFill = CHART_COLORS[0]
    srs.graphicalProperties.line.noFill = True
    sc.series.append(srs)
    sc.legend = None
    ws.add_chart(sc, f"F{section_title_row}")
    row += 2

    _emit_copula_section(
        title="Clayton Copula (lower tail)",
        fn_name="ACT_COPULA_CLAYTON",
        theta_label="Theta (>0 for positive dependence):",
        theta_value=2.0,
    )
    _emit_copula_section(
        title="Frank Copula (symmetric)",
        fn_name="ACT_COPULA_FRANK",
        theta_label="Theta (>0 positive, <0 negative):",
        theta_value=5.0,
    )
    _emit_copula_section(
        title="Gumbel Copula (upper tail)",
        fn_name="ACT_COPULA_GUMBEL",
        theta_label="Theta (>=1):",
        theta_value=2.0,
    )

    return ws


def create_aggregate_claims_sheet(wb):
    """Create the Aggregate Claims (Panjer Recursion) examples sheet."""
    ws = wb.create_sheet("Aggregate Claims")
    set_column_widths(ws, {'A': 12, 'B': 16, 'C': 14, 'D': 14, 'E': 5, 'F': 16, 'G': 14, 'H': 14, 'I': 14})

    row = add_title(ws, "Aggregate Claims - Panjer Recursion")
    row = add_note(ws, "Compute S = X1 + X2 + ... + Xn where N ~ frequency, Xi ~ severity (iid)", row)
    row = add_note(ws, "Workflow: discretize severity -> Panjer recursion -> aggregate statistics", row)
    row += 1

    # =========================================================================
    # SECTION 1: Severity Discretization
    # =========================================================================
    row = add_note(ws, "SECTION 1: SEVERITY DISCRETIZATION", row)
    row = add_note(ws, "Exponential(rate=1) discretized on grid h=0.5, m=20 points", row)
    row += 1

    # Parameters
    ws.cell(row=row, column=1, value="Rate:").font = HEADER_FONT
    ws.cell(row=row, column=2, value=1)
    rate_cell = f"B{row}"
    row += 1
    ws.cell(row=row, column=1, value="Grid spacing (h):").font = HEADER_FONT
    ws.cell(row=row, column=2, value=0.5)
    h_cell = f"B{row}"
    row += 1
    ws.cell(row=row, column=1, value="Grid points (m):").font = HEADER_FONT
    ws.cell(row=row, column=2, value=20)
    m_cell = f"B{row}"
    row += 2

    # Discretized severity table
    row = add_table_header(ws, ["Index (s)", "Value (s*h)", "PMF f(s)"], row)
    sev_data_start = row
    for s in range(20):
        ws.cell(row=row, column=1, value=s).border = THIN_BORDER
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2, value=s * 0.5).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        if s == 0:
            ws.cell(row=row, column=3,
                    value=f'=INDEX(ACT_DISCRETIZE_EXPONENTIAL({rate_cell}, {h_cell}, {m_cell}), {s+1})').border = THIN_BORDER
        else:
            ws.cell(row=row, column=3,
                    value=f'=INDEX(ACT_DISCRETIZE_EXPONENTIAL({rate_cell}, {h_cell}, {m_cell}), {s+1})').border = THIN_BORDER
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=3).number_format = '0.000000'
        row += 1
    sev_data_end = row - 1

    # Severity PMF range for Panjer
    sev_range = f"C{sev_data_start}:C{sev_data_end}"
    row += 1

    # =========================================================================
    # SECTION 2: Panjer Recursion - Poisson frequency
    # =========================================================================
    row = add_note(ws, "SECTION 2: PANJER RECURSION - Poisson(lambda=2) frequency", row)
    row = add_note(ws, "Aggregate PMF via Panjer recursion, max_s=40 -> 41 values", row)
    row += 1

    ws.cell(row=row, column=1, value="Lambda:").font = HEADER_FONT
    ws.cell(row=row, column=2, value=2)
    lambda_cell = f"B{row}"
    row += 1
    ws.cell(row=row, column=1, value="Max index:").font = HEADER_FONT
    ws.cell(row=row, column=2, value=40)
    max_s_cell = f"B{row}"
    row += 2

    row = add_table_header(ws, ["Index (s)", "Aggregate Loss (s*h)", "PMF P(S=s*h)", "CDF P(S<=s*h)"], row)
    agg_data_start = row
    for s in range(41):
        ws.cell(row=row, column=1, value=s).border = THIN_BORDER
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2, value=s * 0.5).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=3,
                value=f'=INDEX(ACT_PANJER_POISSON({lambda_cell}, {sev_range}, {max_s_cell}), {s+1})').border = THIN_BORDER
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=3).number_format = '0.000000'
        # CDF = sum of PMF up to this point
        ws.cell(row=row, column=4,
                value=f'=SUM(C{agg_data_start}:C{row})').border = THIN_BORDER
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=4).number_format = '0.000000'
        row += 1
    agg_data_end = row - 1

    # Aggregate PMF range for statistics
    agg_range = f"C{agg_data_start}:C{agg_data_end}"
    row += 1

    # =========================================================================
    # SECTION 3: Aggregate Statistics
    # =========================================================================
    row = add_note(ws, "SECTION 3: AGGREGATE STATISTICS", row)
    row += 1

    row = add_table_header(ws, ["Statistic", "Formula Result", "Theoretical"], row)
    # Theoretical reference column: plain-text descriptions (no leading "="
    # — that would make openpyxl write them as formulas and Excel would
    # silently strip them on open with "Removed Records: Formula").
    stats = [
        ("Mean",     f'=ACT_AGGREGATE_MEAN({agg_range}, {h_cell})',     "2/1 = 2.0 (lambda/rate)"),
        ("Std Dev",  f'=ACT_AGGREGATE_STDEV({agg_range}, {h_cell})',    "sqrt(lambda*(1/rate^2 + 1/rate^2)) = 2.0"),
        ("Variance", f'=ACT_AGGREGATE_VAR_STAT({agg_range}, {h_cell})', "4.0"),
        ("VaR(95%)", f'=ACT_AGGREGATE_VAR(0.95, {agg_range}, {h_cell})', ""),
        ("TVaR(95%)", f'=ACT_AGGREGATE_TVAR(0.95, {agg_range}, {h_cell})', ""),
        ("CDF at x=3", f'=ACT_AGGREGATE_CDF(3, {agg_range}, {h_cell})', ""),
    ]
    for label, formula, theoretical in stats:
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        cell_f = ws.cell(row=row, column=2, value=formula)
        cell_f.border = THIN_BORDER
        cell_f.number_format = '0.0000'
        ws.cell(row=row, column=3, value=theoretical).border = THIN_BORDER
        ws.cell(row=row, column=3).font = NOTE_FONT
        row += 1
    row += 1

    # =========================================================================
    # SECTION 4: Alternative Frequency - Negative Binomial
    # =========================================================================
    row = add_note(ws, "SECTION 4: ALTERNATIVE FREQUENCY - NegBin(r=2, p=0.5)", row)
    row = add_note(ws, "Same severity, different frequency distribution", row)
    row += 1

    ws.cell(row=row, column=1, value="r:").font = HEADER_FONT
    ws.cell(row=row, column=2, value=2)
    nb_r_cell = f"B{row}"
    row += 1
    ws.cell(row=row, column=1, value="p:").font = HEADER_FONT
    ws.cell(row=row, column=2, value=0.5)
    nb_p_cell = f"B{row}"
    row += 2

    row = add_table_header(ws, ["Index (s)", "Aggregate Loss (s*h)", "PMF (NegBin)", "CDF (NegBin)"], row)
    nb_data_start = row
    for s in range(41):
        ws.cell(row=row, column=1, value=s).border = THIN_BORDER
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2, value=s * 0.5).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=3,
                value=f'=INDEX(ACT_PANJER_NEGBIN({nb_r_cell}, {nb_p_cell}, {sev_range}, {max_s_cell}), {s+1})').border = THIN_BORDER
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=3).number_format = '0.000000'
        ws.cell(row=row, column=4,
                value=f'=SUM(C{nb_data_start}:C{row})').border = THIN_BORDER
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=4).number_format = '0.000000'
        row += 1
    nb_data_end = row - 1

    nb_agg_range = f"C{nb_data_start}:C{nb_data_end}"
    row += 1

    # NegBin statistics
    row = add_note(ws, "NEGBIN AGGREGATE STATISTICS", row)
    row = add_table_header(ws, ["Statistic", "Formula Result", "Notes"], row)
    nb_stats = [
        ("Mean",    f'=ACT_AGGREGATE_MEAN({nb_agg_range}, {h_cell})',   "E[N]*E[X] = r(1-p)/p / rate = 2"),
        ("Std Dev", f'=ACT_AGGREGATE_STDEV({nb_agg_range}, {h_cell})', ""),
        ("VaR(95%)", f'=ACT_AGGREGATE_VAR(0.95, {nb_agg_range}, {h_cell})', ""),
        ("TVaR(95%)", f'=ACT_AGGREGATE_TVAR(0.95, {nb_agg_range}, {h_cell})', ""),
    ]
    for label, formula, notes in nb_stats:
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        cell_f = ws.cell(row=row, column=2, value=formula)
        cell_f.border = THIN_BORDER
        cell_f.number_format = '0.0000'
        ws.cell(row=row, column=3, value=notes).border = THIN_BORDER
        ws.cell(row=row, column=3).font = NOTE_FONT
        row += 1

    # =========================================================================
    # Chart: Aggregate PMF bar chart (Poisson, first 30 points)
    # =========================================================================
    chart = BarChart()
    chart.title = "Aggregate Claims PMF - Poisson(lambda=2) + Exp(rate=1)"
    chart.type = "col"
    chart.grouping = "standard"
    chart.width = 14
    chart.height = 9
    chart.roundedCorners = False
    chart.x_axis.title = "Aggregate Loss (s * h)"
    chart.y_axis.title = "Probability"
    chart.y_axis.scaling.min = 0
    chart.x_axis.tickLblPos = "low"
    chart.y_axis.tickLblPos = "low"
    chart.x_axis.delete = False
    chart.y_axis.delete = False

    # Use first 30 points for readability
    chart_end = agg_data_start + 29
    data = Reference(ws, min_col=3, min_row=agg_data_start - 1, max_row=chart_end)
    cats = Reference(ws, min_col=2, min_row=agg_data_start, max_row=chart_end)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None

    ws.add_chart(chart, "F4")

    return ws


def create_all_functions_test_sheet(wb):
    """Create the All Functions Test sheet from scratch.

    Generates a comprehensive test sheet with every add-in function.
    This replaces the old approach of preserving/patching an existing tab,
    which caused openpyxl drawing relationship corruption.
    """
    ws = wb.create_sheet("All Functions Test")
    set_column_widths(ws, {'A': 35, 'B': 100, 'C': 30})

    section_font = Font(bold=True, size=11)
    subsection_font = Font(bold=True)

    row = 1
    ws.cell(row=row, column=1, value="COMPREHENSIVE FUNCTION TESTS").font = TITLE_FONT
    row += 2

    def section(title):
        nonlocal row
        ws.cell(row=row, column=1, value=title).font = section_font
        row += 1

    def subsection(title):
        nonlocal row
        ws.cell(row=row, column=1, value=title).font = subsection_font
        row += 1

    def test(label, formula, desc=None):
        nonlocal row
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=formula)
        if desc:
            ws.cell(row=row, column=3, value=desc)
        row += 1

    def fit(label, formula, params):
        """FIT function: single-param plain, multi-param TRANSPOSE."""
        nonlocal row
        ws.cell(row=row, column=1, value=label)
        if len(params) == 1:
            ws.cell(row=row, column=2, value=formula)
            ws.cell(row=row, column=3, value=params[0])
        else:
            inner = formula[1:]  # strip leading =
            ws.cell(row=row, column=2, value=f'=TRANSPOSE({inner})')
            ws.cell(row=row, column=3, value=params[0])
        row += 1
        for p in params[1:]:
            ws.cell(row=row, column=3, value=p)
            row += 1

    def array_test(label, formula, desc=None):
        """Array-returning function wrapped in INDEX for test tab."""
        nonlocal row
        ws.cell(row=row, column=1, value=label)
        inner = formula[1:]  # strip =
        ws.cell(row=row, column=2, value=f'=INDEX({inner}, 1, 1)')
        if desc:
            ws.cell(row=row, column=3, value=desc)
        row += 1

    def blank():
        nonlocal row
        row += 1

    # ── DISTRIBUTIONS - Frequency ──
    section("DISTRIBUTIONS - Frequency")
    blank()
    subsection("Poisson Distribution (lambda=5)")
    test("ACT_DIST_POISSON_PDF", "=ACT_DIST_POISSON_PDF(3, 5)", "P(X=3) for Poisson")
    test("ACT_DIST_POISSON_CDF", "=ACT_DIST_POISSON_CDF(3, 5)", "P(X<=3)")
    test("ACT_DIST_POISSON_INV", "=ACT_DIST_POISSON_INV(0.5, 5)", "Median")
    fit("ACT_DIST_POISSON_FIT({1,2,3,4,5,6,7})",
        "=ACT_DIST_POISSON_FIT({1,2,3,4,5,6,7})", ["lambda"])
    blank()
    subsection("Negative Binomial (r=5, p=0.3)")
    test("ACT_DIST_NEGBIN_PDF", "=ACT_DIST_NEGBIN_PDF(5, 5, 0.3)", "P(X=5)")
    test("ACT_DIST_NEGBIN_CDF", "=ACT_DIST_NEGBIN_CDF(5, 5, 0.3)", "P(X<=5)")
    test("ACT_DIST_NEGBIN_INV", "=ACT_DIST_NEGBIN_INV(0.5, 5, 0.3)", "Median")
    fit("ACT_DIST_NEGBIN_FIT({1,2,3,4,5,6,7,8,9,10})",
        "=ACT_DIST_NEGBIN_FIT({1,2,3,4,5,6,7,8,9,10})",
        ["r (successes)", "p (probability)"])
    blank()

    # ── DISTRIBUTIONS - Zero-Truncated ──
    section("DISTRIBUTIONS - Zero-Truncated")
    blank()
    subsection("Zero-Truncated Poisson (lambda=3)")
    test("ACT_DIST_ZTPOISSON_PDF", "=ACT_DIST_ZTPOISSON_PDF(2, 3)", "P(X=2|X>0)")
    test("ACT_DIST_ZTPOISSON_CDF", "=ACT_DIST_ZTPOISSON_CDF(2, 3)", "P(X<=2|X>0)")
    test("ACT_DIST_ZTPOISSON_INV", "=ACT_DIST_ZTPOISSON_INV(0.5, 3)", "Median")
    test("ACT_DIST_ZTPOISSON_MEAN", "=ACT_DIST_ZTPOISSON_MEAN(3)", "E[X|X>0]")
    blank()
    subsection("Zero-Truncated Negative Binomial (r=5, p=0.3)")
    test("ACT_DIST_ZTNEGBIN_PDF", "=ACT_DIST_ZTNEGBIN_PDF(3, 5, 0.3)", "P(X=3|X>0)")
    test("ACT_DIST_ZTNEGBIN_CDF", "=ACT_DIST_ZTNEGBIN_CDF(3, 5, 0.3)", "P(X<=3|X>0)")
    test("ACT_DIST_ZTNEGBIN_INV", "=ACT_DIST_ZTNEGBIN_INV(0.5, 5, 0.3)", "Median")
    test("ACT_DIST_ZTNEGBIN_MEAN", "=ACT_DIST_ZTNEGBIN_MEAN(5, 0.3)", "E[X|X>0]")
    blank()
    subsection("Zero-Truncated Binomial (n=10, p=0.3)")
    test("ACT_DIST_ZTBINOM_PDF", "=ACT_DIST_ZTBINOM_PDF(3, 10, 0.3)", "P(X=3|X>0)")
    test("ACT_DIST_ZTBINOM_CDF", "=ACT_DIST_ZTBINOM_CDF(3, 10, 0.3)", "P(X<=3|X>0)")
    test("ACT_DIST_ZTBINOM_INV", "=ACT_DIST_ZTBINOM_INV(0.5, 10, 0.3)", "Median")
    test("ACT_DIST_ZTBINOM_MEAN", "=ACT_DIST_ZTBINOM_MEAN(10, 0.3)", "E[X|X>0]")
    blank()
    subsection("Zero-Truncated Geometric (p=0.3)")
    test("ACT_DIST_ZTGEOM_PDF", "=ACT_DIST_ZTGEOM_PDF(2, 0.3)", "P(X=2|X>0)")
    test("ACT_DIST_ZTGEOM_CDF", "=ACT_DIST_ZTGEOM_CDF(2, 0.3)", "P(X<=2|X>0)")
    test("ACT_DIST_ZTGEOM_INV", "=ACT_DIST_ZTGEOM_INV(0.5, 0.3)", "Median")
    test("ACT_DIST_ZTGEOM_MEAN", "=ACT_DIST_ZTGEOM_MEAN(0.3)", "E[X|X>0]")
    blank()

    # ── DISTRIBUTIONS - Zero-Modified ──
    section("DISTRIBUTIONS - Zero-Modified")
    blank()
    subsection("Zero-Modified Poisson (lambda=3, p0=0.2)")
    test("ACT_DIST_ZMPOISSON_PDF", "=ACT_DIST_ZMPOISSON_PDF(2, 3, 0.2)", "P(X=2)")
    test("ACT_DIST_ZMPOISSON_CDF", "=ACT_DIST_ZMPOISSON_CDF(2, 3, 0.2)", "P(X<=2)")
    test("ACT_DIST_ZMPOISSON_INV", "=ACT_DIST_ZMPOISSON_INV(0.5, 3, 0.2)", "Median")
    test("ACT_DIST_ZMPOISSON_MEAN", "=ACT_DIST_ZMPOISSON_MEAN(3, 0.2)", "E[X]")
    test("ACT_DIST_ZMPOISSON_VAR", "=ACT_DIST_ZMPOISSON_VAR(3, 0.2)", "Var(X)")
    blank()
    subsection("Zero-Modified Negative Binomial (r=5, p=0.3, p0=0.2)")
    test("ACT_DIST_ZMNEGBIN_PDF", "=ACT_DIST_ZMNEGBIN_PDF(3, 5, 0.3, 0.2)", "P(X=3)")
    test("ACT_DIST_ZMNEGBIN_CDF", "=ACT_DIST_ZMNEGBIN_CDF(3, 5, 0.3, 0.2)", "P(X<=3)")
    test("ACT_DIST_ZMNEGBIN_INV", "=ACT_DIST_ZMNEGBIN_INV(0.5, 5, 0.3, 0.2)", "Median")
    test("ACT_DIST_ZMNEGBIN_MEAN", "=ACT_DIST_ZMNEGBIN_MEAN(5, 0.3, 0.2)", "E[X]")
    test("ACT_DIST_ZMNEGBIN_VAR", "=ACT_DIST_ZMNEGBIN_VAR(5, 0.3, 0.2)", "Var(X)")
    blank()

    # ── DISTRIBUTIONS - Severity ──
    section("DISTRIBUTIONS - Severity")
    blank()
    subsection("Lognormal (mu=0, sigma=1)")
    test("ACT_DIST_LOGNORM_PDF", "=ACT_DIST_LOGNORM_PDF(1, 0, 1)", "f(1)")
    test("ACT_DIST_LOGNORM_CDF", "=ACT_DIST_LOGNORM_CDF(1, 0, 1)", "F(1)")
    test("ACT_DIST_LOGNORM_INV", "=ACT_DIST_LOGNORM_INV(0.5, 0, 1)", "Median = 1")
    fit("ACT_DIST_LOGNORM_FIT({1,2,3,4,5})",
        "=ACT_DIST_LOGNORM_FIT({1,2,3,4,5})", ["mu", "sigma"])
    blank()
    subsection("Gamma (alpha=2, beta=1)")
    test("ACT_DIST_GAMMA_PDF", "=ACT_DIST_GAMMA_PDF(1, 2, 1)", "f(1)")
    test("ACT_DIST_GAMMA_CDF", "=ACT_DIST_GAMMA_CDF(1, 2, 1)", "F(1)")
    test("ACT_DIST_GAMMA_INV", "=ACT_DIST_GAMMA_INV(0.5, 2, 1)", "Median")
    fit("ACT_DIST_GAMMA_FIT({1,2,3,4,5})",
        "=ACT_DIST_GAMMA_FIT({1,2,3,4,5})", ["alpha (shape)", "beta (rate)"])
    blank()
    subsection("Pareto (alpha=2, xm=1)")
    test("ACT_DIST_PARETO_PDF", "=ACT_DIST_PARETO_PDF(2, 2, 1)", "f(2)")
    test("ACT_DIST_PARETO_CDF", "=ACT_DIST_PARETO_CDF(2, 2, 1)", "F(2) = 0.75")
    test("ACT_DIST_PARETO_INV", "=ACT_DIST_PARETO_INV(0.5, 2, 1)", "Median")
    fit("ACT_DIST_PARETO_FIT({1.5,2,2.5,3,4,5}, 1)",
        "=ACT_DIST_PARETO_FIT({1.5,2,2.5,3,4,5}, 1)", ["alpha", "xm"])
    blank()
    subsection("Weibull (k=2, lambda=1)")
    test("ACT_DIST_WEIBULL_PDF", "=ACT_DIST_WEIBULL_PDF(0.5, 2, 1)", "f(0.5)")
    test("ACT_DIST_WEIBULL_CDF", "=ACT_DIST_WEIBULL_CDF(0.5, 2, 1)", "F(0.5)")
    test("ACT_DIST_WEIBULL_INV", "=ACT_DIST_WEIBULL_INV(0.5, 2, 1)", "Median")
    fit("ACT_DIST_WEIBULL_FIT({0.5,1,1.5,2,2.5})",
        "=ACT_DIST_WEIBULL_FIT({0.5,1,1.5,2,2.5})", ["k (shape)", "lambda (scale)"])
    blank()
    subsection("Beta (alpha=2, beta=5)")
    test("ACT_DIST_BETA_PDF", "=ACT_DIST_BETA_PDF(0.3, 2, 5)", "f(0.3)")
    test("ACT_DIST_BETA_CDF", "=ACT_DIST_BETA_CDF(0.3, 2, 5)", "F(0.3)")
    test("ACT_DIST_BETA_INV", "=ACT_DIST_BETA_INV(0.5, 2, 5)", "Median")
    fit("ACT_DIST_BETA_FIT({0.1,0.2,0.3,0.4})",
        "=ACT_DIST_BETA_FIT({0.1,0.2,0.3,0.4})", ["alpha", "beta"])
    blank()
    subsection("Exponential (lambda=0.5)")
    test("ACT_DIST_EXP_PDF", "=ACT_DIST_EXP_PDF(1, 0.5)", "f(1)")
    test("ACT_DIST_EXP_CDF", "=ACT_DIST_EXP_CDF(1, 0.5)", "F(1)")
    test("ACT_DIST_EXP_INV", "=ACT_DIST_EXP_INV(0.5, 0.5)", "Median")
    fit("ACT_DIST_EXP_FIT({1,2,3,4,5})",
        "=ACT_DIST_EXP_FIT({1,2,3,4,5})", ["lambda"])
    blank()
    subsection("GPD (xi=0.5, sigma=1)")
    test("ACT_DIST_GPD_PDF", "=ACT_DIST_GPD_PDF(1, 0.5, 1)", "f(1)")
    test("ACT_DIST_GPD_CDF", "=ACT_DIST_GPD_CDF(1, 0.5, 1)", "F(1)")
    test("ACT_DIST_GPD_INV", "=ACT_DIST_GPD_INV(0.5, 0.5, 1)", "Median")
    fit("ACT_DIST_GPD_FIT({1,2,3,4,5,6,7,8,9,10,12,15})",
        "=ACT_DIST_GPD_FIT({1,2,3,4,5,6,7,8,9,10,12,15})",
        ["xi (shape)", "sigma (scale)"])
    blank()
    subsection("Burr XII (c=2, k=3, lambda=1)")
    test("ACT_DIST_BURR_PDF", "=ACT_DIST_BURR_PDF(1, 2, 3, 1)", "f(1)")
    test("ACT_DIST_BURR_CDF", "=ACT_DIST_BURR_CDF(1, 2, 3, 1)", "F(1)")
    test("ACT_DIST_BURR_INV", "=ACT_DIST_BURR_INV(0.5, 2, 3, 1)", "Median")
    fit("ACT_DIST_BURR_FIT({0.5,1,1.5,2,3})",
        "=ACT_DIST_BURR_FIT({0.5,1,1.5,2,3})", ["c", "k", "lambda"])
    blank()
    subsection("Normal Distribution (mu=0, sigma=1)")
    test("ACT_DIST_NORMAL_PDF", "=ACT_DIST_NORMAL_PDF(0, 0, 1)", "phi(0) = 0.3989")
    test("ACT_DIST_NORMAL_CDF", "=ACT_DIST_NORMAL_CDF(0, 0, 1)", "Phi(0) = 0.5")
    test("ACT_DIST_NORMAL_INV", "=ACT_DIST_NORMAL_INV(0.975, 0, 1)", "97.5th percentile = 1.96")
    blank()
    subsection("Inverse Gaussian (mu=1, lambda=2)")
    test("ACT_DIST_INVGAUSS_PDF", "=ACT_DIST_INVGAUSS_PDF(1, 1, 2)", "f(1)")
    test("ACT_DIST_INVGAUSS_CDF", "=ACT_DIST_INVGAUSS_CDF(1, 1, 2)", "F(1)")
    test("ACT_DIST_INVGAUSS_INV", "=ACT_DIST_INVGAUSS_INV(0.5, 1, 2)", "Median")
    blank()
    subsection("Loglogistic (alpha=2, beta=1)")
    test("ACT_DIST_LOGLOGISTIC_PDF", "=ACT_DIST_LOGLOGISTIC_PDF(1, 2, 1)", "f(1)")
    test("ACT_DIST_LOGLOGISTIC_CDF", "=ACT_DIST_LOGLOGISTIC_CDF(1, 2, 1)", "F(1)")
    test("ACT_DIST_LOGLOGISTIC_INV", "=ACT_DIST_LOGLOGISTIC_INV(0.5, 2, 1)", "Median")
    blank()
    subsection("Lomax / Pareto II (alpha=2, lambda=1)")
    test("ACT_DIST_LOMAX_PDF", "=ACT_DIST_LOMAX_PDF(1, 2, 1)", "f(1)")
    test("ACT_DIST_LOMAX_CDF", "=ACT_DIST_LOMAX_CDF(1, 2, 1)", "F(1) = 0.75")
    test("ACT_DIST_LOMAX_INV", "=ACT_DIST_LOMAX_INV(0.5, 2, 1)", "Median")
    blank()
    subsection("Pareto III (mu=0, sigma=1, gamma=2)")
    test("ACT_DIST_PARETO3_PDF", "=ACT_DIST_PARETO3_PDF(1, 0, 1, 2)", "f(1)")
    test("ACT_DIST_PARETO3_CDF", "=ACT_DIST_PARETO3_CDF(1, 0, 1, 2)", "F(1)")
    test("ACT_DIST_PARETO3_INV", "=ACT_DIST_PARETO3_INV(0.5, 0, 1, 2)", "Median")
    test("ACT_DIST_PARETO3_MEAN", "=ACT_DIST_PARETO3_MEAN(0, 1, 2)", "Mean")
    blank()
    subsection("Pareto IV (mu=0, sigma=1, gamma=2, alpha=3)")
    test("ACT_DIST_PARETO4_PDF", "=ACT_DIST_PARETO4_PDF(1, 0, 1, 2, 3)", "f(1)")
    test("ACT_DIST_PARETO4_CDF", "=ACT_DIST_PARETO4_CDF(1, 0, 1, 2, 3)", "F(1)")
    test("ACT_DIST_PARETO4_INV", "=ACT_DIST_PARETO4_INV(0.5, 0, 1, 2, 3)", "Median")
    blank()
    subsection("Fackler Composite Distributions")
    test("ACT_DIST_LNPARETO_PDF", "=ACT_DIST_LNPARETO_PDF(5, 0, 1, 10)",
         "LN-Pareto PDF at x=5, mu=0, sigma=1, theta=10")
    test("ACT_DIST_LNPARETO_CDF", "=ACT_DIST_LNPARETO_CDF(5, 0, 1, 10)", "LN-Pareto CDF")
    test("ACT_DIST_LNPARETO_ALPHA", "=ACT_DIST_LNPARETO_ALPHA(0, 1, 10)", "Derived tail index alpha")
    test("ACT_DIST_EXPPARETO_PDF", "=ACT_DIST_EXPPARETO_PDF(5, 0.5, 10)",
         "Exp-Pareto PDF at x=5, rate=0.5, theta=10")
    test("ACT_DIST_EXPPARETO_CDF", "=ACT_DIST_EXPPARETO_CDF(5, 0.5, 10)", "Exp-Pareto CDF")
    test("ACT_DIST_EXPPARETO_ALPHA", "=ACT_DIST_EXPPARETO_ALPHA(0.5, 10)", "alpha = theta * rate = 5")
    test("ACT_DIST_POWPARETO_PDF", "=ACT_DIST_POWPARETO_PDF(5, 2, 0.5, 10)",
         "Power-Pareto PDF at x=5, alpha=2, beta=0.5, theta=10")
    test("ACT_DIST_POWPARETO_CDF", "=ACT_DIST_POWPARETO_CDF(5, 2, 0.5, 10)", "Power-Pareto CDF")
    blank()

    # ── LIMITED EXPECTED VALUE ──
    section("LIMITED EXPECTED VALUE (LEV)")
    blank()
    subsection("Poisson LEV (lambda=5, limit=5)")
    test("ACT_DIST_POISSON_LEV", "=ACT_DIST_POISSON_LEV(5, 5)", "E[min(X,5)] for Poisson(5)")
    blank()
    subsection("Negative Binomial LEV (r=5, p=0.3, limit=10)")
    test("ACT_DIST_NEGBIN_LEV", "=ACT_DIST_NEGBIN_LEV(10, 5, 0.3)", "E[min(X,10)] for NegBin")
    blank()
    subsection("Exponential LEV (lambda=0.5, limit=2)")
    test("ACT_DIST_EXP_LEV", "=ACT_DIST_EXP_LEV(2, 0.5)", "E[min(X,2)] = 1.2642")
    blank()
    subsection("Lognormal LEV (mu=0, sigma=1, limit=2)")
    test("ACT_DIST_LOGNORM_LEV", "=ACT_DIST_LOGNORM_LEV(2, 0, 1)", "E[min(X,2)] = 1.1139")
    blank()
    subsection("Gamma LEV (alpha=2, beta=1, limit=2)")
    test("ACT_DIST_GAMMA_LEV", "=ACT_DIST_GAMMA_LEV(2, 2, 1)", "E[min(X,2)] = 1.4587")
    blank()
    subsection("Pareto LEV (alpha=2, xm=1, limit=2)")
    test("ACT_DIST_PARETO_LEV", "=ACT_DIST_PARETO_LEV(2, 2, 1)", "E[min(X,2)] = 1.5")
    blank()
    subsection("Lomax LEV (alpha=2, lambda=1, limit=2)")
    test("ACT_DIST_LOMAX_LEV", "=ACT_DIST_LOMAX_LEV(2, 2, 1)", "E[min(X,2)] = 0.6667")
    blank()
    subsection("GPD LEV (xi=0.5, sigma=1, limit=2)")
    test("ACT_DIST_GPD_LEV", "=ACT_DIST_GPD_LEV(2, 0.5, 1)", "E[min(X,2)] = 1.0")
    blank()
    subsection("Weibull LEV (k=2, lambda=1, limit=1)")
    test("ACT_DIST_WEIBULL_LEV", "=ACT_DIST_WEIBULL_LEV(1, 2, 1)", "E[min(X,1)] = 0.7468")
    blank()
    subsection("Beta LEV (alpha=2, beta=5, limit=0.5)")
    test("ACT_DIST_BETA_LEV", "=ACT_DIST_BETA_LEV(0.5, 2, 5)", "E[min(X,0.5)] = 0.2757")
    blank()
    subsection("Burr XII LEV (c=2, k=3, lambda=1, limit=1)")
    test("ACT_DIST_BURR_LEV", "=ACT_DIST_BURR_LEV(1, 2, 3, 1)", "E[min(X,1)] = 0.5445")
    blank()
    subsection("Inverse Gaussian LEV (mu=1, lambda=2, limit=1)")
    test("ACT_DIST_INVGAUSS_LEV", "=ACT_DIST_INVGAUSS_LEV(1, 1, 2)", "E[min(X,1)]")
    blank()
    subsection("Loglogistic LEV (alpha=2, beta=1, limit=2)")
    test("ACT_DIST_LOGLOGISTIC_LEV", "=ACT_DIST_LOGLOGISTIC_LEV(2, 2, 1)", "E[min(X,2)]")
    blank()

    # ── EXPOSURE CURVES ──
    section("EXPOSURE CURVES")
    blank()
    test("ACT_EXPOSURE_MBBEFD", "=ACT_EXPOSURE_MBBEFD(0.5, 2, 3)", "MBBEFD at d=0.5, b=2, g=3")
    test("ACT_EXPOSURE_SWISSRE", "=ACT_EXPOSURE_SWISSRE(0.5, 1)", "Swiss Re curve 1 (light)")
    test("ACT_EXPOSURE_SWISSRE", "=ACT_EXPOSURE_SWISSRE(0.5, 3)", "Swiss Re curve 3 (medium)")
    test("ACT_EXPOSURE_SWISSRE", "=ACT_EXPOSURE_SWISSRE(0.5, 5)", "Swiss Re curve 5 (heavy)")
    test("ACT_EXPOSURE_LLOYDS", "=ACT_EXPOSURE_LLOYDS(0.5, 1)", "Lloyd's Y1 curve")
    test("ACT_EXPOSURE_LLOYDS", "=ACT_EXPOSURE_LLOYDS(0.5, 2)", "Lloyd's Y2 curve")
    test("ACT_EXPOSURE_LLOYDS", "=ACT_EXPOSURE_LLOYDS(0.5, 3)", "Lloyd's Y3 curve")
    test("ACT_EXPOSURE_LLOYDS", "=ACT_EXPOSURE_LLOYDS(0.5, 4)", "Lloyd's Y4 curve")
    test("ACT_EXPOSURE_POWER", "=ACT_EXPOSURE_POWER(0.5, 2)", "Power curve n=2")
    test("ACT_EXPOSURE_INVERSE_POWER", "=ACT_EXPOSURE_INVERSE_POWER(0.5, 2)", "Inverse power n=2")
    test("ACT_EXPOSURE_PARETO", "=ACT_EXPOSURE_PARETO(0.5, 2)", "Pareto exposure alpha=2")
    test("ACT_EXPOSURE_RIEBESELL", "=ACT_EXPOSURE_RIEBESELL(0.5, 0.8)", "Riebesell c=0.8")
    test("ACT_EXPOSURE_RIEBESELL_INV", "=ACT_EXPOSURE_RIEBESELL_INV(0.7, 0.8)", "Inverse Riebesell")
    blank()

    # ── REINSURANCE ──
    section("REINSURANCE")
    blank()
    test("ACT_XOL_LAYER_LOSS", "=ACT_XOL_LAYER_LOSS(5000000, 1000000, 4000000)",
         "Layer loss: 5M ground-up, 4M xs 1M")
    test("ACT_XOL_EXPECTED_LOSS", "=ACT_XOL_EXPECTED_LOSS(1000000, 4000000, 2, 100000, 0.1)",
         "Expected layer loss")
    test("ACT_ILF_PARETO", "=ACT_ILF_PARETO(2000000, 1000000, 2)", "ILF at 2M, base 1M, alpha=2")
    test("ACT_EXPOSURE_LAYER_RATE", "=ACT_EXPOSURE_LAYER_RATE(0.1, 0.5, 0.05, 2, 3)", "Rate on line")
    blank()

    # ── CAT MODELING ──
    section("CAT MODELING")
    blank()
    elt_rates = "{0.01,0.02,0.05,0.1}"
    elt_losses = "{1000000,500000,200000,50000}"
    array_test("ACT_CAT_ELT_TO_YLT",
               f'=ACT_CAT_ELT_TO_YLT({elt_rates}, {elt_losses}, 100, 42)',
               "ELT to YLT simulation (array)")
    array_test("ACT_CAT_YLT_OEP_CURVE",
               f'=ACT_CAT_YLT_OEP_CURVE({{500000,200000,0,100000,300000}})',
               "OEP curve from annual max losses (array)")
    array_test("ACT_CAT_YLT_AEP_CURVE",
               f'=ACT_CAT_YLT_AEP_CURVE({{500000,200000,0,100000,300000}})',
               "AEP curve from annual aggregate losses (array)")
    array_test("ACT_CAT_OEP_CURVE_RP",
               f'=ACT_CAT_OEP_CURVE_RP({{500000,200000,0,100000,300000}}, {{5,10,25}})',
               "OEP at specific return periods (array)")
    array_test("ACT_CAT_AEP_CURVE_RP",
               f'=ACT_CAT_AEP_CURVE_RP({{500000,200000,0,100000,300000}}, {{5,10,25}})',
               "AEP at specific return periods (array)")
    blank()

    # ── RISK METRICS ──
    section("RISK METRICS")
    blank()
    samples = "{100,200,300,400,500,600,700,800,900,1000}"
    test("ACT_VAR_FROM_SAMPLES",
         f'=ACT_VAR_FROM_SAMPLES({samples}, 0.95)', "95th percentile VaR")
    test("ACT_TVAR_FROM_SAMPLES",
         f'=ACT_TVAR_FROM_SAMPLES({samples}, 0.95)', "95th percentile TVaR")
    blank()

    # ── RETURN PERIOD FUNCTIONS ──
    subsection("Return Period Functions")
    rps = "{100,50,25,10,5}"
    losses = "{1000000,2000000,3000000,5000000,10000000}"
    test("ACT_RETURN_PERIOD_LOSS",
         f'=ACT_RETURN_PERIOD_LOSS({rps}, {losses}, 20, "LOG")', "20-year loss")
    array_test("ACT_RETURN_PERIOD_TABLE",
               f'=ACT_RETURN_PERIOD_TABLE({rps}, {losses}, {{10,25,50,100}})', "RP table (array)")
    test("ACT_AAL_FROM_OEP", f'=ACT_AAL_FROM_OEP({rps}, {losses})', "AAL from OEP curve")
    blank()

    # ── INTERPOLATION ──
    section("INTERPOLATION")
    blank()
    xs = "{1,2,3,4,5}"
    ys = "{10,20,35,55,80}"
    test("ACT_INTERP", f'=ACT_INTERP({xs}, {ys}, 2.5, "LINEAR")', "Linear interpolation")
    test("ACT_INTERP", f'=ACT_INTERP({xs}, {ys}, 0.5, "FLAT")', "Flat extrapolation")
    test("ACT_INTERP", f'=ACT_INTERP({xs}, {ys}, 6, "GRADIENT")', "Gradient extrapolation")
    test("ACT_INTERP_LOG", f'=ACT_INTERP_LOG({xs}, {ys}, 2.5)', "Log-linear interpolation")
    test("ACT_INTERP2D",
         '=ACT_INTERP2D({1,2,3}, {10,20,30}, {100,200,300;110,220,330;120,240,360}, 1.5, 15)',
         "2D bilinear interpolation")
    blank()

    # ── COPULAS ──
    section("COPULAS - Gaussian")
    blank()
    corr = "{1,0.5;0.5,1}"
    array_test("ACT_COPULA_GAUSSIAN", f'=ACT_COPULA_GAUSSIAN({corr}, 5, 42)',
               "Gaussian copula samples (array)")
    array_test("ACT_COPULA_GAUSSIAN_SINGLE", f'=ACT_COPULA_GAUSSIAN_SINGLE({corr}, 42)',
               "Single Gaussian sample")
    blank()
    section("COPULAS - Student-t")
    array_test("ACT_COPULA_STUDENT_T", f'=ACT_COPULA_STUDENT_T({corr}, 5, 5, 42)',
               "Student-t copula (array)")
    array_test("ACT_COPULA_STUDENT_T_SINGLE", f'=ACT_COPULA_STUDENT_T_SINGLE({corr}, 5, 42)',
               "Single Student-t sample")
    blank()
    section("COPULAS - Clayton (lower tail)")
    array_test("ACT_COPULA_CLAYTON", "=ACT_COPULA_CLAYTON(2, 5, 42)", "Clayton copula (array)")
    array_test("ACT_COPULA_CLAYTON_SINGLE", "=ACT_COPULA_CLAYTON_SINGLE(2, 42)",
               "Single Clayton sample")
    test("ACT_COPULA_CLAYTON_CDF", "=ACT_COPULA_CLAYTON_CDF(0.5, 0.5, 2)", "Clayton CDF")
    blank()
    section("COPULAS - Frank (symmetric)")
    array_test("ACT_COPULA_FRANK", "=ACT_COPULA_FRANK(5, 5, 42)", "Frank copula (array)")
    array_test("ACT_COPULA_FRANK_SINGLE", "=ACT_COPULA_FRANK_SINGLE(5, 42)",
               "Single Frank sample")
    test("ACT_COPULA_FRANK_CDF", "=ACT_COPULA_FRANK_CDF(0.5, 0.5, 5)", "Frank CDF")
    blank()
    section("COPULAS - Gumbel (upper tail)")
    array_test("ACT_COPULA_GUMBEL", "=ACT_COPULA_GUMBEL(2, 5, 42)", "Gumbel copula (array)")
    array_test("ACT_COPULA_GUMBEL_SINGLE", "=ACT_COPULA_GUMBEL_SINGLE(2, 42)",
               "Single Gumbel sample")
    test("ACT_COPULA_GUMBEL_CDF", "=ACT_COPULA_GUMBEL_CDF(0.5, 0.5, 2)", "Gumbel CDF")
    blank()
    section("COPULAS - Utilities")
    test("ACT_COPULA_TAU_TO_THETA", '=ACT_COPULA_TAU_TO_THETA(0.5, "CLAYTON")',
         "Kendall tau -> theta")
    test("ACT_COPULA_TAIL_LOWER", '=ACT_COPULA_TAIL_LOWER("CLAYTON", 2)',
         "Lower tail dependence")
    test("ACT_COPULA_TAIL_UPPER", '=ACT_COPULA_TAIL_UPPER("GUMBEL", 2)',
         "Upper tail dependence")
    test("ACT_COPULA_TAIL_LOWER", '=ACT_COPULA_TAIL_LOWER("STUDENT_T", 0.5, 5)',
         "Student-t lower tail")
    test("ACT_COPULA_TAIL_UPPER", '=ACT_COPULA_TAIL_UPPER("STUDENT_T", 0.5, 5)',
         "Student-t upper tail")
    blank()

    # ── CHAIN LADDER ──
    section("CHAIN LADDER - See dedicated sheet")
    blank()
    ws.cell(row=row, column=1,
            value="(Chain ladder functions tested in 'Chain Ladder' sheet with triangle data)")
    row += 1
    cl_funcs = [
        "ACT_CL_FACTORS - Development factors (array)",
        "ACT_CL_LATEST - Latest diagonal (array)",
        "ACT_CL_ULTIMATE - Projected ultimates (array)",
        "ACT_CL_IBNR - IBNR reserves (array)",
        "ACT_BF_ULTIMATE - Bornhuetter-Ferguson (array)",
        "ACT_MACK_FACTOR_SE - Factor std errors (array)",
        "ACT_MACK_RESERVE_SE - Reserve std errors (array)",
        "ACT_CL_BOOTSTRAP - Bootstrap total (array)",
        "ACT_CL_BOOTSTRAP_ORIGIN - Bootstrap by AY (array)",
        "ACT_CAPECOD_ULTIMATE - Cape Cod (array)",
        "ACT_CAPECOD_ELR - Cape Cod ELR",
        "ACT_TRIANGLE_TO_INCREMENTAL - Cum->Inc (array)",
        "ACT_INCREMENTAL_TO_CUMULATIVE - Inc->Cum (array)",
        "ACT_TRIANGLE_DIAGONAL - Extract diagonal (array)",
        "ACT_TRIANGLE_LINK_RATIOS - Link ratios (array)",
        "ACT_CL_CALENDAR_ADJUST - Calendar adjust (array)",
        "ACT_CL_CALENDAR_TOTALS - Calendar totals (array)",
        "ACT_CL_WEIGHTED_AVERAGE - Weighted avg (array)",
    ]
    for func in cl_funcs:
        ws.cell(row=row, column=1, value=func)
        row += 1
    blank()

    # ── VERSION INFO ──
    section("VERSION INFO")
    blank()
    test("ACT_VERSION", "=ACT_VERSION()", "Current version")
    test("ACT_BUILD_DATE", "=ACT_BUILD_DATE()", "Build date")
    test("ACT_GITHUB_URL", "=ACT_GITHUB_URL()", "GitHub URL")
    test("ACT_COMMIT_COUNT", "=ACT_COMMIT_COUNT()", "Number of commits")
    test("ACT_COMMIT_INFO", '=ACT_COMMIT_INFO(1, "hash")', "Latest commit hash")
    test("ACT_COMMIT_INFO", '=ACT_COMMIT_INFO(1, "date")', "Latest commit date")
    test("ACT_COMMIT_INFO", '=ACT_COMMIT_INFO(1, "message")', "Latest commit message")
    array_test("ACT_COMMIT_HISTORY", "=ACT_COMMIT_HISTORY()", "Full history (array)")
    blank()

    # ── AGGREGATE CLAIMS (Panjer Recursion) ──
    section("Aggregate Claims (Panjer Recursion)")
    test("ACT_DISCRETIZE_EXPONENTIAL(1, 0.5, 10)", "=ACT_DISCRETIZE_EXPONENTIAL(1, 0.5, 10)")
    test("ACT_DISCRETIZE_GAMMA(2, 1, 0.5, 10)", "=ACT_DISCRETIZE_GAMMA(2, 1, 0.5, 10)")
    test("ACT_DISCRETIZE_LOGNORMAL(0, 1, 0.5, 10)", "=ACT_DISCRETIZE_LOGNORMAL(0, 1, 0.5, 10)")
    blank()
    test("ACT_PANJER_POISSON(2, {0.5,0.5}, 5)", "=ACT_PANJER_POISSON(2, {0.5,0.5}, 5)")
    test("ACT_PANJER_NEGBIN(2, 0.5, {0.5,0.5}, 5)", "=ACT_PANJER_NEGBIN(2, 0.5, {0.5,0.5}, 5)")
    test("ACT_PANJER_BINOMIAL(5, 0.3, {0.5,0.5}, 5)", "=ACT_PANJER_BINOMIAL(5, 0.3, {0.5,0.5}, 5)")
    blank()
    pmf_expr = "ACT_PANJER_POISSON(2, {0.5,0.5}, 5)"
    test("ACT_AGGREGATE_MEAN(pmf, 1)", f'=ACT_AGGREGATE_MEAN({pmf_expr}, 1)')
    test("ACT_AGGREGATE_STDEV(pmf, 1)", f'=ACT_AGGREGATE_STDEV({pmf_expr}, 1)')
    test("ACT_AGGREGATE_VAR_STAT(pmf, 1)", f'=ACT_AGGREGATE_VAR_STAT({pmf_expr}, 1)')
    test("ACT_AGGREGATE_CDF(2, pmf, 1)", f'=ACT_AGGREGATE_CDF(2, {pmf_expr}, 1)')
    test("ACT_AGGREGATE_VAR(0.95, pmf, 1)", f'=ACT_AGGREGATE_VAR(0.95, {pmf_expr}, 1)')
    test("ACT_AGGREGATE_TVAR(0.95, pmf, 1)", f'=ACT_AGGREGATE_TVAR(0.95, {pmf_expr}, 1)')

    return ws


def create_return_period_sheet(wb):
    """Create the Return Period examples sheet."""
    ws = wb.create_sheet("Return Periods")
    set_column_widths(ws, {'A': 18, 'B': 18, 'C': 20, 'D': 30, 'E': 5})

    row = add_title(ws, "Return Period Functions")
    row = add_note(ws, "Generate losses from return period tables (OEP/AEP curves)", row)
    row += 1

    # Sample EP curve
    row = add_note(ws, "SAMPLE EXCEEDANCE PROBABILITY CURVE", row)
    row = add_table_header(ws, ["Return Period", "OEP Loss"], row)
    ep_data = [(10, 100000), (25, 250000), (50, 500000), (100, 1000000), (250, 2500000)]
    data_start = row
    for rp, loss in ep_data:
        ws.cell(row=row, column=1, value=rp).border = THIN_BORDER
        cell_loss = ws.cell(row=row, column=2, value=loss)
        cell_loss.border = THIN_BORDER
        cell_loss.number_format = '#,##0'
        row += 1
    data_end = row - 1
    row += 1

    rp_range = f"A{data_start}:A{data_end}"
    loss_range = f"B{data_start}:B{data_end}"

    # EP Curve chart with axes
    chart = ScatterChart()
    chart.title = "Exceedance Probability Curve"
    chart.width = 12
    chart.height = 8
    chart.roundedCorners = False
    # Removed: use Excel defaults for chart corners
    chart.x_axis.title = "Return Period (years)"
    chart.y_axis.title = "OEP Loss"
    chart.y_axis.scaling.min = 0
    chart.x_axis.scaling.logBase = 10
    chart.x_axis.tickLblPos = "low"
    chart.y_axis.tickLblPos = "low"
    chart.x_axis.delete = False
    chart.y_axis.delete = False

    from openpyxl.chart import Series
    xvalues = Reference(ws, min_col=1, min_row=data_start, max_row=data_end)
    yvalues = Reference(ws, min_col=2, min_row=data_start, max_row=data_end)
    series = Series(yvalues, xvalues, title_from_data=False)
    series.marker = Marker(symbol='none')
    series.graphicalProperties.line.solidFill = CHART_COLORS[0]
    chart.series.append(series)
    chart.legend = None
    ws.add_chart(chart, "E4")

    # Interpolation examples
    row = add_note(ws, "LOSS INTERPOLATION", row)
    row = add_table_header(ws, ["Target RP", "Loss (LOG)", "Loss (LINEAR)", "Notes"], row)
    targets = [20, 75, 150, 200]
    for target in targets:
        log_formula = f'=ACT_RETURN_PERIOD_LOSS({rp_range}, {loss_range}, A{row}, "LOG")'
        lin_formula = f'=ACT_RETURN_PERIOD_LOSS({rp_range}, {loss_range}, A{row}, "LINEAR")'
        ws.cell(row=row, column=1, value=target).border = THIN_BORDER
        cell_log = ws.cell(row=row, column=2, value=log_formula)
        cell_log.border = THIN_BORDER
        cell_log.number_format = '#,##0'
        cell_lin = ws.cell(row=row, column=3, value=lin_formula)
        cell_lin.border = THIN_BORDER
        cell_lin.number_format = '#,##0'
        ws.cell(row=row, column=4, value="Log interpolation typical for cat curves").border = THIN_BORDER
        row += 1

    row += 1

    # AAL calculation
    row = add_note(ws, "AVERAGE ANNUAL LOSS (AAL) FROM OEP CURVE", row)
    ws.cell(row=row, column=1, value="AAL:").font = HEADER_FONT
    cell_aal = ws.cell(row=row, column=2, value=f"=ACT_AAL_FROM_OEP({rp_range}, {loss_range})")
    cell_aal.number_format = '#,##0'

    return ws


def main():
    """Main function to populate the workbook.

    Creates the entire workbook from scratch to avoid openpyxl drawing
    relationship corruption that occurs when loading existing files with
    many charts/drawings and adding new ones.
    """
    print(f"Creating workbook: {EXCEL_PATH}")

    wb = openpyxl.Workbook()
    # Remove the default "Sheet" created by Workbook()
    del wb["Sheet"]

    # Create All Functions Test tab (previously preserved from existing file)
    print("Creating All Functions Test tab...")
    create_all_functions_test_sheet(wb)

    # Create example sheets
    print("Creating Versions sheet...")
    create_versions_sheet(wb)

    print("Creating Distributions sheet with ALL distributions and charts...")
    create_distributions_sheet(wb)

    print("Creating Exposure Curves sheet with charts...")
    create_exposure_curves_sheet(wb)

    print("Creating Reinsurance sheet...")
    create_reinsurance_sheet(wb)

    print("Creating Cat Modeling sheet...")
    create_cat_modeling_sheet(wb)

    print("Creating Interpolation sheet with chart...")
    create_interpolation_sheet(wb)

    print("Creating Chain Ladder sheet with bar charts...")
    create_chainladder_sheet(wb)

    print("Creating Copulas sheet with scatter plot...")
    create_copulas_sheet(wb)

    print("Creating Aggregate Claims sheet with Panjer recursion...")
    create_aggregate_claims_sheet(wb)

    print("Creating Return Periods sheet with EP curve...")
    create_return_period_sheet(wb)

    # Ensure Versions is the first sheet
    if "Versions" in wb.sheetnames:
        wb.move_sheet("Versions", offset=-wb.sheetnames.index("Versions"))

    # Save
    print(f"Saving workbook...")
    wb.save(EXCEL_PATH)
    print("Done!")



if __name__ == "__main__":
    main()
