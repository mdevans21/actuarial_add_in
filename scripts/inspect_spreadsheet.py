#!/usr/bin/env python3
"""
Inspect the spreadsheet to find formulas and identify issues.
"""

from pathlib import Path
from openpyxl import load_workbook
import re

def main():
    project_root = Path(__file__).parent.parent
    source_file = project_root / "excel" / "actuarial_add_in_v0.0.xlsm"

    wb = load_workbook(source_file, data_only=False, keep_vba=True)

    print("Sheets:", wb.sheetnames)
    print()

    # Find all formulas with ACT_ prefix
    act_functions = set()
    at_formulas = []
    all_formulas = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    all_formulas.append((sheet_name, cell.coordinate, formula))

                    # Find ACT_ functions
                    matches = re.findall(r'ACT_[A-Z_0-9]+', formula)
                    for match in matches:
                        act_functions.add(match)

                    # Find @-prefixed formulas
                    if '@' in formula:
                        at_formulas.append((sheet_name, cell.coordinate, formula))

    print("=" * 60)
    print("ACT_ Functions Used in Spreadsheet:")
    print("=" * 60)
    for func in sorted(act_functions):
        print(f"  {func}")

    print()
    print("=" * 60)
    print("Formulas with @ prefix:")
    print("=" * 60)
    for sheet, coord, formula in at_formulas:
        print(f"  [{sheet}] {coord}: {formula[:80]}...")

    print()
    print(f"Total formulas: {len(all_formulas)}")
    print(f"Total ACT_ functions used: {len(act_functions)}")
    print(f"Formulas with @: {len(at_formulas)}")

if __name__ == "__main__":
    main()
