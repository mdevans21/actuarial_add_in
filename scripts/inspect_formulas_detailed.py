#!/usr/bin/env python3
"""
Detailed inspection of formulas in the spreadsheet.
"""

from pathlib import Path
from openpyxl import load_workbook
import re

def main():
    project_root = Path(__file__).parent.parent
    source_file = project_root / "excel" / "actuarial_add_in_v0.0.xlsm"

    wb = load_workbook(source_file, data_only=False, keep_vba=True)

    # Look for ACT_CL_LATEST and ACT_CL_BOOTSTRAP_ORIGIN usage
    target_functions = ['ACT_CL_LATEST', 'ACT_CL_BOOTSTRAP_ORIGIN']

    print("Searching for specific functions...")
    print("=" * 60)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n[{sheet_name}]")
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    val = str(cell.value)
                    for func in target_functions:
                        if func in val:
                            print(f"  {cell.coordinate}: {val}")

    print("\n" + "=" * 60)
    print("All formulas in Chain Ladder sheet:")
    print("=" * 60)

    ws = wb['Chain Ladder']
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                print(f"  {cell.coordinate}: {cell.value}")

if __name__ == "__main__":
    main()
