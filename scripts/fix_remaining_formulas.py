#!/usr/bin/env python3
"""
Fix remaining formula errors in the All Functions Test sheet.
"""

import openpyxl
import os

EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', 'excel', 'actuarial_add_in.xlsm')

# Formula fixes based on actual function signatures
FORMULA_FIXES = {
    # ACT_CREDIBILITY_K - the inline array syntax might not work well with Excel-DNA
    # Try a simpler form or note that cell references work better
    'ACT_CREDIBILITY_K': {
        'new': '=ACT_CREDIBILITY_K({0.05,0.06,0.04,0.055}, {100,150,120,110})',
        'note': 'Estimate k from data'
    },

    # ACT_BUHLMANN_STRAUB_PARAMS expects 1D arrays, not 2D
    # Current formula uses 2D arrays with semicolons
    'ACT_BUHLMANN_STRAUB_PARAMS': {
        'new': '=ACT_BUHLMANN_STRAUB_PARAMS({0.05,0.06,0.04,0.055}, {100,150,120,110})',
        'note': 'B-S parameters (array)'
    },

    # ACT_BURNING_COST expects 1D arrays
    'ACT_BURNING_COST': {
        'new': '=ACT_BURNING_COST({50000,60000,40000}, {1000000,1100000,1050000})',
        'note': 'Burning cost'
    },


    # ACT_DEDUCTIBLE_CREDIT expects (ler, expenseAdjust=1.0) - only 2 params
    # Current: (10000, 2, 1000) - wrong!
    # LER should be 0-1, use result from ACT_LOSS_ELIMINATION_RATIO
    'ACT_DEDUCTIBLE_CREDIT': {
        'new': '=ACT_DEDUCTIBLE_CREDIT(0.95, 0.9)',
        'note': 'Deductible credit'
    },

    # ACT_PREMIUM_WITH_DEDUCTIBLE expects (groundUpPremium, ler, lossRatioPortion) - 3 params
    # Current: (100000, 10000, 2, 1000) - 4 params, wrong!
    'ACT_PREMIUM_WITH_DEDUCTIBLE': {
        'new': '=ACT_PREMIUM_WITH_DEDUCTIBLE(100000, 0.95, 0.65)',
        'note': 'Premium with deductible'
    },

    # ACT_SHOCK_LOSS_LOAD expects (annualFrequency, shockSeverity, riskMultiplier=1.0)
    # Current: (0.01, 10000000, 100000) - riskMultiplier=100000 is way too high!
    'ACT_SHOCK_LOSS_LOAD': {
        'new': '=ACT_SHOCK_LOSS_LOAD(0.01, 10000000, 1.5)',
        'note': 'Shock loss loading'
    },

    # ACT_FREQUENCY_ESTIMATE expects (observedClaims, exposure, priorFrequency, priorVariance) - 4 params
    # Current: (10, 1000, 5, 2, 0.005) - 5 params!
    'ACT_FREQUENCY_ESTIMATE': {
        'new': '=ACT_FREQUENCY_ESTIMATE(10, 1000, 0.01, 0.0001)',
        'note': 'Credibility frequency est'
    },

    # ACT_EXPOSURE_LAYER_RATE expects (attachmentPct, exhaustionPct, burningCost, b, g) - 5 params
    # Current: (100000, 1000000, 4000000) - 3 params, and they're dollar amounts not proportions!
    'ACT_EXPOSURE_LAYER_RATE': {
        'new': '=ACT_EXPOSURE_LAYER_RATE(0.1, 0.5, 0.05, 2, 3)',
        'note': 'Rate on line'
    },

    # ACT_DIST_GPD_FIT - use typical exceedance data (positive values over threshold)
    'ACT_DIST_GPD_FIT': {
        'new': '=ACT_DIST_GPD_FIT({1,2,3,4,5,6,7,8,9,10,12,15})',
        'note': 'PWM ξ,σ'
    },

    # ACT_ILF_LAYER - now fixed with proper Pareto II LEV formula
    # Parameters: attachment, layerLimit, baseLimit, alpha, scale(optional, default 1000)
    'ACT_ILF_LAYER': {
        'new': '=ACT_ILF_LAYER(100000, 400000, 100000, 2, 50000)',
        'note': 'Layer ILF'
    },
}


def fix_formulas():
    print(f"Opening workbook: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, keep_vba=True)

    if "All Functions Test" not in wb.sheetnames:
        print("Error: 'All Functions Test' sheet not found")
        return

    ws = wb["All Functions Test"]
    fixes_applied = 0

    for row_num in range(1, 250):
        cell_a = ws.cell(row=row_num, column=1).value
        cell_b = ws.cell(row=row_num, column=2).value

        if cell_a and cell_a in FORMULA_FIXES:
            fix = FORMULA_FIXES[cell_a]
            print(f"Row {row_num}: Fixing {cell_a}")
            print(f"  Old: {cell_b}")
            print(f"  New: {fix['new']}")
            ws.cell(row=row_num, column=2).value = fix['new']
            # Also update the note in column C if it exists
            if 'note' in fix:
                ws.cell(row=row_num, column=3).value = fix['note']
            fixes_applied += 1

    print(f"\nApplied {fixes_applied} fixes")

    # Save
    print(f"Saving workbook...")
    wb.save(EXCEL_PATH)
    print("Done!")


if __name__ == "__main__":
    fix_formulas()
