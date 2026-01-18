#!/usr/bin/env python3
"""
Create comprehensive spreadsheet with ALL functions tested.
Fixes @ symbol issues for array-returning functions.
"""

import shutil
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Styles
TITLE_FONT = Font(bold=True, size=14)
HEADER_FONT = Font(bold=True, size=12)
BOLD_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Functions that return arrays (should NOT have @ prefix)
ARRAY_FUNCTIONS = {
    'ACT_CL_FACTORS', 'ACT_CL_LATEST', 'ACT_CL_ULTIMATE', 'ACT_CL_IBNR',
    'ACT_BF_ULTIMATE', 'ACT_MACK_FACTOR_SE', 'ACT_MACK_RESERVE_SE',
    'ACT_BOOTSTRAP_CL', 'ACT_BOOTSTRAP_CL_ORIGIN',
    'ACT_CAPECOD_ULTIMATE', 'ACT_CAPECOD_ELR',
    'ACT_TRIANGLE_TO_INCREMENTAL', 'ACT_INCREMENTAL_TO_CUMULATIVE',
    'ACT_TRIANGLE_DIAGONAL', 'ACT_TRIANGLE_LINK_RATIOS',
    'ACT_CL_CALENDAR_ADJUST', 'ACT_CL_CALENDAR_TOTALS',
    'ACT_CL_WEIGHTED_AVERAGE', 'ACT_BERQUIST_SHERMAN',
    'ACT_COPULA_GAUSSIAN', 'ACT_COPULA_STUDENT_T',
    'ACT_COPULA_CLAYTON', 'ACT_COPULA_FRANK', 'ACT_COPULA_GUMBEL',
    'ACT_RETURN_PERIOD_TABLE', 'ACT_COMMIT_HISTORY',
    'ACT_ILF_TABLE', 'ACT_ILF_TABLE_STANDARD',
    'ACT_BUHLMANN_STRAUB_PARAMS', 'ACT_RETRO_PARAMETERS',
    'ACT_BAYESIAN_UPDATE_FULL',
}


def fix_formulas_in_sheet(ws):
    """Fix @ prefix in array formulas."""
    fixed = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                original = cell.value
                # Check if formula contains @ before an array function
                for func in ARRAY_FUNCTIONS:
                    if f'@{func}' in cell.value:
                        cell.value = cell.value.replace(f'@{func}', func)
                        fixed += 1
                # Also fix =@ at start
                if cell.value.startswith('=@'):
                    cell.value = '=' + cell.value[2:]
                    if cell.value != original:
                        fixed += 1
    return fixed


def add_section_header(ws, row, title):
    """Add a section header."""
    ws.cell(row=row, column=1, value=title)
    ws.cell(row=row, column=1).font = HEADER_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    return row + 1


def add_function_test(ws, row, func_name, formula, description=""):
    """Add a function test row."""
    ws.cell(row=row, column=1, value=func_name).font = BOLD_FONT
    ws.cell(row=row, column=2, value=formula)
    if description:
        ws.cell(row=row, column=3, value=description)
    return row + 1


def create_distributions_tests(ws, start_row):
    """Add distribution function tests."""
    row = start_row
    row = add_section_header(ws, row, "DISTRIBUTION FUNCTIONS")
    row += 1

    # Poisson
    row = add_section_header(ws, row, "Poisson (λ=5)")
    row = add_function_test(ws, row, "ACT_POISSON_PDF", "=ACT_POISSON_PDF(3, 5)", "P(X=3)")
    row = add_function_test(ws, row, "ACT_POISSON_CDF", "=ACT_POISSON_CDF(3, 5)", "P(X≤3)")
    row = add_function_test(ws, row, "ACT_POISSON_INV", "=ACT_POISSON_INV(0.5, 5)", "Median")
    row = add_function_test(ws, row, "ACT_POISSON_FIT", "=ACT_POISSON_FIT({1,2,3,4,5,6,7})", "MLE λ")
    row += 1

    # Negative Binomial
    row = add_section_header(ws, row, "Negative Binomial (r=5, p=0.3)")
    row = add_function_test(ws, row, "ACT_NEGBIN_PDF", "=ACT_NEGBIN_PDF(5, 5, 0.3)", "P(X=5)")
    row = add_function_test(ws, row, "ACT_NEGBIN_CDF", "=ACT_NEGBIN_CDF(5, 5, 0.3)", "P(X≤5)")
    row = add_function_test(ws, row, "ACT_NEGBIN_INV", "=ACT_NEGBIN_INV(0.5, 5, 0.3)", "Median")
    row = add_function_test(ws, row, "ACT_NEGBIN_FIT", "=ACT_NEGBIN_FIT({1,2,3,4,5,6,7})", "MLE r,p")
    row += 1

    # Lognormal
    row = add_section_header(ws, row, "Lognormal (μ=0, σ=1)")
    row = add_function_test(ws, row, "ACT_LOGNORM_PDF", "=ACT_LOGNORM_PDF(1, 0, 1)", "f(1)")
    row = add_function_test(ws, row, "ACT_LOGNORM_CDF", "=ACT_LOGNORM_CDF(1, 0, 1)", "F(1)")
    row = add_function_test(ws, row, "ACT_LOGNORM_INV", "=ACT_LOGNORM_INV(0.5, 0, 1)", "Median")
    row = add_function_test(ws, row, "ACT_LOGNORM_FIT", "=ACT_LOGNORM_FIT({1,2,3,4,5})", "MLE μ,σ")
    row += 1

    # Gamma
    row = add_section_header(ws, row, "Gamma (α=2, β=1)")
    row = add_function_test(ws, row, "ACT_GAMMA_PDF", "=ACT_GAMMA_PDF(1, 2, 1)", "f(1)")
    row = add_function_test(ws, row, "ACT_GAMMA_CDF", "=ACT_GAMMA_CDF(1, 2, 1)", "F(1)")
    row = add_function_test(ws, row, "ACT_GAMMA_INV", "=ACT_GAMMA_INV(0.5, 2, 1)", "Median")
    row = add_function_test(ws, row, "ACT_GAMMA_FIT", "=ACT_GAMMA_FIT({1,2,3,4,5})", "MLE α,β")
    row += 1

    # Pareto
    row = add_section_header(ws, row, "Pareto (α=2, xm=1)")
    row = add_function_test(ws, row, "ACT_PARETO_PDF", "=ACT_PARETO_PDF(2, 2, 1)", "f(2)")
    row = add_function_test(ws, row, "ACT_PARETO_CDF", "=ACT_PARETO_CDF(2, 2, 1)", "F(2)")
    row = add_function_test(ws, row, "ACT_PARETO_INV", "=ACT_PARETO_INV(0.5, 2, 1)", "Median")
    row = add_function_test(ws, row, "ACT_PARETO_FIT", "=ACT_PARETO_FIT({1,2,3,4,5}, 1)", "MLE α")
    row += 1

    # Weibull
    row = add_section_header(ws, row, "Weibull (k=2, λ=1)")
    row = add_function_test(ws, row, "ACT_WEIBULL_PDF", "=ACT_WEIBULL_PDF(1, 2, 1)", "f(1)")
    row = add_function_test(ws, row, "ACT_WEIBULL_CDF", "=ACT_WEIBULL_CDF(1, 2, 1)", "F(1)")
    row = add_function_test(ws, row, "ACT_WEIBULL_INV", "=ACT_WEIBULL_INV(0.5, 2, 1)", "Median")
    row = add_function_test(ws, row, "ACT_WEIBULL_FIT", "=ACT_WEIBULL_FIT({1,2,3,4,5})", "MLE k,λ")
    row += 1

    # Beta
    row = add_section_header(ws, row, "Beta (α=2, β=5)")
    row = add_function_test(ws, row, "ACT_BETA_PDF", "=ACT_BETA_PDF(0.3, 2, 5)", "f(0.3)")
    row = add_function_test(ws, row, "ACT_BETA_CDF", "=ACT_BETA_CDF(0.3, 2, 5)", "F(0.3)")
    row = add_function_test(ws, row, "ACT_BETA_INV", "=ACT_BETA_INV(0.5, 2, 5)", "Median")
    row = add_function_test(ws, row, "ACT_BETA_FIT", "=ACT_BETA_FIT({0.1,0.2,0.3,0.4,0.5})", "MLE α,β")
    row += 1

    # Exponential
    row = add_section_header(ws, row, "Exponential (λ=0.5)")
    row = add_function_test(ws, row, "ACT_EXP_PDF", "=ACT_EXP_PDF(1, 0.5)", "f(1)")
    row = add_function_test(ws, row, "ACT_EXP_CDF", "=ACT_EXP_CDF(1, 0.5)", "F(1)")
    row = add_function_test(ws, row, "ACT_EXP_INV", "=ACT_EXP_INV(0.5, 0.5)", "Median")
    row = add_function_test(ws, row, "ACT_EXP_FIT", "=ACT_EXP_FIT({1,2,3,4,5})", "MLE λ")
    row += 1

    # GPD
    row = add_section_header(ws, row, "GPD (ξ=0.5, σ=1)")
    row = add_function_test(ws, row, "ACT_GPD_PDF", "=ACT_GPD_PDF(1, 0.5, 1)", "f(1)")
    row = add_function_test(ws, row, "ACT_GPD_CDF", "=ACT_GPD_CDF(1, 0.5, 1)", "F(1)")
    row = add_function_test(ws, row, "ACT_GPD_INV", "=ACT_GPD_INV(0.5, 0.5, 1)", "Median")
    row = add_function_test(ws, row, "ACT_GPD_FIT", "=ACT_GPD_FIT({1,2,3,4,5})", "MLE ξ,σ")
    row += 1

    # Burr
    row = add_section_header(ws, row, "Burr XII (c=2, k=3, λ=1)")
    row = add_function_test(ws, row, "ACT_BURR_PDF", "=ACT_BURR_PDF(1, 2, 3, 1)", "f(1)")
    row = add_function_test(ws, row, "ACT_BURR_CDF", "=ACT_BURR_CDF(1, 2, 3, 1)", "F(1)")
    row = add_function_test(ws, row, "ACT_BURR_INV", "=ACT_BURR_INV(0.5, 2, 3, 1)", "Median")
    row = add_function_test(ws, row, "ACT_BURR_FIT", "=ACT_BURR_FIT({1,2,3,4,5})", "MLE c,k,λ")
    row += 1

    return row


def create_exposure_curves_tests(ws, start_row):
    """Add exposure curve tests."""
    row = start_row
    row = add_section_header(ws, row, "EXPOSURE CURVES")
    row += 1

    row = add_function_test(ws, row, "ACT_MBBEFD", "=ACT_MBBEFD(0.5, 2, 3)", "MBBEFD at d=0.5")
    row = add_function_test(ws, row, "ACT_SWISSRE_CURVE", "=ACT_SWISSRE_CURVE(0.5, 3)", "Swiss Re curve 3")
    row = add_function_test(ws, row, "ACT_LLOYDS_CURVE", "=ACT_LLOYDS_CURVE(0.5, 2)", "Lloyd's Y2 curve")
    row = add_function_test(ws, row, "ACT_POWER_CURVE", "=ACT_POWER_CURVE(0.5, 2)", "Power curve n=2")
    row = add_function_test(ws, row, "ACT_INVERSE_POWER_CURVE", "=ACT_INVERSE_POWER_CURVE(0.5, 2)", "Inverse power")
    row = add_function_test(ws, row, "ACT_PARETO_EXPOSURE", "=ACT_PARETO_EXPOSURE(0.5, 2)", "Pareto exposure α=2")
    row = add_function_test(ws, row, "ACT_RIEBESELL_CURVE", "=ACT_RIEBESELL_CURVE(0.5, 0.8)", "Riebesell c=0.8")
    row = add_function_test(ws, row, "ACT_RIEBESELL_CURVE_INV", "=ACT_RIEBESELL_CURVE_INV(0.5, 0.8)", "Riebesell inverse")
    row += 1

    return row


def create_reinsurance_tests(ws, start_row):
    """Add reinsurance tests."""
    row = start_row
    row = add_section_header(ws, row, "REINSURANCE FUNCTIONS")
    row += 1

    row = add_function_test(ws, row, "ACT_XOL_LAYER_LOSS", "=ACT_XOL_LAYER_LOSS(5000000, 1000000, 4000000)", "XOL layer loss")
    row = add_function_test(ws, row, "ACT_XOL_EXPECTED_LOSS", "=ACT_XOL_EXPECTED_LOSS(1000000, 4000000, 2, 1000000, 0.01)", "Expected layer loss")
    row = add_function_test(ws, row, "ACT_QS_CEDED", "=ACT_QS_CEDED(1000000, 0.5)", "Quota share 50%")
    row = add_function_test(ws, row, "ACT_AGGREGATE_LAYER", "=ACT_AGGREGATE_LAYER(15000000, 2000000, 10000000)", "Aggregate layer")
    row = add_function_test(ws, row, "ACT_ILF_PARETO", "=ACT_ILF_PARETO(2000000, 1000000, 2)", "ILF Pareto")
    row = add_function_test(ws, row, "ACT_LAYER_RATE_ON_LINE", "=ACT_LAYER_RATE_ON_LINE(100000, 1000000, 4000000)", "Rate on line")
    row += 1

    row = add_section_header(ws, row, "Return Period Functions")
    row = add_function_test(ws, row, "ACT_RETURN_PERIOD_LOSS", "=ACT_RETURN_PERIOD_LOSS({100,50,25,10,5}, {1000000,2000000,3000000,5000000,10000000}, 20, \"LOG\")", "RP=20 loss")
    ws.cell(row=row, column=1, value="ACT_RETURN_PERIOD_TABLE").font = BOLD_FONT
    ws.cell(row=row, column=2, value="=ACT_RETURN_PERIOD_TABLE({100,50,25,10,5}, {1000000,2000000,3000000,5000000,10000000}, {5,10,25,50,100})")
    row += 1
    row = add_function_test(ws, row, "ACT_AAL_FROM_OEP", "=ACT_AAL_FROM_OEP({100,50,25,10,5}, {1000000,2000000,3000000,5000000,10000000})", "AAL from OEP")
    row += 1

    return row


def create_interpolation_tests(ws, start_row):
    """Add interpolation tests."""
    row = start_row
    row = add_section_header(ws, row, "INTERPOLATION FUNCTIONS")
    row += 1

    row = add_function_test(ws, row, "ACT_INTERP (linear)", "=ACT_INTERP({1,2,3,4,5}, {10,20,35,55,80}, 2.5, \"LINEAR\")", "Linear interp")
    row = add_function_test(ws, row, "ACT_INTERP (flat)", "=ACT_INTERP({1,2,3,4,5}, {10,20,35,55,80}, 0.5, \"FLAT\")", "Flat extrap")
    row = add_function_test(ws, row, "ACT_INTERP (gradient)", "=ACT_INTERP({1,2,3,4,5}, {10,20,35,55,80}, 6, \"GRADIENT\")", "Gradient extrap")
    row = add_function_test(ws, row, "ACT_INTERP_LOG", "=ACT_INTERP_LOG({1,2,3,4,5}, {10,20,35,55,80}, 2.5)", "Log-linear interp")
    row = add_function_test(ws, row, "ACT_INTERP2D", "=ACT_INTERP2D({1,2,3}, {10,20,30}, {100,200,300;110,220,330;120,240,360}, 1.5, 15)", "2D interp")
    row += 1

    return row


def create_credibility_tests(ws, start_row):
    """Add credibility tests."""
    row = start_row
    row = add_section_header(ws, row, "CREDIBILITY FUNCTIONS")
    row += 1

    # Bühlmann
    row = add_section_header(ws, row, "Bühlmann Credibility")
    row = add_function_test(ws, row, "ACT_CREDIBILITY_BUHLMANN", "=ACT_CREDIBILITY_BUHLMANN(10, 5)", "Z = n/(n+k)")
    row = add_function_test(ws, row, "ACT_CREDIBILITY_ESTIMATE", "=ACT_CREDIBILITY_ESTIMATE(10, 5, 0.05, 0.03)", "Blended estimate")
    row = add_function_test(ws, row, "ACT_CREDIBILITY_K", "=ACT_CREDIBILITY_K({0.05,0.06,0.04}, {100,150,120})", "Estimate k")
    row += 1

    # Bühlmann-Straub
    row = add_section_header(ws, row, "Bühlmann-Straub")
    row = add_function_test(ws, row, "ACT_CREDIBILITY_BUHLMANN_STRAUB", "=ACT_CREDIBILITY_BUHLMANN_STRAUB(1000, 500)", "Z with weights")
    ws.cell(row=row, column=1, value="ACT_BUHLMANN_STRAUB_PARAMS").font = BOLD_FONT
    ws.cell(row=row, column=2, value="=ACT_BUHLMANN_STRAUB_PARAMS({0.05,0.06,0.04;0.055,0.058,0.045}, {100,150,120;110,140,130})")
    row += 2

    # Experience Rating
    row = add_section_header(ws, row, "Experience Rating")
    row = add_function_test(ws, row, "ACT_EXPERIENCE_MOD", "=ACT_EXPERIENCE_MOD(0.7, 50000, 70000)", "Exp mod factor")
    row = add_function_test(ws, row, "ACT_EXPERIENCE_MOD_SPLIT", "=ACT_EXPERIENCE_MOD_SPLIT(0.7, 50000, 70000, 25000, 0.3)", "Split exp mod")
    row += 1

    # Full/Partial Credibility
    row = add_section_header(ws, row, "Full/Partial Credibility")
    row = add_function_test(ws, row, "ACT_FULL_CREDIBILITY_STANDARD", "=ACT_FULL_CREDIBILITY_STANDARD(0.05, 0.90, 1.5)", "n₀ for full cred")
    row = add_function_test(ws, row, "ACT_PARTIAL_CREDIBILITY", "=ACT_PARTIAL_CREDIBILITY(500, 1082)", "Z = sqrt(n/n₀)")
    row = add_function_test(ws, row, "ACT_LINEAR_CREDIBILITY", "=ACT_LINEAR_CREDIBILITY(500, 1082)", "Z = n/n₀")
    row = add_function_test(ws, row, "ACT_ASYMPTOTIC_CREDIBILITY", "=ACT_ASYMPTOTIC_CREDIBILITY(10, 5, 0.5)", "Z = n^p/(n^p+k)")
    row += 1

    # ILF
    row = add_section_header(ws, row, "ILF Functions")
    row = add_function_test(ws, row, "ACT_ILF_LAYER", "=ACT_ILF_LAYER(1000000, 4000000, 100000, 2)", "Layer ILF")
    ws.cell(row=row, column=1, value="ACT_ILF_TABLE").font = BOLD_FONT
    ws.cell(row=row, column=2, value="=ACT_ILF_TABLE(100000, {100000,250000,500000,1000000}, 2)")
    row += 1
    ws.cell(row=row, column=1, value="ACT_ILF_TABLE_STANDARD").font = BOLD_FONT
    ws.cell(row=row, column=2, value="=ACT_ILF_TABLE_STANDARD(100000, 5000000, 2)")
    row += 2

    # Deductibles & Loading
    row = add_section_header(ws, row, "Deductibles & Loading")
    row = add_function_test(ws, row, "ACT_LOSS_ELIMINATION_RATIO", "=ACT_LOSS_ELIMINATION_RATIO(10000, 2, 1000)", "LER Pareto")
    row = add_function_test(ws, row, "ACT_LOSS_ELIMINATION_RATIO_LOGNORMAL", "=ACT_LOSS_ELIMINATION_RATIO_LOGNORMAL(10000, 10, 1)", "LER Lognormal")
    row = add_function_test(ws, row, "ACT_DEDUCTIBLE_CREDIT", "=ACT_DEDUCTIBLE_CREDIT(10000, 2, 1000)", "Ded credit")
    row = add_function_test(ws, row, "ACT_PREMIUM_WITH_DEDUCTIBLE", "=ACT_PREMIUM_WITH_DEDUCTIBLE(100000, 10000, 2, 1000)", "Premium with ded")
    row = add_function_test(ws, row, "ACT_LARGE_LOSS_LOAD", "=ACT_LARGE_LOSS_LOAD(0.1, 500000, 100000)", "Large loss load")
    row = add_function_test(ws, row, "ACT_SHOCK_LOSS_LOAD", "=ACT_SHOCK_LOSS_LOAD(0.01, 10000000, 100000)", "Shock loss load")
    row += 1

    # Burning Cost & Rating
    row = add_section_header(ws, row, "Burning Cost & Rating")
    row = add_function_test(ws, row, "ACT_BURNING_COST", "=ACT_BURNING_COST({50000,60000,40000}, {1000000,1100000,1050000})", "Burning cost")
    row = add_function_test(ws, row, "ACT_CREDIBILITY_RATE", "=ACT_CREDIBILITY_RATE(0.05, 0.03, 0.6)", "Cred rate")
    row = add_function_test(ws, row, "ACT_COMPOSITE_RATE", "=ACT_COMPOSITE_RATE({0.05,0.06,0.04}, {0.5,0.3,0.2})", "Composite rate")
    row = add_function_test(ws, row, "ACT_MULTIYEAR_CREDIBILITY", "=ACT_MULTIYEAR_CREDIBILITY(3, 100, 0.5)", "Multi-year Z")
    row += 1

    # Schedule & Retro Rating
    row = add_section_header(ws, row, "Schedule & Retrospective Rating")
    row = add_function_test(ws, row, "ACT_SCHEDULE_RATING", "=ACT_SCHEDULE_RATING(100000, {0.1,-0.05,0.15})", "Schedule rating")
    row = add_function_test(ws, row, "ACT_SCHEDULE_RATING_CAPPED", "=ACT_SCHEDULE_RATING_CAPPED(100000, {0.1,-0.05,0.15}, {0.15,0.1,0.2})", "Capped schedule")
    row = add_function_test(ws, row, "ACT_RETRO_PREMIUM", "=ACT_RETRO_PREMIUM(100000, 50000, 0.3, 1.1, 0.6, 1.2, 1.03)", "Retro premium")
    ws.cell(row=row, column=1, value="ACT_RETRO_PARAMETERS").font = BOLD_FONT
    ws.cell(row=row, column=2, value="=ACT_RETRO_PARAMETERS(0.6, 1.2, 0.05, 1.1)")
    row += 1
    row = add_function_test(ws, row, "ACT_RETRO_LOSS_LIMIT_CREDIT", "=ACT_RETRO_LOSS_LIMIT_CREDIT(250000, 100000, 2, 1000)", "Loss limit credit")
    row += 1

    # Bayesian & Poisson
    row = add_section_header(ws, row, "Bayesian & Poisson Credibility")
    row = add_function_test(ws, row, "ACT_BAYESIAN_UPDATE", "=ACT_BAYESIAN_UPDATE(0.05, 0.01, 0.06, 0.005)", "Bayes update")
    ws.cell(row=row, column=1, value="ACT_BAYESIAN_UPDATE_FULL").font = BOLD_FONT
    ws.cell(row=row, column=2, value="=ACT_BAYESIAN_UPDATE_FULL(0.05, 0.01, 0.06, 0.005)")
    row += 1
    row = add_function_test(ws, row, "ACT_CREDIBILITY_POISSON", "=ACT_CREDIBILITY_POISSON(10, 5, 2)", "Poisson cred")
    row = add_function_test(ws, row, "ACT_FREQUENCY_ESTIMATE", "=ACT_FREQUENCY_ESTIMATE(10, 1000, 5, 2, 0.005)", "Freq estimate")
    row += 1

    return row


def create_copula_tests(ws, start_row):
    """Add copula tests."""
    row = start_row
    row = add_section_header(ws, row, "COPULA FUNCTIONS")
    row += 1

    row = add_section_header(ws, row, "Gaussian Copula")
    ws.cell(row=row, column=1, value="ACT_COPULA_GAUSSIAN").font = BOLD_FONT
    ws.cell(row=row, column=2, value="=ACT_COPULA_GAUSSIAN({1,0.5;0.5,1}, 5, 42)")
    ws.cell(row=row, column=3, value="5 samples, seed=42")
    row += 1
    row = add_function_test(ws, row, "ACT_COPULA_GAUSSIAN_SINGLE", "=ACT_COPULA_GAUSSIAN_SINGLE({1,0.5;0.5,1}, 42)", "Single sample")
    row += 1

    row = add_section_header(ws, row, "Student-t Copula")
    ws.cell(row=row, column=1, value="ACT_COPULA_STUDENT_T").font = BOLD_FONT
    ws.cell(row=row, column=2, value="=ACT_COPULA_STUDENT_T({1,0.5;0.5,1}, 5, 5, 42)")
    ws.cell(row=row, column=3, value="df=5, 5 samples")
    row += 1
    row = add_function_test(ws, row, "ACT_COPULA_STUDENT_T_SINGLE", "=ACT_COPULA_STUDENT_T_SINGLE({1,0.5;0.5,1}, 5, 42)", "Single sample")
    row += 1

    row = add_section_header(ws, row, "Archimedean Copulas")
    ws.cell(row=row, column=1, value="ACT_COPULA_CLAYTON").font = BOLD_FONT
    ws.cell(row=row, column=2, value="=ACT_COPULA_CLAYTON(2, 5, 42)")
    ws.cell(row=row, column=3, value="θ=2, lower tail dep")
    row += 1
    row = add_function_test(ws, row, "ACT_COPULA_CLAYTON_SINGLE", "=ACT_COPULA_CLAYTON_SINGLE(2, 42)", "Single pair")
    row = add_function_test(ws, row, "ACT_COPULA_CLAYTON_CDF", "=ACT_COPULA_CLAYTON_CDF(0.5, 0.5, 2)", "C(0.5,0.5)")
    row += 1

    ws.cell(row=row, column=1, value="ACT_COPULA_FRANK").font = BOLD_FONT
    ws.cell(row=row, column=2, value="=ACT_COPULA_FRANK(5, 5, 42)")
    ws.cell(row=row, column=3, value="θ=5, no tail dep")
    row += 1
    row = add_function_test(ws, row, "ACT_COPULA_FRANK_SINGLE", "=ACT_COPULA_FRANK_SINGLE(5, 42)", "Single pair")
    row = add_function_test(ws, row, "ACT_COPULA_FRANK_CDF", "=ACT_COPULA_FRANK_CDF(0.5, 0.5, 5)", "C(0.5,0.5)")
    row += 1

    ws.cell(row=row, column=1, value="ACT_COPULA_GUMBEL").font = BOLD_FONT
    ws.cell(row=row, column=2, value="=ACT_COPULA_GUMBEL(2, 5, 42)")
    ws.cell(row=row, column=3, value="θ=2, upper tail dep")
    row += 1
    row = add_function_test(ws, row, "ACT_COPULA_GUMBEL_SINGLE", "=ACT_COPULA_GUMBEL_SINGLE(2, 42)", "Single pair")
    row = add_function_test(ws, row, "ACT_COPULA_GUMBEL_CDF", "=ACT_COPULA_GUMBEL_CDF(0.5, 0.5, 2)", "C(0.5,0.5)")
    row += 1

    row = add_section_header(ws, row, "Copula Utilities")
    row = add_function_test(ws, row, "ACT_COPULA_TAU_TO_THETA", "=ACT_COPULA_TAU_TO_THETA(0.5, \"CLAYTON\")", "τ→θ Clayton")
    row = add_function_test(ws, row, "ACT_COPULA_TAIL_LOWER", "=ACT_COPULA_TAIL_LOWER(\"CLAYTON\", 2)", "Lower tail dep")
    row = add_function_test(ws, row, "ACT_COPULA_TAIL_UPPER", "=ACT_COPULA_TAIL_UPPER(\"GUMBEL\", 2)", "Upper tail dep")
    row += 1

    return row


def create_chain_ladder_sheet(wb):
    """Create chain ladder sheet with proper formulas."""
    ws = wb.create_sheet("Chain Ladder Tests")

    row = 1
    ws.cell(row=row, column=1, value="CHAIN LADDER FUNCTIONS").font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 2

    # Sample triangle
    ws.cell(row=row, column=1, value="Sample Triangle (Taylor-Ashe style)").font = HEADER_FONT
    row += 1

    triangle_start_row = row
    triangle = [
        [3578, 4219, 4651, 4672, 4729, 4752, 4756, 4758, 4759, 4760],
        [3239, 4232, 4641, 4736, 4854, 4903, 4930, 4948, 4953, None],
        [3303, 4406, 4788, 4895, 5021, 5078, 5108, 5125, None, None],
        [3529, 4655, 5040, 5173, 5313, 5378, 5413, None, None, None],
        [3701, 4838, 5235, 5387, 5542, 5614, None, None, None, None],
        [3756, 4880, 5279, 5442, 5608, None, None, None, None, None],
        [3809, 4942, 5347, 5517, None, None, None, None, None, None],
        [3882, 5027, 5441, None, None, None, None, None, None, None],
        [3960, 5123, None, None, None, None, None, None, None, None],
        [4043, None, None, None, None, None, None, None, None, None],
    ]

    for i, tri_row in enumerate(triangle):
        for j, val in enumerate(tri_row):
            if val is not None:
                ws.cell(row=row + i, column=2 + j, value=val)

    triangle_range = f"B{triangle_start_row}:K{triangle_start_row + 9}"
    row = triangle_start_row + 11

    # Development Factors - ARRAY FUNCTION (no @)
    ws.cell(row=row, column=1, value="Development Factors").font = HEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value="ACT_CL_FACTORS").font = BOLD_FONT
    ws.cell(row=row, column=2, value=f"=ACT_CL_FACTORS({triangle_range})")
    ws.cell(row=row, column=3, value="← Array output - should spill right")
    row += 2

    # Latest Diagonal - ARRAY FUNCTION (no @)
    ws.cell(row=row, column=1, value="Latest Diagonal").font = HEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value="ACT_CL_LATEST").font = BOLD_FONT
    ws.cell(row=row, column=2, value=f"=ACT_CL_LATEST({triangle_range})")
    ws.cell(row=row, column=3, value="← Array output - should spill down")
    row += 2

    # Ultimate - ARRAY FUNCTION (no @)
    ws.cell(row=row, column=1, value="Chain Ladder Ultimate").font = HEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value="ACT_CL_ULTIMATE").font = BOLD_FONT
    ws.cell(row=row, column=2, value=f"=ACT_CL_ULTIMATE({triangle_range})")
    ws.cell(row=row, column=3, value="← Array output")
    row += 2

    # IBNR - ARRAY FUNCTION (no @)
    ws.cell(row=row, column=1, value="IBNR Reserves").font = HEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value="ACT_CL_IBNR").font = BOLD_FONT
    ws.cell(row=row, column=2, value=f"=ACT_CL_IBNR({triangle_range})")
    ws.cell(row=row, column=3, value="← Array output")
    row += 2

    # Mack SE - ARRAY FUNCTION (no @)
    ws.cell(row=row, column=1, value="Mack Reserve Standard Errors").font = HEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value="ACT_MACK_RESERVE_SE").font = BOLD_FONT
    ws.cell(row=row, column=2, value=f"=ACT_MACK_RESERVE_SE({triangle_range})")
    ws.cell(row=row, column=3, value="← Array output")
    row += 2

    ws.cell(row=row, column=1, value="ACT_MACK_FACTOR_SE").font = BOLD_FONT
    ws.cell(row=row, column=2, value=f"=ACT_MACK_FACTOR_SE({triangle_range})")
    ws.cell(row=row, column=3, value="← Array output")
    row += 2

    # Bootstrap - ARRAY FUNCTION (no @)
    ws.cell(row=row, column=1, value="Bootstrap Chain Ladder").font = HEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value="ACT_BOOTSTRAP_CL").font = BOLD_FONT
    ws.cell(row=row, column=2, value=f"=ACT_BOOTSTRAP_CL({triangle_range}, 1000, 42)")
    ws.cell(row=row, column=3, value="← Array output (1000 iterations)")
    row += 2

    ws.cell(row=row, column=1, value="ACT_BOOTSTRAP_CL_ORIGIN").font = BOLD_FONT
    ws.cell(row=row, column=2, value=f"=ACT_BOOTSTRAP_CL_ORIGIN({triangle_range}, 1000, 42)")
    ws.cell(row=row, column=3, value="← Array output by origin year")
    row += 2

    # B-F
    ws.cell(row=row, column=1, value="Bornhuetter-Ferguson").font = HEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value="ACT_BF_ULTIMATE").font = BOLD_FONT
    factors_ref = f"ACT_CL_FACTORS({triangle_range})"
    ws.cell(row=row, column=2, value=f"=ACT_BF_ULTIMATE({triangle_range}, {factors_ref}, {{5000,5100,5200,5300,5400,5500,5600,5700,5800,5900}})")
    ws.cell(row=row, column=3, value="← Array output (with a-priori)")
    row += 2

    # Cape Cod
    ws.cell(row=row, column=1, value="Cape Cod Method").font = HEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value="ACT_CAPECOD_ULTIMATE").font = BOLD_FONT
    ws.cell(row=row, column=2, value=f"=ACT_CAPECOD_ULTIMATE({triangle_range}, {factors_ref}, {{6000,6100,6200,6300,6400,6500,6600,6700,6800,6900}})")
    row += 1
    ws.cell(row=row, column=1, value="ACT_CAPECOD_ELR").font = BOLD_FONT
    ws.cell(row=row, column=2, value=f"=ACT_CAPECOD_ELR({triangle_range}, {factors_ref}, {{6000,6100,6200,6300,6400,6500,6600,6700,6800,6900}})")
    row += 2

    # Triangle utilities
    ws.cell(row=row, column=1, value="Triangle Utilities").font = HEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value="ACT_TRIANGLE_TO_INCREMENTAL").font = BOLD_FONT
    ws.cell(row=row, column=2, value=f"=ACT_TRIANGLE_TO_INCREMENTAL({triangle_range})")
    ws.cell(row=row, column=3, value="← Array output")
    row += 1
    ws.cell(row=row, column=1, value="ACT_TRIANGLE_DIAGONAL").font = BOLD_FONT
    ws.cell(row=row, column=2, value=f"=ACT_TRIANGLE_DIAGONAL({triangle_range})")
    ws.cell(row=row, column=3, value="← Array output")
    row += 1
    ws.cell(row=row, column=1, value="ACT_TRIANGLE_LINK_RATIOS").font = BOLD_FONT
    ws.cell(row=row, column=2, value=f"=ACT_TRIANGLE_LINK_RATIOS({triangle_range})")
    ws.cell(row=row, column=3, value="← Array output")
    row += 1
    ws.cell(row=row, column=1, value="ACT_CL_CALENDAR_TOTALS").font = BOLD_FONT
    ws.cell(row=row, column=2, value=f"=ACT_CL_CALENDAR_TOTALS({triangle_range})")
    ws.cell(row=row, column=3, value="← Array output")
    row += 2

    # Berquist-Sherman
    ws.cell(row=row, column=1, value="ACT_BERQUIST_SHERMAN").font = BOLD_FONT
    ws.cell(row=row, column=2, value=f"=ACT_BERQUIST_SHERMAN({triangle_range}, {{1.05,1.04,1.03,1.02,1.01,1.0,0.99,0.98,0.97,0.96}})")
    ws.cell(row=row, column=3, value="← Array output (adjusted triangle)")
    row += 2

    # Calendar year adjustment
    ws.cell(row=row, column=1, value="ACT_CL_CALENDAR_ADJUST").font = BOLD_FONT
    ws.cell(row=row, column=2, value=f"=ACT_CL_CALENDAR_ADJUST({triangle_range}, 0.03)")
    ws.cell(row=row, column=3, value="← Array output (3% annual trend)")
    row += 2

    # Weighted average
    ws.cell(row=row, column=1, value="ACT_CL_WEIGHTED_AVERAGE").font = BOLD_FONT
    ws.cell(row=row, column=2, value=f"=ACT_CL_WEIGHTED_AVERAGE({{5000,5100,5200}}, {{4900,5050,5150}}, {{0.6,0.4}})")
    ws.cell(row=row, column=3, value="← Array output")
    row += 1

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 30

    return ws


def create_versions_sheet(wb):
    """Create versions sheet."""
    ws = wb.create_sheet("Versions", 0)

    COMMITS = [
        ("b2a74d4", "2026-01-18", "Add Phase 6: Credibility and experience rating functions"),
        ("8e5109b", "2026-01-18", "Add Phase 5: Reserving enhancements"),
        ("6a2d830", "2026-01-18", "Add Phase 3: Archimedean copulas"),
        ("2f71e98", "2026-01-18", "Add Phase 2: Parameter estimation"),
        ("79ef7bc", "2026-01-18", "Implement Phase 1: distributions, copulas, help text"),
    ]

    row = 1
    ws.cell(row=row, column=1, value="Actuarial Add-In v0.1.0").font = TITLE_FONT
    row += 2

    ws.cell(row=row, column=1, value="Build Date:").font = BOLD_FONT
    ws.cell(row=row, column=2, value="2026-01-18")
    row += 1

    ws.cell(row=row, column=1, value="GitHub:").font = BOLD_FONT
    ws.cell(row=row, column=2, value="https://github.com/mdevans21/actuarial_add_in")
    ws.cell(row=row, column=2).hyperlink = "https://github.com/mdevans21/actuarial_add_in"
    row += 2

    ws.cell(row=row, column=1, value="Version Functions:").font = HEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value="ACT_VERSION").font = BOLD_FONT
    ws.cell(row=row, column=2, value="=ACT_VERSION()")
    row += 1
    ws.cell(row=row, column=1, value="ACT_BUILD_DATE").font = BOLD_FONT
    ws.cell(row=row, column=2, value="=ACT_BUILD_DATE()")
    row += 1
    ws.cell(row=row, column=1, value="ACT_GITHUB_URL").font = BOLD_FONT
    ws.cell(row=row, column=2, value="=ACT_GITHUB_URL()")
    row += 1
    ws.cell(row=row, column=1, value="ACT_COMMIT_COUNT").font = BOLD_FONT
    ws.cell(row=row, column=2, value="=ACT_COMMIT_COUNT()")
    row += 1
    ws.cell(row=row, column=1, value="ACT_COMMIT_INFO").font = BOLD_FONT
    ws.cell(row=row, column=2, value="=ACT_COMMIT_INFO(1, \"message\")")
    ws.cell(row=row, column=3, value="Latest commit message")
    row += 1
    ws.cell(row=row, column=1, value="ACT_COMMIT_HISTORY").font = BOLD_FONT
    ws.cell(row=row, column=2, value="=ACT_COMMIT_HISTORY()")
    ws.cell(row=row, column=3, value="← Array output")
    row += 2

    ws.cell(row=row, column=1, value="Note on Array Formulas:").font = HEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value="Functions returning arrays should NOT have @ prefix.")
    row += 1
    ws.cell(row=row, column=1, value="If you see @, delete it or re-enter the formula.")
    row += 1
    ws.cell(row=row, column=1, value="In Excel 365, arrays spill automatically.")
    row += 1

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 25

    return ws


def main():
    project_root = Path(__file__).parent.parent
    target_file = project_root / "excel" / "actuarial_add_in_v0.1.xlsm"

    # Start fresh
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Create Versions sheet
    create_versions_sheet(wb)

    # Create main test sheet
    ws = wb.create_sheet("All Functions")
    row = 1
    ws.cell(row=row, column=1, value="ACTUARIAL ADD-IN - ALL FUNCTIONS").font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 2

    row = create_distributions_tests(ws, row)
    row = create_exposure_curves_tests(ws, row)
    row = create_reinsurance_tests(ws, row)
    row = create_interpolation_tests(ws, row)
    row = create_credibility_tests(ws, row)
    row = create_copula_tests(ws, row)

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 25

    # Create Chain Ladder sheet
    create_chain_ladder_sheet(wb)

    # Save
    wb.save(target_file)
    print(f"Created: {target_file}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
