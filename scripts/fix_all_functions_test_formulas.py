#!/usr/bin/env python3
"""
Fix formula errors in the All Functions Test sheet.
"""

import openpyxl
from openpyxl.styles import Font
import os
import shutil

OLD_EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', 'excel', 'actuarial_add_in_v0.2.xlsm')
NEW_EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', 'excel', 'actuarial_add_in.xlsm')

# Formula fixes: (function_name, old_formula_part, new_formula)
# We match by function name in column A and replace the formula in column B
FORMULA_FIXES = {
    # Category 1: Wrong parameter order
    'ACT_EXPERIENCE_MOD': {
        'old': '=ACT_EXPERIENCE_MOD(0.7, 50000, 70000)',
        'new': '=ACT_EXPERIENCE_MOD(50000, 70000, 0.7)',
        'note': 'Experience mod factor'
    },
    'ACT_FULL_CREDIBILITY_STANDARD': {
        'old': '=ACT_FULL_CREDIBILITY_STANDARD(0.05, 0.90, 1.5)',
        'new': '=ACT_FULL_CREDIBILITY_STANDARD(0.90, 0.05, 1.5)',
        'note': 'Full cred standard n₀'
    },
    'ACT_MULTIYEAR_CREDIBILITY': {
        'old': '=ACT_MULTIYEAR_CREDIBILITY(3, 100, 0.5)',
        'new': '=ACT_MULTIYEAR_CREDIBILITY(0.5, 3, 0.1)',
        'note': 'Multi-year credibility'
    },
    'ACT_RETRO_PREMIUM': {
        'old': '=ACT_RETRO_PREMIUM(100000, 50000, 0.3, 1.1, 0.6, 1.2, 1.03)',
        'new': '=ACT_RETRO_PREMIUM(100000, 50000, 1.1, 0.3, 1.03, 0.6, 1.2)',
        'note': 'Retro premium'
    },

    # Category 2: Wrong number of parameters
    'ACT_CREDIBILITY_ESTIMATE': {
        'old': '=ACT_CREDIBILITY_ESTIMATE(10, 5, 0.05, 0.03)',
        'new': '=ACT_CREDIBILITY_ESTIMATE(0.05, 0.03, 0.67)',
        'note': 'Blended estimate'
    },
    'ACT_RETRO_PARAMETERS': {
        'old': '=ACT_RETRO_PARAMETERS(0.6, 1.2, 0.05, 1.1)',
        'new': '=ACT_RETRO_PARAMETERS(0.6, 0.6, 1.2, 1.03, 0.05)',
        'note': 'Retro parameters (array)'
    },

    # Category 3: Invalid parameter values
    'ACT_ILF_LAYER': {
        'old': '=ACT_ILF_LAYER(1000000, 4000000, 100000, 2)',
        'new': '=ACT_ILF_LAYER(100000, 400000, 10000, 2)',
        'note': 'Layer ILF'
    },
    'ACT_RETRO_LOSS_LIMIT_CREDIT': {
        'old': '=ACT_RETRO_LOSS_LIMIT_CREDIT(250000, 100000, 2, 1000)',
        'new': '=ACT_RETRO_LOSS_LIMIT_CREDIT(250000, 0.65, 2, 10000)',
        'note': 'Loss limit credit'
    },

    # Category 4: Credibility rate - add missing LDF/trend explicitly
    'ACT_CREDIBILITY_RATE': {
        'old': '=ACT_CREDIBILITY_RATE(0.05, 0.03, 0.6)',
        'new': '=ACT_CREDIBILITY_RATE(0.05, 0.03, 0.6, 1.0, 1.0)',
        'note': 'Credibility-weighted rate'
    },
}

# Fixes for rows where the same function appears multiple times (use row-specific matching)
ROW_SPECIFIC_FIXES = [
    # ACT_COPULA_TAIL_LOWER for Student-t (there are two of these)
    {
        'func': 'ACT_COPULA_TAIL_LOWER',
        'old_contains': 'STUDENT_T',
        'new': '=ACT_COPULA_TAIL_LOWER("STUDENT_T", 0.5, 5)',
        'note': 'Student-t lower tail'
    },
    # ACT_COPULA_TAIL_UPPER for Student-t
    {
        'func': 'ACT_COPULA_TAIL_UPPER',
        'old_contains': 'STUDENT_T',
        'new': '=ACT_COPULA_TAIL_UPPER("STUDENT_T", 0.5, 5)',
        'note': 'Student-t upper tail'
    },
]


def fix_formulas():
    print(f"Opening workbook: {OLD_EXCEL_PATH}")
    wb = openpyxl.load_workbook(OLD_EXCEL_PATH, keep_vba=True)

    if "All Functions Test" not in wb.sheetnames:
        print("Error: 'All Functions Test' sheet not found")
        return

    ws = wb["All Functions Test"]
    fixes_applied = 0

    # First pass: simple function name matching
    for row_num in range(1, 250):
        cell_a = ws.cell(row=row_num, column=1).value
        cell_b = ws.cell(row=row_num, column=2).value

        if cell_a and cell_a in FORMULA_FIXES:
            fix = FORMULA_FIXES[cell_a]
            if cell_b and fix['old'] in str(cell_b):
                print(f"Row {row_num}: Fixing {cell_a}")
                print(f"  Old: {cell_b}")
                print(f"  New: {fix['new']}")
                ws.cell(row=row_num, column=2).value = fix['new']
                fixes_applied += 1

    # Second pass: row-specific fixes for duplicate function names
    for row_num in range(1, 250):
        cell_a = ws.cell(row=row_num, column=1).value
        cell_b = ws.cell(row=row_num, column=2).value

        for fix in ROW_SPECIFIC_FIXES:
            if cell_a == fix['func'] and cell_b and fix['old_contains'] in str(cell_b):
                # Check if already fixed
                if fix['new'] != str(cell_b):
                    print(f"Row {row_num}: Fixing {cell_a} ({fix['note']})")
                    print(f"  Old: {cell_b}")
                    print(f"  New: {fix['new']}")
                    ws.cell(row=row_num, column=2).value = fix['new']
                    fixes_applied += 1

    print(f"\nApplied {fixes_applied} fixes")

    # Save to new filename (removing _v0.2)
    print(f"\nSaving to new path: {NEW_EXCEL_PATH}")
    wb.save(NEW_EXCEL_PATH)

    # Remove old file
    if os.path.exists(OLD_EXCEL_PATH):
        os.remove(OLD_EXCEL_PATH)
        print(f"Removed old file: {OLD_EXCEL_PATH}")

    # Also remove any temp file
    temp_file = OLD_EXCEL_PATH.replace('.xlsm', '.xlsm').replace('actuarial_add_in_v0.2', '~$actuarial_add_in_v0.2')
    if os.path.exists(temp_file):
        try:
            os.remove(temp_file)
        except:
            pass

    print("Done!")


if __name__ == "__main__":
    fix_formulas()
