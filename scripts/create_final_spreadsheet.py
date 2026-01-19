#!/usr/bin/env python3
"""
Create final spreadsheet: copy original, fix @ issues, add comprehensive test sheet.
"""

import shutil
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import re

# Styles
TITLE_FONT = Font(bold=True, size=14)
HEADER_FONT = Font(bold=True, size=12)
BOLD_FONT = Font(bold=True)

# Functions that return arrays (should NOT have @ prefix)
ARRAY_FUNCTIONS = [
    'ACT_CL_FACTORS', 'ACT_CL_LATEST', 'ACT_CL_ULTIMATE', 'ACT_CL_IBNR',
    'ACT_BF_ULTIMATE', 'ACT_MACK_FACTOR_SE', 'ACT_MACK_RESERVE_SE',
    'ACT_CL_BOOTSTRAP', 'ACT_CL_BOOTSTRAP_ORIGIN',
    'ACT_CAPECOD_ULTIMATE', 'ACT_CAPECOD_ELR',
    'ACT_TRIANGLE_TO_INCREMENTAL', 'ACT_INCREMENTAL_TO_CUMULATIVE',
    'ACT_TRIANGLE_DIAGONAL', 'ACT_TRIANGLE_LINK_RATIOS',
    'ACT_CL_CALENDAR_ADJUST', 'ACT_CL_CALENDAR_TOTALS',
    'ACT_CL_WEIGHTED_AVERAGE', 'ACT_BERQUIST_SHERMAN',
    'ACT_COPULA_GAUSSIAN', 'ACT_COPULA_STUDENT_T',
    'ACT_COPULA_CLAYTON', 'ACT_COPULA_FRANK', 'ACT_COPULA_GUMBEL',
    'ACT_RETURN_PERIOD_TABLE', 'ACT_COMMIT_HISTORY',
    'ACT_ILF_TABLE', 'ACT_ILF_TABLE_STANDARD',
    'ACT_BUHLMANN_STRAUB_PARAMS', 'ACT_RETRO_PARAMETERS',
    'ACT_BAYESIAN_UPDATE_FULL',
]


def fix_at_formulas(ws):
    """Remove @ prefix from array function formulas."""
    fixed = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and '=' in cell.value:
                original = cell.value
                new_value = original

                # Fix @ before array functions
                for func in ARRAY_FUNCTIONS:
                    if f'@{func}' in new_value:
                        new_value = new_value.replace(f'@{func}', func)

                # Also fix =@ at start for any ACT function
                if new_value.startswith('=@ACT_'):
                    new_value = '=' + new_value[2:]

                if new_value != original:
                    cell.value = new_value
                    fixed += 1
                    print(f"  Fixed: {cell.coordinate}: {original} -> {new_value}")

    return fixed


def create_versions_sheet(wb):
    """Create or replace versions sheet."""
    if "Versions" in wb.sheetnames:
        del wb["Versions"]

    ws = wb.create_sheet("Versions", 0)

    row = 1
    ws.cell(row=row, column=1, value="Actuarial Add-In v0.1.0").font = TITLE_FONT
    row += 2

    ws.cell(row=row, column=1, value="Build Date:").font = BOLD_FONT
    ws.cell(row=row, column=2, value="2026-01-18")
    row += 1

    ws.cell(row=row, column=1, value="GitHub:").font = BOLD_FONT
    ws.cell(row=row, column=2, value="https://github.com/mdevans21/actuarial_add_in")
    row += 2

    ws.cell(row=row, column=1, value="Version Functions:").font = HEADER_FONT
    row += 1
    ws.cell(row=row, column=2, value="=ACT_VERSION()")
    row += 1
    ws.cell(row=row, column=2, value="=ACT_BUILD_DATE()")
    row += 1
    ws.cell(row=row, column=2, value="=ACT_COMMIT_HISTORY()")
    ws.cell(row=row, column=3, value="(array output)")
    row += 2

    ws.cell(row=row, column=1, value="IMPORTANT: Array Formulas").font = HEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value="If formulas show @ prefix (e.g., =@ACT_CL_FACTORS), the @ should be removed.")
    row += 1
    ws.cell(row=row, column=1, value="Array functions will spill their results in Excel 365.")
    row += 1
    ws.cell(row=row, column=1, value="For older Excel, use INDEX() to extract values, e.g.: =INDEX(ACT_CL_FACTORS(...), 1)")
    row += 2

    ws.cell(row=row, column=1, value="Array-returning functions (no @ prefix):").font = HEADER_FONT
    row += 1
    for func in sorted(ARRAY_FUNCTIONS):
        ws.cell(row=row, column=1, value=f"  {func}")
        row += 1

    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20

    return ws


def add_comprehensive_tests_sheet(wb):
    """Add sheet with all functions tested."""
    if "All Functions Test" in wb.sheetnames:
        del wb["All Functions Test"]

    ws = wb.create_sheet("All Functions Test")

    row = 1
    ws.cell(row=row, column=1, value="COMPREHENSIVE FUNCTION TESTS").font = TITLE_FONT
    row += 2

    # All test cases organized by category
    tests = [
        # Distributions
        ("DISTRIBUTIONS", None),
        ("Poisson (λ=5)", None),
        ("ACT_POISSON_PDF", "=ACT_POISSON_PDF(3, 5)"),
        ("ACT_POISSON_CDF", "=ACT_POISSON_CDF(3, 5)"),
        ("ACT_POISSON_INV", "=ACT_POISSON_INV(0.5, 5)"),
        ("", None),
        ("Negative Binomial", None),
        ("ACT_NEGBIN_PDF", "=ACT_NEGBIN_PDF(5, 5, 0.3)"),
        ("ACT_NEGBIN_CDF", "=ACT_NEGBIN_CDF(5, 5, 0.3)"),
        ("ACT_NEGBIN_INV", "=ACT_NEGBIN_INV(0.5, 5, 0.3)"),
        ("", None),
        ("Lognormal (μ=0, σ=1)", None),
        ("ACT_LOGNORM_PDF", "=ACT_LOGNORM_PDF(1, 0, 1)"),
        ("ACT_LOGNORM_CDF", "=ACT_LOGNORM_CDF(1, 0, 1)"),
        ("ACT_LOGNORM_INV", "=ACT_LOGNORM_INV(0.5, 0, 1)"),
        ("", None),
        ("Gamma (α=2, β=1)", None),
        ("ACT_GAMMA_PDF", "=ACT_GAMMA_PDF(1, 2, 1)"),
        ("ACT_GAMMA_CDF", "=ACT_GAMMA_CDF(1, 2, 1)"),
        ("ACT_GAMMA_INV", "=ACT_GAMMA_INV(0.5, 2, 1)"),
        ("", None),
        ("Pareto (α=2, xm=1)", None),
        ("ACT_PARETO_PDF", "=ACT_PARETO_PDF(2, 2, 1)"),
        ("ACT_PARETO_CDF", "=ACT_PARETO_CDF(2, 2, 1)"),
        ("ACT_PARETO_INV", "=ACT_PARETO_INV(0.5, 2, 1)"),
        ("", None),
        ("Weibull (k=2, λ=1)", None),
        ("ACT_WEIBULL_PDF", "=ACT_WEIBULL_PDF(1, 2, 1)"),
        ("ACT_WEIBULL_CDF", "=ACT_WEIBULL_CDF(1, 2, 1)"),
        ("ACT_WEIBULL_INV", "=ACT_WEIBULL_INV(0.5, 2, 1)"),
        ("", None),
        ("Beta (α=2, β=5)", None),
        ("ACT_BETA_PDF", "=ACT_BETA_PDF(0.3, 2, 5)"),
        ("ACT_BETA_CDF", "=ACT_BETA_CDF(0.3, 2, 5)"),
        ("ACT_BETA_INV", "=ACT_BETA_INV(0.5, 2, 5)"),
        ("", None),
        ("Exponential (λ=0.5)", None),
        ("ACT_EXP_PDF", "=ACT_EXP_PDF(1, 0.5)"),
        ("ACT_EXP_CDF", "=ACT_EXP_CDF(1, 0.5)"),
        ("ACT_EXP_INV", "=ACT_EXP_INV(0.5, 0.5)"),
        ("", None),
        ("GPD (ξ=0.5, σ=1)", None),
        ("ACT_GPD_PDF", "=ACT_GPD_PDF(1, 0.5, 1)"),
        ("ACT_GPD_CDF", "=ACT_GPD_CDF(1, 0.5, 1)"),
        ("ACT_GPD_INV", "=ACT_GPD_INV(0.5, 0.5, 1)"),
        ("", None),
        ("Burr XII (c=2, k=3, λ=1)", None),
        ("ACT_BURR_PDF", "=ACT_BURR_PDF(1, 2, 3, 1)"),
        ("ACT_BURR_CDF", "=ACT_BURR_CDF(1, 2, 3, 1)"),
        ("ACT_BURR_INV", "=ACT_BURR_INV(0.5, 2, 3, 1)"),
        ("", None),

        # Exposure Curves
        ("EXPOSURE CURVES", None),
        ("ACT_MBBEFD", "=ACT_MBBEFD(0.5, 2, 3)"),
        ("ACT_SWISSRE_CURVE", "=ACT_SWISSRE_CURVE(0.5, 3)"),
        ("ACT_LLOYDS_CURVE", "=ACT_LLOYDS_CURVE(0.5, 2)"),
        ("ACT_POWER_CURVE", "=ACT_POWER_CURVE(0.5, 2)"),
        ("ACT_INVERSE_POWER_CURVE", "=ACT_INVERSE_POWER_CURVE(0.5, 2)"),
        ("ACT_PARETO_EXPOSURE", "=ACT_PARETO_EXPOSURE(0.5, 2)"),
        ("ACT_RIEBESELL_CURVE", "=ACT_RIEBESELL_CURVE(0.5, 0.8)"),
        ("", None),

        # Reinsurance
        ("REINSURANCE", None),
        ("ACT_XOL_LAYER_LOSS", "=ACT_XOL_LAYER_LOSS(5000000, 1000000, 4000000)"),
        ("ACT_QS_CEDED", "=ACT_QS_CEDED(1000000, 0.5)"),
        ("ACT_AGGREGATE_LAYER", "=ACT_AGGREGATE_LAYER(15000000, 2000000, 10000000)"),
        ("ACT_ILF_PARETO", "=ACT_ILF_PARETO(2000000, 1000000, 2)"),
        ("ACT_LAYER_RATE_ON_LINE", "=ACT_LAYER_RATE_ON_LINE(100000, 1000000, 4000000)"),
        ("ACT_AAL_FROM_OEP", "=ACT_AAL_FROM_OEP({100,50,25,10,5}, {1000000,2000000,3000000,5000000,10000000})"),
        ("", None),

        # Interpolation
        ("INTERPOLATION", None),
        ("ACT_INTERP", "=ACT_INTERP({1,2,3,4,5}, {10,20,35,55,80}, 2.5, \"LINEAR\")"),
        ("ACT_INTERP_LOG", "=ACT_INTERP_LOG({1,2,3,4,5}, {10,20,35,55,80}, 2.5)"),
        ("", None),

        # Credibility (sample)
        ("CREDIBILITY", None),
        ("ACT_CREDIBILITY_BUHLMANN", "=ACT_CREDIBILITY_BUHLMANN(10, 5)"),
        ("ACT_CREDIBILITY_ESTIMATE", "=ACT_CREDIBILITY_ESTIMATE(10, 5, 0.05, 0.03)"),
        ("ACT_EXPERIENCE_MOD", "=ACT_EXPERIENCE_MOD(0.7, 50000, 70000)"),
        ("ACT_FULL_CREDIBILITY_STANDARD", "=ACT_FULL_CREDIBILITY_STANDARD(0.05, 0.90, 1.5)"),
        ("ACT_PARTIAL_CREDIBILITY", "=ACT_PARTIAL_CREDIBILITY(500, 1082)"),
        ("ACT_BURNING_COST", "=ACT_BURNING_COST({50000,60000,40000}, {1000000,1100000,1050000})"),
        ("ACT_LOSS_ELIMINATION_RATIO", "=ACT_LOSS_ELIMINATION_RATIO(10000, 2, 1000)"),
        ("ACT_DEDUCTIBLE_CREDIT", "=ACT_DEDUCTIBLE_CREDIT(10000, 2, 1000)"),
        ("", None),

        # Copulas
        ("COPULAS", None),
        ("ACT_COPULA_GAUSSIAN_SINGLE", "=ACT_COPULA_GAUSSIAN_SINGLE({1,0.5;0.5,1}, 42)"),
        ("ACT_COPULA_STUDENT_T_SINGLE", "=ACT_COPULA_STUDENT_T_SINGLE({1,0.5;0.5,1}, 5, 42)"),
        ("ACT_COPULA_CLAYTON_SINGLE", "=ACT_COPULA_CLAYTON_SINGLE(2, 42)"),
        ("ACT_COPULA_FRANK_SINGLE", "=ACT_COPULA_FRANK_SINGLE(5, 42)"),
        ("ACT_COPULA_GUMBEL_SINGLE", "=ACT_COPULA_GUMBEL_SINGLE(2, 42)"),
        ("ACT_COPULA_CLAYTON_CDF", "=ACT_COPULA_CLAYTON_CDF(0.5, 0.5, 2)"),
        ("ACT_COPULA_TAU_TO_THETA", "=ACT_COPULA_TAU_TO_THETA(0.5, \"CLAYTON\")"),
        ("ACT_COPULA_TAIL_LOWER", "=ACT_COPULA_TAIL_LOWER(\"CLAYTON\", 2)"),
        ("ACT_COPULA_TAIL_UPPER", "=ACT_COPULA_TAIL_UPPER(\"GUMBEL\", 2)"),
    ]

    for name, formula in tests:
        if formula is None:
            if name:
                ws.cell(row=row, column=1, value=name).font = HEADER_FONT
        else:
            ws.cell(row=row, column=1, value=name).font = BOLD_FONT
            ws.cell(row=row, column=2, value=formula)
        row += 1

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 80

    return ws


def main():
    project_root = Path(__file__).parent.parent
    source_file = project_root / "excel" / "actuarial_add_in_v0.0.xlsm"
    target_file = project_root / "excel" / "actuarial_add_in_v0.1.xlsm"

    # Copy original
    shutil.copy2(source_file, target_file)
    print(f"Copied {source_file.name}")

    # Load and modify
    wb = load_workbook(target_file, keep_vba=True)

    # Fix @ formulas in all existing sheets
    total_fixed = 0
    for sheet_name in wb.sheetnames:
        print(f"\nProcessing sheet: {sheet_name}")
        fixed = fix_at_formulas(wb[sheet_name])
        total_fixed += fixed
        if fixed:
            print(f"  Fixed {fixed} formulas")

    # Create/update Versions sheet
    create_versions_sheet(wb)
    print("\nCreated Versions sheet")

    # Add comprehensive tests sheet
    add_comprehensive_tests_sheet(wb)
    print("Created All Functions Test sheet")

    # Save
    wb.save(target_file)
    print(f"\nSaved: {target_file}")
    print(f"Total @ formulas fixed: {total_fixed}")
    print(f"Final sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
