#!/usr/bin/env python3
"""
Create complete spreadsheet with ALL functions tested and proper Versions sheet.
"""

import shutil
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Styles
TITLE_FONT = Font(bold=True, size=14)
HEADER_FONT = Font(bold=True, size=12)
BOLD_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Commit history
COMMITS = [
    ("b2a74d4", "2026-01-18", "Add Phase 6: Credibility and experience rating functions"),
    ("8e5109b", "2026-01-18", "Add Phase 5: Reserving enhancements - Cape Cod, triangle utilities, calendar year"),
    ("6a2d830", "2026-01-18", "Add Phase 3: Archimedean copulas and utility functions"),
    ("2f71e98", "2026-01-18", "Add Phase 2: Parameter estimation functions for distribution fitting"),
    ("79ef7bc", "2026-01-18", "Implement Phase 1 from next_steps.md: distributions, copulas, and help text"),
    ("bc30112", "2026-01-18", "Add comprehensive review and roadmap for actuarial add-in"),
    ("fb0a7fe", "2026-01-17", "Merge branch 'main'"),
    ("83a9008", "2026-01-17", "Add chain ladder enhancements and update examples"),
    ("b16e651", "2026-01-17", "Update README.md"),
    ("7bb88e5", "2026-01-17", "Update README.md"),
    ("7458388", "2026-01-17", "Add charts to Excel sheets and update Chain Ladder with Taylor-Ashe data"),
    ("dc52de8", "2026-01-17", "Add documentation, examples, and agent instructions"),
    ("1fbd1ef", "2026-01-17", "Rename PRD file to fix double extension"),
    ("d2323be", "2026-01-17", "Fix Poisson inverse CDF implementation"),
    ("62c42d7", "2026-01-17", "Initial implementation of Actuarial Excel Add-in"),
]

# Array-returning functions
ARRAY_FUNCTIONS = [
    'ACT_CL_FACTORS', 'ACT_CL_LATEST', 'ACT_CL_ULTIMATE', 'ACT_CL_IBNR',
    # Exposure curves use ACT_EXPOSURE_* prefix
    # Distributions use ACT_DIST_* prefix
    'ACT_BF_ULTIMATE', 'ACT_MACK_FACTOR_SE', 'ACT_MACK_RESERVE_SE',
    'ACT_CL_BOOTSTRAP', 'ACT_CL_BOOTSTRAP_ORIGIN',
    'ACT_CAPECOD_ULTIMATE', 'ACT_TRIANGLE_TO_INCREMENTAL',
    'ACT_INCREMENTAL_TO_CUMULATIVE', 'ACT_TRIANGLE_DIAGONAL',
    'ACT_TRIANGLE_LINK_RATIOS', 'ACT_CL_CALENDAR_ADJUST',
    'ACT_CL_CALENDAR_TOTALS', 'ACT_CL_WEIGHTED_AVERAGE', 'ACT_BERQUIST_SHERMAN',
    'ACT_COPULA_GAUSSIAN', 'ACT_COPULA_STUDENT_T',
    'ACT_COPULA_CLAYTON', 'ACT_COPULA_FRANK', 'ACT_COPULA_GUMBEL',
    'ACT_RETURN_PERIOD_TABLE', 'ACT_COMMIT_HISTORY',
    'ACT_ILF_TABLE', 'ACT_ILF_TABLE_STANDARD',
    'ACT_BUHLMANN_STRAUB_PARAMS', 'ACT_RETRO_PARAMETERS',
    'ACT_BAYESIAN_UPDATE_FULL',
]


def create_versions_sheet(wb):
    """Create proper Versions sheet with commit history."""
    if "Versions" in wb.sheetnames:
        del wb["Versions"]

    ws = wb.create_sheet("Versions", 0)
    row = 1

    # Title
    ws.cell(row=row, column=1, value="Actuarial Add-In Version History").font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 2

    # Version info
    ws.cell(row=row, column=1, value="Current Version:").font = BOLD_FONT
    ws.cell(row=row, column=2, value="0.1.0")
    row += 1
    ws.cell(row=row, column=1, value="Build Date:").font = BOLD_FONT
    ws.cell(row=row, column=2, value="2026-01-18")
    row += 1
    ws.cell(row=row, column=1, value="GitHub:").font = BOLD_FONT
    ws.cell(row=row, column=2, value="https://github.com/mdevans21/actuarial_add_in")
    ws.cell(row=row, column=2).hyperlink = "https://github.com/mdevans21/actuarial_add_in"
    ws.cell(row=row, column=2).style = "Hyperlink"
    row += 2

    # Version functions
    ws.cell(row=row, column=1, value="Version Functions:").font = HEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value="ACT_VERSION()").font = BOLD_FONT
    ws.cell(row=row, column=2, value="=ACT_VERSION()")
    row += 1
    ws.cell(row=row, column=1, value="ACT_BUILD_DATE()").font = BOLD_FONT
    ws.cell(row=row, column=2, value="=ACT_BUILD_DATE()")
    row += 1
    ws.cell(row=row, column=1, value="ACT_GITHUB_URL()").font = BOLD_FONT
    ws.cell(row=row, column=2, value="=ACT_GITHUB_URL()")
    row += 1
    ws.cell(row=row, column=1, value="ACT_COMMIT_COUNT()").font = BOLD_FONT
    ws.cell(row=row, column=2, value="=ACT_COMMIT_COUNT()")
    row += 1
    ws.cell(row=row, column=1, value="ACT_COMMIT_INFO(n, field)").font = BOLD_FONT
    ws.cell(row=row, column=2, value='=ACT_COMMIT_INFO(1, "message")')
    row += 2

    # Commit History Table
    ws.cell(row=row, column=1, value="Commit History").font = HEADER_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 1

    # Headers
    for col, header in enumerate(["Commit", "Date", "Message"], 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = BOLD_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    row += 1

    # Commit data
    for commit_hash, date, message in COMMITS:
        ws.cell(row=row, column=1, value=commit_hash).border = THIN_BORDER
        ws.cell(row=row, column=2, value=date).border = THIN_BORDER
        ws.cell(row=row, column=3, value=message).border = THIN_BORDER
        row += 1

    row += 1

    # Note about @ prefix
    ws.cell(row=row, column=1, value="Note on Array Formulas:").font = HEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value="If you see @ before function names in Excel 365, delete it.")
    row += 1
    ws.cell(row=row, column=1, value="Array functions need empty cells to spill their results.")
    row += 1
    ws.cell(row=row, column=1, value="Or use INDEX() to extract individual values.")
    row += 2

    # List of array functions
    ws.cell(row=row, column=1, value="Array-returning functions:").font = HEADER_FONT
    row += 1
    for i, func in enumerate(sorted(ARRAY_FUNCTIONS)):
        col = 1 + (i % 3)
        if col == 1 and i > 0:
            row += 1
        ws.cell(row=row, column=col, value=func)
    row += 2

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 60

    return ws


def create_all_functions_sheet(wb):
    """Create sheet with ALL functions tested."""
    if "All Functions Test" in wb.sheetnames:
        del wb["All Functions Test"]

    ws = wb.create_sheet("All Functions Test")
    row = 1

    ws.cell(row=row, column=1, value="COMPREHENSIVE FUNCTION TESTS - ALL 127 FUNCTIONS").font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    # All functions organized by category
    all_tests = [
        # ===== DISTRIBUTIONS =====
        ("DISTRIBUTIONS - Frequency", None, None),
        ("", None, None),
        ("Poisson Distribution (λ=5)", None, None),
        ("ACT_POISSON_PDF", "=ACT_POISSON_PDF(3, 5)", "P(X=3) for Poisson"),
        ("ACT_POISSON_CDF", "=ACT_POISSON_CDF(3, 5)", "P(X≤3)"),
        ("ACT_POISSON_INV", "=ACT_POISSON_INV(0.5, 5)", "Median"),
        ("ACT_POISSON_FIT", "=ACT_POISSON_FIT({1,2,3,4,5,6,7})", "MLE estimate of λ"),
        ("", None, None),
        ("Negative Binomial (r=5, p=0.3)", None, None),
        ("ACT_NEGBIN_PDF", "=ACT_NEGBIN_PDF(5, 5, 0.3)", "P(X=5)"),
        ("ACT_NEGBIN_CDF", "=ACT_NEGBIN_CDF(5, 5, 0.3)", "P(X≤5)"),
        ("ACT_NEGBIN_INV", "=ACT_NEGBIN_INV(0.5, 5, 0.3)", "Median"),
        ("ACT_NEGBIN_FIT", "=ACT_NEGBIN_FIT({1,2,3,4,5,6,7,8,9,10})", "MLE r,p"),
        ("", None, None),

        ("DISTRIBUTIONS - Severity", None, None),
        ("", None, None),
        ("Lognormal (μ=0, σ=1)", None, None),
        ("ACT_LOGNORM_PDF", "=ACT_LOGNORM_PDF(1, 0, 1)", "f(1)"),
        ("ACT_LOGNORM_CDF", "=ACT_LOGNORM_CDF(1, 0, 1)", "F(1)"),
        ("ACT_LOGNORM_INV", "=ACT_LOGNORM_INV(0.5, 0, 1)", "Median = 1"),
        ("ACT_LOGNORM_FIT", "=ACT_LOGNORM_FIT({1,2,3,4,5})", "MLE μ,σ"),
        ("", None, None),
        ("Gamma (α=2, β=1)", None, None),
        ("ACT_GAMMA_PDF", "=ACT_GAMMA_PDF(1, 2, 1)", "f(1)"),
        ("ACT_GAMMA_CDF", "=ACT_GAMMA_CDF(1, 2, 1)", "F(1)"),
        ("ACT_GAMMA_INV", "=ACT_GAMMA_INV(0.5, 2, 1)", "Median"),
        ("ACT_GAMMA_FIT", "=ACT_GAMMA_FIT({1,2,3,4,5})", "MLE α,β"),
        ("", None, None),
        ("Pareto (α=2, xm=1)", None, None),
        ("ACT_PARETO_PDF", "=ACT_PARETO_PDF(2, 2, 1)", "f(2)"),
        ("ACT_PARETO_CDF", "=ACT_PARETO_CDF(2, 2, 1)", "F(2) = 0.75"),
        ("ACT_PARETO_INV", "=ACT_PARETO_INV(0.5, 2, 1)", "Median = √2"),
        ("ACT_PARETO_FIT", "=ACT_PARETO_FIT({1.5,2,2.5,3,4,5}, 1)", "MLE α"),
        ("", None, None),
        ("Weibull (k=2, λ=1)", None, None),
        ("ACT_WEIBULL_PDF", "=ACT_WEIBULL_PDF(0.5, 2, 1)", "f(0.5)"),
        ("ACT_WEIBULL_CDF", "=ACT_WEIBULL_CDF(0.5, 2, 1)", "F(0.5)"),
        ("ACT_WEIBULL_INV", "=ACT_WEIBULL_INV(0.5, 2, 1)", "Median"),
        ("ACT_WEIBULL_FIT", "=ACT_WEIBULL_FIT({0.5,1,1.5,2,2.5})", "MLE k,λ"),
        ("", None, None),
        ("Beta (α=2, β=5)", None, None),
        ("ACT_BETA_PDF", "=ACT_BETA_PDF(0.3, 2, 5)", "f(0.3)"),
        ("ACT_BETA_CDF", "=ACT_BETA_CDF(0.3, 2, 5)", "F(0.3)"),
        ("ACT_BETA_INV", "=ACT_BETA_INV(0.5, 2, 5)", "Median"),
        ("ACT_BETA_FIT", "=ACT_BETA_FIT({0.1,0.2,0.3,0.4})", "MLE α,β"),
        ("", None, None),
        ("Exponential (λ=0.5)", None, None),
        ("ACT_EXP_PDF", "=ACT_EXP_PDF(1, 0.5)", "f(1)"),
        ("ACT_EXP_CDF", "=ACT_EXP_CDF(1, 0.5)", "F(1)"),
        ("ACT_EXP_INV", "=ACT_EXP_INV(0.5, 0.5)", "Median"),
        ("ACT_EXP_FIT", "=ACT_EXP_FIT({1,2,3,4,5})", "MLE λ"),
        ("", None, None),
        ("GPD (ξ=0.5, σ=1)", None, None),
        ("ACT_GPD_PDF", "=ACT_GPD_PDF(1, 0.5, 1)", "f(1)"),
        ("ACT_GPD_CDF", "=ACT_GPD_CDF(1, 0.5, 1)", "F(1)"),
        ("ACT_GPD_INV", "=ACT_GPD_INV(0.5, 0.5, 1)", "Median"),
        ("ACT_GPD_FIT", "=ACT_GPD_FIT({1,2,3,4,5})", "MLE ξ,σ"),
        ("", None, None),
        ("Burr XII (c=2, k=3, λ=1)", None, None),
        ("ACT_BURR_PDF", "=ACT_BURR_PDF(1, 2, 3, 1)", "f(1)"),
        ("ACT_BURR_CDF", "=ACT_BURR_CDF(1, 2, 3, 1)", "F(1)"),
        ("ACT_BURR_INV", "=ACT_BURR_INV(0.5, 2, 3, 1)", "Median"),
        ("ACT_BURR_FIT", "=ACT_BURR_FIT({0.5,1,1.5,2,3})", "MLE c,k,λ"),
        ("", None, None),

        # ===== EXPOSURE CURVES =====
        ("EXPOSURE CURVES", None, None),
        ("", None, None),
        ("ACT_MBBEFD", "=ACT_MBBEFD(0.5, 2, 3)", "MBBEFD at d=0.5, b=2, g=3"),
        ("ACT_SWISSRE_CURVE", "=ACT_SWISSRE_CURVE(0.5, 1)", "Swiss Re curve 1 (light)"),
        ("ACT_SWISSRE_CURVE", "=ACT_SWISSRE_CURVE(0.5, 3)", "Swiss Re curve 3 (medium)"),
        ("ACT_SWISSRE_CURVE", "=ACT_SWISSRE_CURVE(0.5, 5)", "Swiss Re curve 5 (heavy)"),
        ("ACT_LLOYDS_CURVE", "=ACT_LLOYDS_CURVE(0.5, 1)", "Lloyd's Y1 curve"),
        ("ACT_LLOYDS_CURVE", "=ACT_LLOYDS_CURVE(0.5, 2)", "Lloyd's Y2 curve"),
        ("ACT_LLOYDS_CURVE", "=ACT_LLOYDS_CURVE(0.5, 3)", "Lloyd's Y3 curve"),
        ("ACT_LLOYDS_CURVE", "=ACT_LLOYDS_CURVE(0.5, 4)", "Lloyd's Y4 curve"),
        ("ACT_POWER_CURVE", "=ACT_POWER_CURVE(0.5, 2)", "Power curve n=2"),
        ("ACT_INVERSE_POWER_CURVE", "=ACT_INVERSE_POWER_CURVE(0.5, 2)", "Inverse power n=2"),
        ("ACT_PARETO_EXPOSURE", "=ACT_PARETO_EXPOSURE(0.5, 2)", "Pareto exposure α=2"),
        ("ACT_RIEBESELL_CURVE", "=ACT_RIEBESELL_CURVE(0.5, 0.8)", "Riebesell c=0.8"),
        ("ACT_RIEBESELL_CURVE_INV", "=ACT_RIEBESELL_CURVE_INV(0.7, 0.8)", "Inverse Riebesell"),
        ("", None, None),

        # ===== REINSURANCE =====
        ("REINSURANCE", None, None),
        ("", None, None),
        ("ACT_XOL_LAYER_LOSS", "=ACT_XOL_LAYER_LOSS(5000000, 1000000, 4000000)", "Layer loss: 5M ground-up, 4M xs 1M"),
        ("ACT_XOL_EXPECTED_LOSS", "=ACT_XOL_EXPECTED_LOSS(1000000, 4000000, 2, 100000, 0.1)", "Expected layer loss"),
        ("ACT_QS_CEDED", "=ACT_QS_CEDED(1000000, 0.5)", "50% quota share"),
        ("ACT_AGGREGATE_LAYER", "=ACT_AGGREGATE_LAYER(15000000, 2000000, 10000000)", "Aggregate: 10M xs 2M AAD"),
        ("ACT_ILF_PARETO", "=ACT_ILF_PARETO(2000000, 1000000, 2)", "ILF at 2M, base 1M, α=2"),
        ("ACT_LAYER_RATE_ON_LINE", "=ACT_LAYER_RATE_ON_LINE(100000, 1000000, 4000000)", "Rate on line"),
        ("", None, None),
        ("Return Period Functions", None, None),
        ("ACT_RETURN_PERIOD_LOSS", "=ACT_RETURN_PERIOD_LOSS({100,50,25,10,5}, {1000000,2000000,3000000,5000000,10000000}, 20, \"LOG\")", "20-year loss"),
        ("ACT_RETURN_PERIOD_TABLE", "=ACT_RETURN_PERIOD_TABLE({100,50,25,10,5}, {1000000,2000000,3000000,5000000,10000000}, {10,25,50,100})", "RP table (array)"),
        ("ACT_AAL_FROM_OEP", "=ACT_AAL_FROM_OEP({100,50,25,10,5}, {1000000,2000000,3000000,5000000,10000000})", "AAL from OEP curve"),
        ("", None, None),

        # ===== INTERPOLATION =====
        ("INTERPOLATION", None, None),
        ("", None, None),
        ("ACT_INTERP", "=ACT_INTERP({1,2,3,4,5}, {10,20,35,55,80}, 2.5, \"LINEAR\")", "Linear interpolation"),
        ("ACT_INTERP", "=ACT_INTERP({1,2,3,4,5}, {10,20,35,55,80}, 0.5, \"FLAT\")", "Flat extrapolation"),
        ("ACT_INTERP", "=ACT_INTERP({1,2,3,4,5}, {10,20,35,55,80}, 6, \"GRADIENT\")", "Gradient extrapolation"),
        ("ACT_INTERP_LOG", "=ACT_INTERP_LOG({1,2,3,4,5}, {10,20,35,55,80}, 2.5)", "Log-linear interpolation"),
        ("ACT_INTERP2D", "=ACT_INTERP2D({1,2,3}, {10,20,30}, {100,200,300;110,220,330;120,240,360}, 1.5, 15)", "2D bilinear interpolation"),
        ("", None, None),

        # ===== CREDIBILITY =====
        ("CREDIBILITY - Bühlmann", None, None),
        ("", None, None),
        ("ACT_CREDIBILITY_BUHLMANN", "=ACT_CREDIBILITY_BUHLMANN(10, 5)", "Z = n/(n+k) = 10/15"),
        ("ACT_CREDIBILITY_ESTIMATE", "=ACT_CREDIBILITY_ESTIMATE(10, 5, 0.05, 0.03)", "Blended estimate"),
        ("ACT_CREDIBILITY_K", "=ACT_CREDIBILITY_K({0.05,0.06,0.04}, {100,150,120})", "Estimate k from data"),
        ("ACT_CREDIBILITY_BUHLMANN_STRAUB", "=ACT_CREDIBILITY_BUHLMANN_STRAUB(1000, 500)", "B-S credibility"),
        ("ACT_BUHLMANN_STRAUB_PARAMS", "=ACT_BUHLMANN_STRAUB_PARAMS({0.05,0.06,0.04;0.055,0.058,0.045}, {100,150,120;110,140,130})", "B-S parameters (array)"),
        ("", None, None),
        ("CREDIBILITY - Experience Rating", None, None),
        ("ACT_EXPERIENCE_MOD", "=ACT_EXPERIENCE_MOD(0.7, 50000, 70000)", "Experience mod factor"),
        ("ACT_EXPERIENCE_MOD_SPLIT", "=ACT_EXPERIENCE_MOD_SPLIT(0.7, 50000, 70000, 25000, 0.3)", "Split plan exp mod"),
        ("", None, None),
        ("CREDIBILITY - Full/Partial", None, None),
        ("ACT_FULL_CREDIBILITY_STANDARD", "=ACT_FULL_CREDIBILITY_STANDARD(0.05, 0.90, 1.5)", "Full cred standard n₀"),
        ("ACT_PARTIAL_CREDIBILITY", "=ACT_PARTIAL_CREDIBILITY(500, 1082)", "Partial cred (sqrt rule)"),
        ("ACT_LINEAR_CREDIBILITY", "=ACT_LINEAR_CREDIBILITY(500, 1082)", "Linear credibility"),
        ("ACT_ASYMPTOTIC_CREDIBILITY", "=ACT_ASYMPTOTIC_CREDIBILITY(10, 5, 0.5)", "Asymptotic cred"),
        ("", None, None),
        ("CREDIBILITY - ILF", None, None),
        ("ACT_ILF_LAYER", "=ACT_ILF_LAYER(1000000, 4000000, 100000, 2)", "Layer ILF"),
        ("ACT_ILF_TABLE", "=ACT_ILF_TABLE(100000, {250000,500000,1000000,2000000}, 2)", "ILF table (array)"),
        ("ACT_ILF_TABLE_STANDARD", "=ACT_ILF_TABLE_STANDARD(100000, 5000000, 2)", "Standard ILF table (array)"),
        ("", None, None),
        ("CREDIBILITY - Deductibles", None, None),
        ("ACT_LOSS_ELIMINATION_RATIO", "=ACT_LOSS_ELIMINATION_RATIO(10000, 2, 1000)", "LER Pareto"),
        ("ACT_LOSS_ELIMINATION_RATIO_LOGNORMAL", "=ACT_LOSS_ELIMINATION_RATIO_LOGNORMAL(10000, 10, 1)", "LER Lognormal"),
        ("ACT_DEDUCTIBLE_CREDIT", "=ACT_DEDUCTIBLE_CREDIT(10000, 2, 1000)", "Deductible credit"),
        ("ACT_PREMIUM_WITH_DEDUCTIBLE", "=ACT_PREMIUM_WITH_DEDUCTIBLE(100000, 10000, 2, 1000)", "Premium with deductible"),
        ("", None, None),
        ("CREDIBILITY - Loading", None, None),
        ("ACT_LARGE_LOSS_LOAD", "=ACT_LARGE_LOSS_LOAD(0.1, 500000, 100000)", "Large loss loading"),
        ("ACT_SHOCK_LOSS_LOAD", "=ACT_SHOCK_LOSS_LOAD(0.01, 10000000, 100000)", "Shock loss loading"),
        ("", None, None),
        ("CREDIBILITY - Burning Cost", None, None),
        ("ACT_BURNING_COST", "=ACT_BURNING_COST({50000,60000,40000}, {1000000,1100000,1050000})", "Burning cost"),
        ("ACT_CREDIBILITY_RATE", "=ACT_CREDIBILITY_RATE(0.05, 0.03, 0.6)", "Credibility-weighted rate"),
        ("ACT_COMPOSITE_RATE", "=ACT_COMPOSITE_RATE({0.05,0.06,0.04}, {0.5,0.3,0.2})", "Composite rate"),
        ("ACT_MULTIYEAR_CREDIBILITY", "=ACT_MULTIYEAR_CREDIBILITY(3, 100, 0.5)", "Multi-year credibility"),
        ("", None, None),
        ("CREDIBILITY - Schedule Rating", None, None),
        ("ACT_SCHEDULE_RATING", "=ACT_SCHEDULE_RATING(100000, {0.1,-0.05,0.15})", "Schedule rating"),
        ("ACT_SCHEDULE_RATING_CAPPED", "=ACT_SCHEDULE_RATING_CAPPED(100000, {0.1,-0.05,0.15}, {0.15,0.1,0.2})", "Capped schedule rating"),
        ("", None, None),
        ("CREDIBILITY - Retrospective", None, None),
        ("ACT_RETRO_PREMIUM", "=ACT_RETRO_PREMIUM(100000, 50000, 0.3, 1.1, 0.6, 1.2, 1.03)", "Retro premium"),
        ("ACT_RETRO_PARAMETERS", "=ACT_RETRO_PARAMETERS(0.6, 1.2, 0.05, 1.1)", "Retro parameters (array)"),
        ("ACT_RETRO_LOSS_LIMIT_CREDIT", "=ACT_RETRO_LOSS_LIMIT_CREDIT(250000, 100000, 2, 1000)", "Loss limit credit"),
        ("", None, None),
        ("CREDIBILITY - Bayesian", None, None),
        ("ACT_BAYESIAN_UPDATE", "=ACT_BAYESIAN_UPDATE(0.05, 0.01, 0.06, 0.005)", "Bayesian posterior mean"),
        ("ACT_BAYESIAN_UPDATE_FULL", "=ACT_BAYESIAN_UPDATE_FULL(0.05, 0.01, 0.06, 0.005)", "Posterior mean & var (array)"),
        ("", None, None),
        ("CREDIBILITY - Poisson", None, None),
        ("ACT_CREDIBILITY_POISSON", "=ACT_CREDIBILITY_POISSON(10, 5, 2)", "Poisson credibility"),
        ("ACT_FREQUENCY_ESTIMATE", "=ACT_FREQUENCY_ESTIMATE(10, 1000, 5, 2, 0.005)", "Credibility frequency est"),
        ("", None, None),

        # ===== COPULAS =====
        ("COPULAS - Gaussian", None, None),
        ("", None, None),
        ("ACT_COPULA_GAUSSIAN", "=ACT_COPULA_GAUSSIAN({1,0.5;0.5,1}, 5, 42)", "Gaussian copula samples (array)"),
        ("ACT_COPULA_GAUSSIAN_SINGLE", "=ACT_COPULA_GAUSSIAN_SINGLE({1,0.5;0.5,1}, 42)", "Single Gaussian sample"),
        ("", None, None),
        ("COPULAS - Student-t", None, None),
        ("ACT_COPULA_STUDENT_T", "=ACT_COPULA_STUDENT_T({1,0.5;0.5,1}, 5, 5, 42)", "Student-t copula (array)"),
        ("ACT_COPULA_STUDENT_T_SINGLE", "=ACT_COPULA_STUDENT_T_SINGLE({1,0.5;0.5,1}, 5, 42)", "Single Student-t sample"),
        ("", None, None),
        ("COPULAS - Clayton (lower tail)", None, None),
        ("ACT_COPULA_CLAYTON", "=ACT_COPULA_CLAYTON(2, 5, 42)", "Clayton copula (array)"),
        ("ACT_COPULA_CLAYTON_SINGLE", "=ACT_COPULA_CLAYTON_SINGLE(2, 42)", "Single Clayton sample"),
        ("ACT_COPULA_CLAYTON_CDF", "=ACT_COPULA_CLAYTON_CDF(0.5, 0.5, 2)", "Clayton CDF"),
        ("", None, None),
        ("COPULAS - Frank (symmetric)", None, None),
        ("ACT_COPULA_FRANK", "=ACT_COPULA_FRANK(5, 5, 42)", "Frank copula (array)"),
        ("ACT_COPULA_FRANK_SINGLE", "=ACT_COPULA_FRANK_SINGLE(5, 42)", "Single Frank sample"),
        ("ACT_COPULA_FRANK_CDF", "=ACT_COPULA_FRANK_CDF(0.5, 0.5, 5)", "Frank CDF"),
        ("", None, None),
        ("COPULAS - Gumbel (upper tail)", None, None),
        ("ACT_COPULA_GUMBEL", "=ACT_COPULA_GUMBEL(2, 5, 42)", "Gumbel copula (array)"),
        ("ACT_COPULA_GUMBEL_SINGLE", "=ACT_COPULA_GUMBEL_SINGLE(2, 42)", "Single Gumbel sample"),
        ("ACT_COPULA_GUMBEL_CDF", "=ACT_COPULA_GUMBEL_CDF(0.5, 0.5, 2)", "Gumbel CDF"),
        ("", None, None),
        ("COPULAS - Utilities", None, None),
        ("ACT_COPULA_TAU_TO_THETA", '=ACT_COPULA_TAU_TO_THETA(0.5, "CLAYTON")', "Kendall τ → θ"),
        ("ACT_COPULA_TAIL_LOWER", '=ACT_COPULA_TAIL_LOWER("CLAYTON", 2)', "Lower tail dependence"),
        ("ACT_COPULA_TAIL_UPPER", '=ACT_COPULA_TAIL_UPPER("GUMBEL", 2)', "Upper tail dependence"),
        ("ACT_COPULA_TAIL_LOWER", '=ACT_COPULA_TAIL_LOWER("STUDENT_T", 5, 0.5)', "Student-t lower tail"),
        ("ACT_COPULA_TAIL_UPPER", '=ACT_COPULA_TAIL_UPPER("STUDENT_T", 5, 0.5)', "Student-t upper tail"),
        ("", None, None),

        # ===== CHAIN LADDER =====
        ("CHAIN LADDER - See dedicated sheet", None, None),
        ("", None, None),
        ("(Chain ladder functions tested in 'Chain Ladder' sheet with triangle data)", None, None),
        ("ACT_CL_FACTORS - Development factors (array)", None, None),
        ("ACT_CL_LATEST - Latest diagonal (array)", None, None),
        ("ACT_CL_ULTIMATE - Projected ultimates (array)", None, None),
        ("ACT_CL_IBNR - IBNR reserves (array)", None, None),
        ("ACT_BF_ULTIMATE - Bornhuetter-Ferguson (array)", None, None),
        ("ACT_MACK_FACTOR_SE - Factor std errors (array)", None, None),
        ("ACT_MACK_RESERVE_SE - Reserve std errors (array)", None, None),
        ("ACT_CL_BOOTSTRAP - Bootstrap total (array)", None, None),
        ("ACT_CL_BOOTSTRAP_ORIGIN - Bootstrap by AY (array)", None, None),
        ("ACT_CAPECOD_ULTIMATE - Cape Cod (array)", None, None),
        ("ACT_CAPECOD_ELR - Cape Cod ELR", None, None),
        ("ACT_BERQUIST_SHERMAN - Berquist-Sherman (array)", None, None),
        ("ACT_TRIANGLE_TO_INCREMENTAL - Cum→Inc (array)", None, None),
        ("ACT_INCREMENTAL_TO_CUMULATIVE - Inc→Cum (array)", None, None),
        ("ACT_TRIANGLE_DIAGONAL - Extract diagonal (array)", None, None),
        ("ACT_TRIANGLE_LINK_RATIOS - Link ratios (array)", None, None),
        ("ACT_CL_CALENDAR_ADJUST - Calendar adjust (array)", None, None),
        ("ACT_CL_CALENDAR_TOTALS - Calendar totals (array)", None, None),
        ("ACT_CL_WEIGHTED_AVERAGE - Weighted avg (array)", None, None),
        ("", None, None),

        # ===== VERSION =====
        ("VERSION INFO", None, None),
        ("", None, None),
        ("ACT_VERSION", "=ACT_VERSION()", "Current version"),
        ("ACT_BUILD_DATE", "=ACT_BUILD_DATE()", "Build date"),
        ("ACT_GITHUB_URL", "=ACT_GITHUB_URL()", "GitHub URL"),
        ("ACT_COMMIT_COUNT", "=ACT_COMMIT_COUNT()", "Number of commits"),
        ("ACT_COMMIT_INFO", '=ACT_COMMIT_INFO(1, "hash")', "Latest commit hash"),
        ("ACT_COMMIT_INFO", '=ACT_COMMIT_INFO(1, "date")', "Latest commit date"),
        ("ACT_COMMIT_INFO", '=ACT_COMMIT_INFO(1, "message")', "Latest commit message"),
        ("ACT_COMMIT_HISTORY", "=ACT_COMMIT_HISTORY()", "Full history (array)"),
    ]

    for name, formula, description in all_tests:
        if formula is None and description is None:
            # Section header
            if name:
                ws.cell(row=row, column=1, value=name).font = HEADER_FONT
        else:
            ws.cell(row=row, column=1, value=name).font = BOLD_FONT
            if formula:
                ws.cell(row=row, column=2, value=formula)
            if description:
                ws.cell(row=row, column=3, value=description)
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 100
    ws.column_dimensions['C'].width = 30

    return ws


def fix_chain_ladder_sheet(wb):
    """Ensure Chain Ladder sheet has proper array formula areas."""
    if 'Chain Ladder' not in wb.sheetnames:
        return

    ws = wb['Chain Ladder']

    # Clear cells to the right of ACT_CL_FACTORS at B21 (needs to spill to J21)
    for col in range(3, 12):
        cell = ws.cell(row=21, column=col)
        if cell.value is None or (isinstance(cell.value, (int, float)) and cell.value == 0):
            pass  # Already empty
        else:
            cell.value = None

    # Same for B41
    for col in range(3, 12):
        cell = ws.cell(row=41, column=col)
        if cell.value is None or (isinstance(cell.value, (int, float)) and cell.value == 0):
            pass
        else:
            cell.value = None

    # Clear area for ACT_BOOTSTRAP_CL at A89
    for r in range(89, 98):
        for c in range(1, 4):
            if r == 89 and c == 1:
                continue
            ws.cell(row=r, column=c).value = None

    # Clear area for ACT_BOOTSTRAP_CL_ORIGIN at A92 - but it's at row 92
    # Actually need more room - this returns 11 rows x 8 cols
    for r in range(92, 105):
        for c in range(1, 10):
            if r == 92 and c == 1:
                continue
            ws.cell(row=r, column=c).value = None

    print("Fixed Chain Ladder array formula areas")


def main():
    project_root = Path(__file__).parent.parent
    source_file = project_root / "excel" / "actuarial_add_in_v0.0.xlsm"
    target_file = project_root / "excel" / "actuarial_add_in_v0.1.1.xlsm"

    # Remove target if exists (in case of previous failed run)
    if target_file.exists():
        try:
            target_file.unlink()
        except PermissionError:
            print(f"ERROR: {target_file} is open. Please close Excel and try again.")
            return

    # Copy original
    shutil.copy2(source_file, target_file)
    print(f"Copied {source_file.name}")

    # Load
    wb = load_workbook(target_file, keep_vba=True)

    # Create proper Versions sheet with commit history
    create_versions_sheet(wb)
    print("Created Versions sheet with commit history")

    # Create comprehensive All Functions Test sheet
    create_all_functions_sheet(wb)
    print("Created All Functions Test sheet with ALL functions")

    # Fix Chain Ladder sheet
    fix_chain_ladder_sheet(wb)

    # Remove old Array Formula Demo if exists
    if "Array Formula Demo" in wb.sheetnames:
        del wb["Array Formula Demo"]

    # Save
    wb.save(target_file)
    print(f"\nSaved: {target_file}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
