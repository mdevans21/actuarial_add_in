#!/usr/bin/env python3
"""
Update the spreadsheet with missing function examples and fix formula issues.
"""

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import re

def fix_at_formulas(ws):
    """Remove @ prefix from array formulas that shouldn't have it."""
    fixed = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                if '@ACT_' in cell.value or '=@' in cell.value:
                    # Remove @ prefix before ACT_ functions
                    new_formula = re.sub(r'@(ACT_[A-Z_]+)', r'\1', cell.value)
                    new_formula = new_formula.replace('=@', '=')
                    if new_formula != cell.value:
                        cell.value = new_formula
                        fixed += 1
    return fixed

def add_chain_ladder_examples(wb):
    """Add examples for ACT_CL_LATEST and ACT_BOOTSTRAP_CL_ORIGIN."""
    ws = wb['Chain Ladder']

    # Find the last used row
    max_row = ws.max_row

    # Add section for new functions
    start_row = max_row + 3

    # Header
    ws.cell(row=start_row, column=1, value="Additional Chain Ladder Functions")
    ws.cell(row=start_row, column=1).font = Font(bold=True, size=12)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=5)
    start_row += 2

    # ACT_CL_LATEST example
    ws.cell(row=start_row, column=1, value="ACT_CL_LATEST")
    ws.cell(row=start_row, column=1).font = Font(bold=True)
    ws.cell(row=start_row, column=2, value="Returns the latest diagonal values")
    start_row += 1

    # Note: We reference the existing triangle in the sheet (typically at B5:F9)
    ws.cell(row=start_row, column=1, value="Latest Values:")
    # Use the triangle range that exists in the sheet
    ws.cell(row=start_row, column=2, value="=ACT_CL_LATEST(B5:F9)")
    start_row += 2

    # ACT_BOOTSTRAP_CL_ORIGIN example
    ws.cell(row=start_row, column=1, value="ACT_BOOTSTRAP_CL_ORIGIN")
    ws.cell(row=start_row, column=1).font = Font(bold=True)
    ws.cell(row=start_row, column=2, value="Bootstrap reserves by origin year")
    start_row += 1

    ws.cell(row=start_row, column=1, value="Origin Year Stats:")
    ws.cell(row=start_row, column=2, value="=ACT_BOOTSTRAP_CL_ORIGIN(B5:F9, 1000, 42)")
    start_row += 1

    return start_row

def main():
    project_root = Path(__file__).parent.parent
    source_file = project_root / "excel" / "actuarial_add_in_v0.1.xlsm"

    if not source_file.exists():
        source_file = project_root / "excel" / "actuarial_add_in_v0.0.xlsm"

    target_file = project_root / "excel" / "actuarial_add_in_v0.1.xlsm"

    wb = load_workbook(source_file, keep_vba=True)

    # Fix @ formulas in all sheets
    total_fixed = 0
    for sheet_name in wb.sheetnames:
        if sheet_name != "Versions":
            fixed = fix_at_formulas(wb[sheet_name])
            if fixed > 0:
                print(f"Fixed {fixed} @-prefixed formulas in {sheet_name}")
                total_fixed += fixed

    # Add chain ladder examples if sheet exists
    if 'Chain Ladder' in wb.sheetnames:
        add_chain_ladder_examples(wb)
        print("Added ACT_CL_LATEST and ACT_BOOTSTRAP_CL_ORIGIN examples")

    # Ensure Versions sheet exists and is first
    if "Versions" not in wb.sheetnames:
        from create_versioned_spreadsheet import create_versions_sheet
        create_versions_sheet(wb)
        print("Created Versions sheet")

    wb.save(target_file)
    print(f"\nSaved: {target_file}")
    print(f"Total @-prefix formulas fixed: {total_fixed}")

if __name__ == "__main__":
    main()
