"""
Add LEV column to Distributions tab in Excel workbook.
Columns: x, PDF, CDF, LEV, Notes
Sets uniform column widths.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font

def add_lev_to_distributions():
    wb = load_workbook('excel/actuarial_add_in.xlsm', keep_vba=True)
    ws = wb['Distributions']

    bold_font = Font(bold=True)

    # Distribution configurations: (header_row, lev_function, params)
    distributions = [
        (20, 'ACT_DIST_POISSON_LEV', '5'),           # Poisson (lambda=5)
        (35, 'ACT_DIST_NEGBIN_LEV', '5, 0.3'),       # NegBin (r=5, p=0.3)
        (50, 'ACT_DIST_LOGNORM_LEV', '0, 1'),        # Lognormal (mu=0, sigma=1)
        (65, 'ACT_DIST_GAMMA_LEV', '2, 1'),          # Gamma (alpha=2, beta=1)
        (80, 'ACT_DIST_EXP_LEV', '1'),               # Exponential (lambda=1)
        (95, 'ACT_DIST_WEIBULL_LEV', '2, 1'),        # Weibull (k=2, lambda=1)
        (110, 'ACT_DIST_BETA_LEV', '2, 5'),          # Beta (alpha=2, beta=5)
        (125, 'ACT_DIST_PARETO_LEV', '2, 1'),        # Pareto (alpha=2, xm=1)
        (140, 'ACT_DIST_LOMAX_LEV', '2, 1'),         # Lomax (alpha=2, lambda=1)
        (155, 'ACT_DIST_GPD_LEV', '0.5, 1'),         # GPD (xi=0.5, sigma=1)
        (170, 'ACT_DIST_BURR_LEV', '2, 1, 1'),       # Burr (c=2, k=1, lambda=1)
    ]

    for header_row, lev_func, params in distributions:
        col_header_row = header_row + 1

        # First pass: save existing Notes values from column D
        notes_values = {}
        for data_row in range(header_row + 2, header_row + 13):
            notes_val = ws.cell(data_row, 4).value
            if notes_val:
                notes_values[data_row] = notes_val
            # Clear column E (in case there's old data)
            ws.cell(data_row, 5, None)

        # Update column headers: x, PDF, CDF, LEV, Notes
        ws.cell(col_header_row, 4).value = 'LEV'
        ws.cell(col_header_row, 4).font = bold_font
        ws.cell(col_header_row, 5).value = 'Notes'
        ws.cell(col_header_row, 5).font = bold_font

        # Second pass: add LEV formulas in column D, move Notes to column E
        for data_row in range(header_row + 2, header_row + 13):
            x_val = ws.cell(data_row, 1).value
            if x_val is not None and x_val != '':
                x_ref = f'A{data_row}'
                formula = f'={lev_func}({x_ref}, {params})'
                ws.cell(data_row, 4, formula)
            else:
                ws.cell(data_row, 4, None)

            # Restore notes to column E
            if data_row in notes_values:
                ws.cell(data_row, 5, notes_values[data_row])

    # Set uniform column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 8

    wb.save('excel/actuarial_add_in.xlsm')
    print(f'Reformatted {len(distributions)} distribution sections')

if __name__ == "__main__":
    add_lev_to_distributions()
