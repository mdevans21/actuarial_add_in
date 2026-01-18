#!/usr/bin/env python3
"""
Create a new version of the spreadsheet with a Versions tab.
"""

import shutil
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Commit history
COMMITS = [
    ("b2a74d4", "2026-01-18", "Add Phase 6: Credibility and experience rating functions"),
    ("8e5109b", "2026-01-18", "Add Phase 5: Reserving enhancements - Cape Cod, triangle utilities, calendar year"),
    ("6a2d830", "2026-01-18", "Add Phase 3: Archimedean copulas and utility functions"),
    ("2f71e98", "2026-01-18", "Add Phase 2: Parameter estimation functions for distribution fitting"),
    ("79ef7bc", "2026-01-18", "Implement Phase 1 from next_steps.md: distributions, copulas, and help text"),
    ("bc30112", "2026-01-18", "Add comprehensive review and roadmap for actuarial add-in"),
    ("fb0a7fe", "2026-01-17", "Merge branch 'main' of https://github.com/mdevans21/actuarial_add_in"),
    ("83a9008", "2026-01-17", "Add chain ladder enhancements and update examples"),
    ("b16e651", "2026-01-17", "Update README.md"),
    ("7bb88e5", "2026-01-17", "Update README.md"),
    ("7458388", "2026-01-17", "Add charts to Excel sheets and update Chain Ladder with Taylor-Ashe data"),
    ("dc52de8", "2026-01-17", "Add documentation, examples, and agent instructions"),
    ("1fbd1ef", "2026-01-17", "Rename PRD file to fix double extension"),
    ("d2323be", "2026-01-17", "Fix Poisson inverse CDF implementation"),
    ("62c42d7", "2026-01-17", "Initial implementation of Actuarial Excel Add-in"),
]

VERSION = "0.1.0"
BUILD_DATE = "2026-01-18"
GITHUB_URL = "https://github.com/mdevans21/actuarial_add_in"


def create_versions_sheet(wb):
    """Create and populate the Versions sheet."""
    # Create sheet at the beginning
    ws = wb.create_sheet("Versions", 0)

    # Styles
    title_font = Font(bold=True, size=16)
    header_font = Font(bold=True, size=14)
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    row = 1

    # Title
    ws.cell(row=row, column=1, value="Actuarial Add-In Version History")
    ws.cell(row=row, column=1).font = title_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 2

    # Version info
    ws.cell(row=row, column=1, value="Current Version:").font = bold_font
    ws.cell(row=row, column=2, value=VERSION)
    row += 1

    ws.cell(row=row, column=1, value="Build Date:").font = bold_font
    ws.cell(row=row, column=2, value=BUILD_DATE)
    row += 1

    ws.cell(row=row, column=1, value="GitHub:").font = bold_font
    ws.cell(row=row, column=2, value=GITHUB_URL)
    ws.cell(row=row, column=2).hyperlink = GITHUB_URL
    ws.cell(row=row, column=2).style = "Hyperlink"
    row += 2

    # Note about @ formulas
    ws.cell(row=row, column=1, value="Note on Array Formulas")
    ws.cell(row=row, column=1).font = header_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1

    ws.cell(row=row, column=1, value="If you see @ before function names, this is Excel's implicit intersection operator.")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1

    ws.cell(row=row, column=1, value="For array functions, either:")
    row += 1
    ws.cell(row=row, column=1, value="  1. Select output range and press Ctrl+Shift+Enter")
    row += 1
    ws.cell(row=row, column=1, value="  2. Use INDEX() to extract individual values")
    row += 1
    ws.cell(row=row, column=1, value="  3. Let the formula spill into adjacent cells (Excel 365)")
    row += 2

    # Commit history header
    ws.cell(row=row, column=1, value="Commit History")
    ws.cell(row=row, column=1).font = header_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1

    # Column headers
    headers = ["Commit", "Date", "Message"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = thin_border
    row += 1

    # Commit data
    for commit_hash, date, message in COMMITS:
        ws.cell(row=row, column=1, value=commit_hash).border = thin_border
        ws.cell(row=row, column=2, value=date).border = thin_border
        ws.cell(row=row, column=3, value=message).border = thin_border
        row += 1

    # Auto-fit columns
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 70

    return ws


def main():
    project_root = Path(__file__).parent.parent
    source_file = project_root / "excel" / "actuarial_add_in_v0.0.xlsm"
    target_file = project_root / "excel" / "actuarial_add_in_v0.1.xlsm"

    if not source_file.exists():
        print(f"Error: Source file not found: {source_file}")
        return 1

    # Copy the file first (to preserve macros)
    shutil.copy2(source_file, target_file)
    print(f"Copied {source_file.name} to {target_file.name}")

    # Open and modify
    wb = load_workbook(target_file, keep_vba=True)

    # Check if Versions sheet already exists
    if "Versions" in wb.sheetnames:
        del wb["Versions"]
        print("Removed existing Versions sheet")

    # Create new Versions sheet
    create_versions_sheet(wb)
    print("Created Versions sheet")

    # Save
    wb.save(target_file)
    print(f"Saved: {target_file}")

    return 0


if __name__ == "__main__":
    exit(main())
