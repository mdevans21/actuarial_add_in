#!/usr/bin/env python3
"""Verify the v0.1 spreadsheet."""

from pathlib import Path
from openpyxl import load_workbook

def main():
    project_root = Path(__file__).parent.parent
    target_file = project_root / "excel" / "actuarial_add_in_v0.1.xlsm"

    wb = load_workbook(target_file, data_only=False, keep_vba=True)

    print(f"Sheets: {wb.sheetnames}")
    print(f"Total sheets: {len(wb.sheetnames)}")
    print()

    # Check Versions sheet
    if "Versions" in wb.sheetnames:
        ws = wb["Versions"]
        print("Versions sheet contents (first 20 rows):")
        for row_num, row in enumerate(ws.iter_rows(max_row=20, values_only=True), 1):
            values = [str(v) if v else "" for v in row[:3]]
            print(f"  {row_num}: {values}")
    else:
        print("WARNING: Versions sheet not found!")

if __name__ == "__main__":
    main()
