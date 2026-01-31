#!/usr/bin/env python3
"""
Fix formatting issues in the All Functions Test sheet:
1. Add blank row before Normal Distribution section
2. Make subheadings bold (Normal, Lomax, Fackler)
"""

import openpyxl
from openpyxl.styles import Font
import os

EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', 'excel', 'actuarial_add_in.xlsm')

def fix_all_functions_formatting():
    print(f"Opening workbook: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, keep_vba=True)

    if "All Functions Test" not in wb.sheetnames:
        print("Error: 'All Functions Test' sheet not found")
        return

    ws = wb["All Functions Test"]

    # Find the rows containing our target subheadings
    normal_row = None
    lomax_row = None
    fackler_row = None

    for row_num in range(1, 150):
        cell_val = ws.cell(row=row_num, column=1).value
        if cell_val:
            if "Normal Distribution" in str(cell_val) and "μ" in str(cell_val):
                normal_row = row_num
            elif "Lomax" in str(cell_val) and "Pareto II" in str(cell_val):
                lomax_row = row_num
            elif "Fackler Composite" in str(cell_val):
                fackler_row = row_num

    print(f"Found Normal Distribution at row {normal_row}")
    print(f"Found Lomax / Pareto II at row {lomax_row}")
    print(f"Found Fackler Composite at row {fackler_row}")

    if not all([normal_row, lomax_row, fackler_row]):
        print("Error: Could not find all subheading rows")
        return

    # Check if there's already a blank row before Normal Distribution
    prev_row_val = ws.cell(row=normal_row - 1, column=1).value
    need_blank_row = prev_row_val is not None and str(prev_row_val).strip() != ""

    if need_blank_row:
        print(f"Inserting blank row before row {normal_row}")
        ws.insert_rows(normal_row)
        # After insertion, row numbers shift
        normal_row += 1
        lomax_row += 1
        fackler_row += 1
    else:
        print("Blank row already exists before Normal Distribution")

    # Apply bold formatting to subheadings
    subheading_font = Font(bold=True)

    print(f"Making row {normal_row} bold")
    ws.cell(row=normal_row, column=1).font = subheading_font

    print(f"Making row {lomax_row} bold")
    ws.cell(row=lomax_row, column=1).font = subheading_font

    print(f"Making row {fackler_row} bold")
    ws.cell(row=fackler_row, column=1).font = subheading_font

    # Save
    print("Saving workbook...")
    wb.save(EXCEL_PATH)
    print("Done!")

if __name__ == "__main__":
    fix_all_functions_formatting()
