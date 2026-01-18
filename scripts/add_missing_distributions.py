#!/usr/bin/env python3
"""
Add missing distributions to the Distributions tab.
Missing: GPD, Weibull, Beta, Exponential, Burr
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# Styles
TITLE_FONT = Font(bold=True, size=12)
HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def add_distribution_section(ws, start_row, title, params_desc, x_values, pdf_formula, cdf_formula, inv_example=None):
    """Add a distribution section with PDF and CDF examples."""
    row = start_row

    # Title
    ws.cell(row=row, column=1, value=title).font = TITLE_FONT
    row += 1

    # Headers
    headers = ["x", "PDF", "CDF"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    row += 1

    # Data rows
    for x_val in x_values:
        ws.cell(row=row, column=1, value=x_val).border = THIN_BORDER
        ws.cell(row=row, column=2, value=pdf_formula.format(x=f"A{row}")).border = THIN_BORDER
        ws.cell(row=row, column=3, value=cdf_formula.format(x=f"A{row}")).border = THIN_BORDER
        row += 1

    # INV example if provided
    if inv_example:
        ws.cell(row=row, column=1, value=inv_example["label"])
        ws.cell(row=row, column=2, value=inv_example["formula"])
        row += 1

    row += 1  # Empty row before next section
    return row


def main():
    # Load the workbook
    wb = load_workbook('excel/actuarial_add_in_v0.1.1.xlsm', keep_vba=True)
    ws = wb['Distributions']

    # Find the last used row
    start_row = ws.max_row + 2

    # 1. GPD Distribution
    start_row = add_distribution_section(
        ws, start_row,
        "GPD DISTRIBUTION (xi=0.5, sigma=1) - Extreme Value Theory",
        "xi=0.5 (heavy tail), sigma=1",
        [0.5, 1, 1.5, 2, 2.5, 3, 4, 5],
        "=ACT_GPD_PDF({x}, 0.5, 1)",
        "=ACT_GPD_CDF({x}, 0.5, 1)",
        {"label": "99th percentile (VaR):", "formula": "=ACT_GPD_INV(0.99, 0.5, 1)"}
    )

    # 2. Weibull Distribution
    start_row = add_distribution_section(
        ws, start_row,
        "WEIBULL DISTRIBUTION (k=1.5, lambda=2)",
        "k=1.5 (shape), lambda=2 (scale)",
        [0.5, 1, 1.5, 2, 2.5, 3, 4, 5],
        "=ACT_WEIBULL_PDF({x}, 1.5, 2)",
        "=ACT_WEIBULL_CDF({x}, 1.5, 2)",
        {"label": "Median:", "formula": "=ACT_WEIBULL_INV(0.5, 1.5, 2)"}
    )

    # 3. Beta Distribution
    start_row = add_distribution_section(
        ws, start_row,
        "BETA DISTRIBUTION (alpha=2, beta=5) - Loss Ratios",
        "alpha=2, beta=5 (right-skewed, mean=0.286)",
        [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8],
        "=ACT_BETA_PDF({x}, 2, 5)",
        "=ACT_BETA_CDF({x}, 2, 5)",
        {"label": "75th percentile:", "formula": "=ACT_BETA_INV(0.75, 2, 5)"}
    )

    # 4. Exponential Distribution
    start_row = add_distribution_section(
        ws, start_row,
        "EXPONENTIAL DISTRIBUTION (lambda=0.5) - Mean=2",
        "lambda=0.5 (rate), mean=1/lambda=2",
        [0.5, 1, 1.5, 2, 2.5, 3, 4, 5],
        "=ACT_EXP_PDF({x}, 0.5)",
        "=ACT_EXP_CDF({x}, 0.5)",
        {"label": "Median:", "formula": "=ACT_EXP_INV(0.5, 0.5)"}
    )

    # 5. Burr Distribution
    start_row = add_distribution_section(
        ws, start_row,
        "BURR TYPE XII DISTRIBUTION (c=2, k=3, lambda=1)",
        "c=2 (shape1), k=3 (shape2), lambda=1 (scale)",
        [0.5, 1, 1.5, 2, 2.5, 3, 4, 5],
        "=ACT_BURR_PDF({x}, 2, 3, 1)",
        "=ACT_BURR_CDF({x}, 2, 3, 1)",
        {"label": "95th percentile:", "formula": "=ACT_BURR_INV(0.95, 2, 3, 1)"}
    )

    # Auto-fit columns
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25

    # Save to new version
    output_file = 'excel/actuarial_add_in_v0.2.xlsm'
    wb.save(output_file)
    print(f"Saved to {output_file}")
    print("Added 5 missing distributions: GPD, Weibull, Beta, Exponential, Burr")
    print(f"Distributions tab now has {ws.max_row} rows")


if __name__ == "__main__":
    main()
