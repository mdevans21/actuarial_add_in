"""
Add LEV functions to All Functions Test tab in Excel workbook.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font

def add_lev_functions():
    wb = load_workbook('excel/actuarial_add_in.xlsm', keep_vba=True)
    ws = wb['All Functions Test']

    # Find the row with "EXPOSURE CURVES" to insert before it
    exposure_row = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == "EXPOSURE CURVES":
            exposure_row = row
            break

    if not exposure_row:
        print("Could not find EXPOSURE CURVES section")
        return

    print(f"Found EXPOSURE CURVES at row {exposure_row}")

    # LEV functions to add - using same parameters as other distribution tests
    lev_functions = [
        # Header
        ("LIMITED EXPECTED VALUE (LEV)", "", ""),
        ("", "", ""),
        # Frequency distributions
        ("Poisson LEV (λ=5, limit=5)", "", ""),
        ("ACT_DIST_POISSON_LEV", "=ACT_DIST_POISSON_LEV(5, 5)", "E[min(X,5)] for Poisson(5)"),
        ("", "", ""),
        ("Negative Binomial LEV (r=5, p=0.3, limit=10)", "", ""),
        ("ACT_DIST_NEGBIN_LEV", "=ACT_DIST_NEGBIN_LEV(10, 5, 0.3)", "E[min(X,10)] for NegBin"),
        ("", "", ""),
        # Severity distributions
        ("Exponential LEV (λ=0.5, limit=2)", "", ""),
        ("ACT_DIST_EXP_LEV", "=ACT_DIST_EXP_LEV(2, 0.5)", "E[min(X,2)] = 1.2642"),
        ("", "", ""),
        ("Lognormal LEV (μ=0, σ=1, limit=2)", "", ""),
        ("ACT_DIST_LOGNORM_LEV", "=ACT_DIST_LOGNORM_LEV(2, 0, 1)", "E[min(X,2)] = 1.1139"),
        ("", "", ""),
        ("Gamma LEV (α=2, β=1, limit=2)", "", ""),
        ("ACT_DIST_GAMMA_LEV", "=ACT_DIST_GAMMA_LEV(2, 2, 1)", "E[min(X,2)] = 1.4587"),
        ("", "", ""),
        ("Pareto LEV (α=2, xm=1, limit=2)", "", ""),
        ("ACT_DIST_PARETO_LEV", "=ACT_DIST_PARETO_LEV(2, 2, 1)", "E[min(X,2)] = 1.5"),
        ("", "", ""),
        ("Lomax LEV (α=2, λ=1, limit=2)", "", ""),
        ("ACT_DIST_LOMAX_LEV", "=ACT_DIST_LOMAX_LEV(2, 2, 1)", "E[min(X,2)] = 0.6667"),
        ("", "", ""),
        ("GPD LEV (ξ=0.5, σ=1, limit=2)", "", ""),
        ("ACT_DIST_GPD_LEV", "=ACT_DIST_GPD_LEV(2, 0.5, 1)", "E[min(X,2)] = 1.0"),
        ("", "", ""),
        ("Weibull LEV (k=2, λ=1, limit=1)", "", ""),
        ("ACT_DIST_WEIBULL_LEV", "=ACT_DIST_WEIBULL_LEV(1, 2, 1)", "E[min(X,1)] = 0.7468"),
        ("", "", ""),
        ("Beta LEV (α=2, β=5, limit=0.5)", "", ""),
        ("ACT_DIST_BETA_LEV", "=ACT_DIST_BETA_LEV(0.5, 2, 5)", "E[min(X,0.5)] = 0.2757"),
        ("", "", ""),
        ("Burr XII LEV (c=2, k=3, λ=1, limit=1)", "", ""),
        ("ACT_DIST_BURR_LEV", "=ACT_DIST_BURR_LEV(1, 2, 3, 1)", "E[min(X,1)] = 0.5445"),
        ("", "", ""),
    ]

    # Insert rows for LEV functions
    num_rows_needed = len(lev_functions)
    print(f"Inserting {num_rows_needed} rows at row {exposure_row}")
    ws.insert_rows(exposure_row, num_rows_needed)

    # Add the LEV functions
    bold_font = Font(bold=True)
    for i, (col_a, col_b, col_c) in enumerate(lev_functions):
        row = exposure_row + i
        ws.cell(row, 1, col_a)
        ws.cell(row, 2, col_b)
        ws.cell(row, 3, col_c)

        # Make section headers bold
        if col_a and not col_a.startswith("ACT_") and col_b == "":
            ws.cell(row, 1).font = bold_font

    # Save
    wb.save('excel/actuarial_add_in.xlsm')
    print(f"Saved workbook with {num_rows_needed} new LEV function rows")

if __name__ == "__main__":
    add_lev_functions()
