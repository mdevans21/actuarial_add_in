#!/usr/bin/env python3
"""
Fix array formula placement to allow proper spilling in Excel 365.
"""

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")


def fix_chain_ladder_sheet(wb):
    """Fix Chain Ladder sheet to allow array formulas to spill."""
    ws = wb['Chain Ladder']

    # The issue is that B21 has =ACT_CL_FACTORS(...) which returns 9 values
    # but needs cells B21:J21 to be empty for spilling

    # Check B21 - Development Factors (should spill horizontally)
    cell_b21 = ws['B21']
    if cell_b21.value and 'ACT_CL_FACTORS' in str(cell_b21.value):
        # Clear cells to the right to allow spilling
        for col in range(3, 12):  # C through K
            ws.cell(row=21, column=col).value = None

    # Check B41 - also ACT_CL_FACTORS
    cell_b41 = ws['B41']
    if cell_b41.value and 'ACT_CL_FACTORS' in str(cell_b41.value):
        for col in range(3, 12):
            ws.cell(row=41, column=col).value = None

    # Check A89 - ACT_CL_BOOTSTRAP (returns array with statistics)
    cell_a89 = ws['A89']
    if cell_a89.value and 'ACT_CL_BOOTSTRAP' in str(cell_a89.value):
        # This returns a 2-column array (labels and values)
        # Clear area below and to the right
        for r in range(89, 100):
            for c in range(1, 4):
                if r != 89 or c != 1:  # Keep the formula cell
                    ws.cell(row=r, column=c).value = None

    # Check A92 - ACT_CL_BOOTSTRAP_ORIGIN (returns larger array)
    cell_a92 = ws['A92']
    if cell_a92.value and 'ACT_CL_BOOTSTRAP_ORIGIN' in str(cell_a92.value):
        # This returns a multi-row, multi-column array
        # Clear area below and to the right
        for r in range(92, 110):
            for c in range(1, 10):
                if r != 92 or c != 1:
                    ws.cell(row=r, column=c).value = None

    print("Fixed Chain Ladder array formula areas")


def add_array_formula_examples(wb):
    """Add a sheet demonstrating array formulas properly."""
    if "Array Formula Demo" in wb.sheetnames:
        del wb["Array Formula Demo"]

    ws = wb.create_sheet("Array Formula Demo")

    row = 1
    ws.cell(row=row, column=1, value="ARRAY FORMULA DEMONSTRATION").font = Font(bold=True, size=14)
    row += 2

    ws.cell(row=row, column=1, value="These formulas return arrays that spill in Excel 365.").font = HEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value="If you see @ prefix, delete it - the formula should spill into empty cells.")
    row += 2

    # Small test triangle
    ws.cell(row=row, column=1, value="Test Triangle:").font = HEADER_FONT
    row += 1
    tri_start = row
    triangle = [
        [100, 150, 170, 180, 185],
        [110, 165, 190, 200, None],
        [120, 180, 210, None, None],
        [130, 195, None, None, None],
        [140, None, None, None, None],
    ]
    for i, tri_row in enumerate(triangle):
        for j, val in enumerate(tri_row):
            if val is not None:
                ws.cell(row=row + i, column=2 + j, value=val)
    tri_range = f"B{tri_start}:F{tri_start + 4}"
    row = tri_start + 6

    # Development Factors - spills horizontally
    ws.cell(row=row, column=1, value="ACT_CL_FACTORS (spills right →)").font = HEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value="Formula:")
    ws.cell(row=row, column=2, value=f"=ACT_CL_FACTORS({tri_range})")
    row += 1
    ws.cell(row=row, column=1, value="Result:")
    # Leave B column empty for formula, clear C-F for spilling
    row += 2

    # Latest diagonal - spills vertically
    ws.cell(row=row, column=1, value="ACT_CL_LATEST (spills down ↓)").font = HEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value="Formula:")
    ws.cell(row=row, column=2, value=f"=ACT_CL_LATEST({tri_range})")
    row += 1
    ws.cell(row=row, column=1, value="Result:")
    row += 6  # Leave room for 5 values

    # Ultimate - spills vertically
    ws.cell(row=row, column=1, value="ACT_CL_ULTIMATE (spills down ↓)").font = HEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value="Formula:")
    ws.cell(row=row, column=2, value=f"=ACT_CL_ULTIMATE({tri_range})")
    row += 1
    ws.cell(row=row, column=1, value="Result:")
    row += 6

    # Bootstrap - spills as table
    ws.cell(row=row, column=1, value="ACT_CL_BOOTSTRAP (spills as table ↓→)").font = HEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value="Formula:")
    ws.cell(row=row, column=2, value=f"=ACT_CL_BOOTSTRAP({tri_range}, 100, 42)")
    row += 1
    ws.cell(row=row, column=1, value="Result:")
    row += 10

    # Copula - spills as table
    ws.cell(row=row, column=1, value="ACT_COPULA_GAUSSIAN (spills as table)").font = HEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value="Formula:")
    ws.cell(row=row, column=2, value="=ACT_COPULA_GAUSSIAN({1,0.5;0.5,1}, 5, 42)")
    row += 1
    ws.cell(row=row, column=1, value="Result:")
    row += 7

    # Commit history
    ws.cell(row=row, column=1, value="ACT_COMMIT_HISTORY (spills as table)").font = HEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value="Formula:")
    ws.cell(row=row, column=2, value="=ACT_COMMIT_HISTORY()")
    row += 1
    ws.cell(row=row, column=1, value="Result:")
    row += 1

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 50

    print("Added Array Formula Demo sheet")


def main():
    project_root = Path(__file__).parent.parent
    target_file = project_root / "excel" / "actuarial_add_in_v0.1.xlsm"

    wb = load_workbook(target_file, keep_vba=True)

    fix_chain_ladder_sheet(wb)
    add_array_formula_examples(wb)

    wb.save(target_file)
    print(f"\nSaved: {target_file}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
